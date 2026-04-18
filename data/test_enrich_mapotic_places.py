#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sys
import tempfile
import threading
import time
import unittest
from pathlib import Path
from unittest import mock

import requests
from openpyxl import Workbook
from PIL import Image

THIS_DIR = Path(__file__).resolve().parent
if str(THIS_DIR) not in sys.path:
    sys.path.insert(0, str(THIS_DIR))

import enrich_mapotic_places as module


def make_image(
    *,
    original: str,
    title: str,
    source: str = "Example Source",
    link: str = "https://example.com/gallery",
    width: int = 1800,
    height: int = 1100,
    position: int = 1,
) -> dict[str, object]:
    return {
        "original": original,
        "title": title,
        "source": source,
        "link": link,
        "original_width": width,
        "original_height": height,
        "position": position,
        "is_product": False,
    }


def make_blank_image_bytes(width: int = 1200, height: int = 800) -> bytes:
    image = Image.new("RGB", (width, height), color="white")
    from io import BytesIO

    buffer = BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


class FakeSerpClient:
    def __init__(self, payloads: dict[str, object]) -> None:
        self.payloads = payloads
        self.requests: list[tuple[str, dict[str, object]]] = []
        self.cache_hits = 0
        self._request_counts: dict[str, int] = {}

    def request(self, namespace: str, params: dict[str, object]) -> dict[str, object]:
        self.requests.append((namespace, params))
        payload = self.payloads[namespace]
        if isinstance(payload, list):
            index = self._request_counts.get(namespace, 0)
            self._request_counts[namespace] = index + 1
            return payload[index]
        return payload


class FakeGooglePlacesClient:
    def __init__(self, payload: dict[str, object]) -> None:
        self.payload = payload
        self.lookups: list[tuple[str, str]] = []

    def lookup_place(self, place_id: str, country_iso2: str = "") -> dict[str, object]:
        self.lookups.append((place_id, country_iso2))
        return dict(self.payload)


class ImageLogicTests(unittest.TestCase):
    def setUp(self) -> None:
        module.place_serp_bundle_cache.clear()
        module.place_serp_bundle_inflight.clear()

    def test_build_unified_image_query_normalizes_query_inputs(self) -> None:
        query = module.build_unified_image_query(
            {
                "CITY": " Austin  ",
                "COUNTRY": " United   States ",
            },
            " Test   Venue ",
        )

        self.assertTrue(query.startswith('"Test Venue" austin united states '))
        self.assertIn(
            '("concert" OR "crowd" OR "stage" OR "exterior" OR "facade" OR "entrance")',
            query,
        )
        self.assertTrue(query.endswith(module.IMAGE_SEARCH_NEGATIVE_CLAUSE))

    def test_retry_delay_for_openai_429_uses_longer_default_backoff(self) -> None:
        response = requests.Response()
        response.status_code = 429
        error = requests.HTTPError("rate limited", response=response)

        self.assertEqual(module.retry_delay_for_exception(error, 1), 15.0)
        self.assertEqual(module.retry_delay_for_exception(error, 2), 30.0)

    def test_retry_delay_for_openai_429_honors_retry_after_header(self) -> None:
        response = requests.Response()
        response.status_code = 429
        response.headers["Retry-After"] = "42"
        error = requests.HTTPError("rate limited", response=response)

        self.assertEqual(module.retry_delay_for_exception(error, 1), 42.0)

    def test_google_places_cost_tracker_counts_network_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tracker = module.CostTracker()
            client = module.GooglePlacesClient(
                api_key="google-secret",
                cache_dir=Path(tmpdir),
                request_pause=0.0,
                cost_tracker=tracker,
            )
            response = mock.Mock()
            response.status_code = 200
            response.raise_for_status.return_value = None
            response.json.return_value = {
                "displayName": {"text": "Test Venue"},
                "formattedAddress": "123 Main St",
                "location": {"latitude": 30.0, "longitude": -97.0},
            }
            client.session.get = mock.Mock(return_value=response)

            client.lookup_place("abc123", country_iso2="US")
            client.lookup_place("abc123", country_iso2="US")

        snapshot = tracker.snapshot()
        self.assertEqual(snapshot["by_service"]["google_places"]["calls"], 1)
        self.assertEqual(snapshot["by_service"]["google_places"]["cost"], 0.017)
        self.assertEqual(snapshot["total"], 0.017)
        self.assertEqual(client.cache_hits, 1)

    def test_google_places_lookup_uses_place_id_and_field_mask(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            client = module.GooglePlacesClient(
                api_key="google-secret",
                cache_dir=Path(tmpdir),
                request_pause=0.0,
            )
            response = mock.Mock()
            response.status_code = 200
            response.raise_for_status.return_value = None
            response.json.return_value = {
                "id": "abc123",
                "displayName": {"text": "Test Venue"},
                "formattedAddress": "123 Main St, Austin, TX 78701, USA",
                "location": {"latitude": 30.2672, "longitude": -97.7431},
                "websiteUri": "https://testvenue.example.com",
                "regularOpeningHours": {
                    "openNow": True,
                    "weekdayDescriptions": ["Monday: 9:00 AM - 5:00 PM"],
                },
            }
            client.session.get = mock.Mock(return_value=response)

            payload = client.lookup_place("abc123", country_iso2="US")

        self.assertEqual(payload["title"], "Test Venue")
        self.assertEqual(payload["website"], "https://testvenue.example.com")
        client.session.get.assert_called_once()
        call = client.session.get.call_args
        self.assertEqual(call.args[0], f"{module.GOOGLE_PLACES_URL}/abc123")
        self.assertEqual(call.kwargs["headers"]["X-Goog-Api-Key"], "google-secret")
        self.assertEqual(call.kwargs["headers"]["X-Goog-FieldMask"], module.GOOGLE_PLACE_FIELD_MASK)
        self.assertEqual(call.kwargs["params"]["languageCode"], "en")
        self.assertEqual(call.kwargs["params"]["regionCode"], "US")

    @mock.patch("enrich_mapotic_places.time.sleep")
    def test_openai_client_429_retries_with_retry_after_delay(self, mocked_sleep: mock.Mock) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            client = module.OpenAIClient(
                api_key="openai-secret",
                cache_dir=Path(tmpdir),
                request_pause=0.0,
            )

            throttled_response = mock.Mock()
            throttled_response.status_code = 429
            throttled_response.headers = {"Retry-After": "21"}
            throttled_error = requests.HTTPError("rate limited", response=throttled_response)
            throttled_response.raise_for_status.side_effect = throttled_error

            success_response = mock.Mock()
            success_response.status_code = 200
            success_response.raise_for_status.return_value = None
            success_response.json.return_value = {
                "output": [
                    {
                        "type": "message",
                        "content": [
                            {
                                "type": "output_text",
                                "text": '{"descriptions":[{"place_id":"abc123","description":"A vivid venue description with musicroadtrip.com and enough specific detail to satisfy the validator without inventing facts."}]}',
                            }
                        ],
                    }
                ]
            }

            client.session.post = mock.Mock(side_effect=[throttled_response, success_response])

            result = client.describe_batch(
                [{"place_id": "abc123", "name": "Test Venue", "style_hint": "scene-setting"}]
            )

        self.assertIn("abc123", result)
        self.assertEqual(client.network_calls, 1)
        mocked_sleep.assert_any_call(21.0)

    def test_rejects_flyer_poster_candidate(self) -> None:
        candidate = module.build_image_candidate(
            make_image(
                original="https://cdn.example.com/flyer.jpg",
                title="Venue flyer poster with lineup text",
            ),
            query_type="live",
        )
        self.assertIn("disallowed_image_type_or_content", candidate["rejection_reasons"])

    def test_rejects_visible_text_overlay_candidate(self) -> None:
        candidate = module.build_image_candidate(
            make_image(
                original="https://cdn.example.com/text-overlay.jpg",
                title="Concert crowd with visible writing and text overlay",
            ),
            query_type="live",
        )
        self.assertIn("visible_text_or_writing", candidate["rejection_reasons"])

    def test_rejects_watermark_candidate(self) -> None:
        candidate = module.build_image_candidate(
            make_image(
                original="https://cdn.example.com/watermark.jpg",
                title="Wide venue shot with watermark",
            ),
            query_type="live",
        )
        self.assertIn("watermark_or_logo_overlay", candidate["rejection_reasons"])

    def test_rejects_food_and_drink_candidate(self) -> None:
        candidate = module.build_image_candidate(
            make_image(
                original="https://cdn.example.com/cocktails.jpg",
                title="Cocktails and bottle service at the venue",
            ),
            query_type="live",
        )
        self.assertIn("food_or_drink_focus", candidate["rejection_reasons"])

    def test_rejects_small_or_narrow_candidate(self) -> None:
        candidate = module.build_image_candidate(
            make_image(
                original="https://cdn.example.com/narrow.jpg",
                title="Wide live concert crowd inside venue stage room",
                width=1199,
                height=900,
            ),
            query_type="live",
        )
        self.assertIn("low_resolution", candidate["rejection_reasons"])
        self.assertIn("narrow_or_vertical_framing", candidate["rejection_reasons"])

    def test_prefers_wide_live_performance_venue_shot(self) -> None:
        concert_results = [
            make_image(
                original="https://cdn.example.com/live-wide.jpg",
                title="Wide live concert crowd inside venue stage room",
                position=2,
                width=2200,
                height=1200,
            ),
            make_image(
                original="https://cdn.example.com/artist-portrait.jpg",
                title="Artist portrait on stage",
                position=1,
            ),
        ]
        exterior_results = [
            make_image(
                original="https://cdn.example.com/exterior.jpg",
                title="Venue exterior facade entrance",
                position=1,
            )
        ]

        selected = module.select_image_fields(concert_results, exterior_results, place_results={})

        self.assertEqual(selected["Main Image URL"], "https://cdn.example.com/live-wide.jpg")
        self.assertTrue(
            selected["Additional Image URL(s)"].startswith("https://cdn.example.com/live-wide.jpg$")
        )
        self.assertEqual(selected["image_status"], "main_live_performance_with_ordered_gallery")

    def test_falls_back_to_exterior_when_live_shot_is_rejected(self) -> None:
        concert_results = [
            make_image(
                original="https://cdn.example.com/poster.jpg",
                title="Event poster flyer with lineup text",
            )
        ]
        exterior_results = [
            make_image(
                original="https://cdn.example.com/facade.jpg",
                title="Venue exterior facade entrance marquee",
            )
        ]

        selected = module.select_image_fields(concert_results, exterior_results, place_results={})

        self.assertEqual(selected["Main Image URL"], "https://cdn.example.com/facade.jpg")
        self.assertEqual(selected["image_status"], "main_exterior_fallback_with_ordered_gallery")

    def test_additional_gallery_dedupes_near_duplicates(self) -> None:
        concert_results = [
            make_image(
                original="https://cdn.example.com/venue-main.jpg?size=large",
                title="Wide live concert crowd inside venue stage room",
            ),
            make_image(
                original="https://cdn.example.com/venue-main.jpg?size=small",
                title="Wide live concert crowd inside venue stage room",
                position=2,
            ),
            make_image(
                original="https://cdn.example.com/stage-layout.jpg",
                title="Stage layout audience room at the venue",
                position=3,
            ),
        ]
        exterior_results = [
            make_image(
                original="https://cdn.example.com/venue-main.jpg",
                title="Venue exterior facade",
                position=1,
            ),
            make_image(
                original="https://cdn.example.com/front-door.jpg",
                title="Venue exterior facade entrance marquee",
                position=2,
            ),
        ]

        selected = module.select_image_fields(concert_results, exterior_results, place_results={})
        parts = selected["Additional Image URL(s)"].split("$")

        self.assertEqual(parts[0], selected["Main Image URL"])
        self.assertEqual(parts.count("https://cdn.example.com/venue-main.jpg?size=large"), 1)
        self.assertIn("https://cdn.example.com/front-door.jpg", parts)
        self.assertIn("https://cdn.example.com/stage-layout.jpg", parts)

    def test_additional_gallery_output_is_dollar_delimited(self) -> None:
        value = module.format_additional_image_value(
            [
                "https://cdn.example.com/1.jpg",
                "https://cdn.example.com/2.jpg",
                "https://cdn.example.com/3.jpg",
            ]
        )
        self.assertEqual(
            value,
            "https://cdn.example.com/1.jpg$https://cdn.example.com/2.jpg$https://cdn.example.com/3.jpg",
        )
        self.assertEqual(module.validate_additional_image_value(value), [])

    def test_delete_sentinel_rules(self) -> None:
        self.assertEqual(module.format_additional_image_value([], use_delete_sentinel=True), "$DELETE")
        self.assertEqual(module.validate_additional_image_value("$DELETE"), [])
        self.assertIn(
            "delete_sentinel_mixed_with_urls",
            module.validate_additional_image_value("$DELETE$https://cdn.example.com/1.jpg"),
        )

    def test_first_additional_image_matches_main(self) -> None:
        concert_results = [
            make_image(
                original="https://cdn.example.com/main-shot.jpg",
                title="Wide live concert crowd inside venue stage room",
            ),
            make_image(
                original="https://cdn.example.com/interior-room.jpg",
                title="Interior audience room at the venue",
                position=2,
            ),
        ]
        exterior_results = [
            make_image(
                original="https://cdn.example.com/exterior-marquee.jpg",
                title="Venue exterior facade entrance marquee",
            )
        ]

        selected = module.select_image_fields(concert_results, exterior_results, place_results={})
        first_additional = selected["Additional Image URL(s)"].split("$")[0]

        self.assertEqual(first_additional, selected["Main Image URL"])

    def test_non_direct_asset_url_is_rejected(self) -> None:
        candidate = module.build_image_candidate(
            make_image(
                original="https://example.com/gallery/poster.html",
                title="Wide live concert crowd inside venue stage room",
            ),
            query_type="live",
        )
        self.assertIn("not_direct_public_image_asset", candidate["rejection_reasons"])

    @mock.patch("enrich_mapotic_places.pytesseract.image_to_data")
    def test_binary_inspection_rejects_visible_text_overlay(self, mocked_ocr: mock.Mock) -> None:
        mocked_ocr.return_value = {
            "text": ["BIG", "SUMMER", "SHOW", "TONIGHT", "TICKETS"],
            "left": [50, 180, 360, 520, 760],
            "top": [40, 42, 38, 44, 40],
            "width": [100, 150, 120, 180, 140],
            "height": [40, 40, 40, 40, 40],
        }

        inspected = module.inspect_image_bytes(make_blank_image_bytes(), content_type="image/png")

        self.assertIn("ocr_visible_text", inspected["rejection_reasons"])

    @mock.patch("enrich_mapotic_places.pytesseract.image_to_data")
    def test_binary_inspection_rejects_watermark(self, mocked_ocr: mock.Mock) -> None:
        mocked_ocr.return_value = {
            "text": ["Getty", "Images"],
            "left": [10, 80],
            "top": [10, 12],
            "width": [60, 70],
            "height": [20, 20],
        }

        inspected = module.inspect_image_bytes(make_blank_image_bytes(), content_type="image/png")

        self.assertIn("ocr_watermark_text", inspected["rejection_reasons"])

    def test_social_search_fallback_enriches_missing_socials(self) -> None:
        serp = FakeSerpClient(
            {
                "google_social_bundle": {
                    "organic_results": [
                        {"link": "https://www.instagram.com/testvenue/"},
                        {"link": "https://www.facebook.com/testvenue"},
                        {"link": "https://x.com/testvenue"},
                        {"link": "https://www.tiktok.com/@testvenue"},
                        {"link": "https://www.youtube.com/@testvenue"},
                    ]
                }
            }
        )
        results = module.search_missing_socials(
            serp=serp,
            row={"CITY": "  Austin  ", "COUNTRY": " United   States ", "COUNTRY_ISO2": "US"},
            place_name="  Test   Venue ",
            current_fields={},
        )

        self.assertEqual(results["Instagram"], "https://www.instagram.com/testvenue/")
        self.assertEqual(results["Facebook"], "https://www.facebook.com/testvenue")
        self.assertEqual(results["X (Twitter)"], "https://x.com/testvenue")
        self.assertEqual(results["TikTok"], "https://www.tiktok.com/@testvenue")
        self.assertEqual(results["YouTube"], "https://www.youtube.com/@testvenue")
        self.assertEqual(len(serp.requests), 1)
        namespace, params = serp.requests[0]
        self.assertEqual(namespace, "google_social_bundle")
        self.assertEqual(
            params["q"],
            '"Test Venue" austin united states site:instagram.com OR site:facebook.com OR '
            "site:tiktok.com OR site:x.com OR site:twitter.com OR site:youtube.com",
        )

    def test_social_search_returns_full_bundle_from_single_response(self) -> None:
        serp = FakeSerpClient(
            {
                "google_social_bundle": {
                    "organic_results": [
                        {"link": "https://www.instagram.com/testvenue/"},
                        {"link": "https://www.facebook.com/testvenue"},
                        {"link": "https://x.com/testvenue"},
                    ]
                }
            }
        )

        results = module.search_missing_socials(
            serp=serp,
            row={"CITY": "Austin", "COUNTRY": "United States", "COUNTRY_ISO2": "US"},
            place_name="Test Venue",
            current_fields={"Instagram": "https://www.instagram.com/already-known/"},
        )

        self.assertEqual(results["Instagram"], "https://www.instagram.com/testvenue/")
        self.assertEqual(results["Facebook"], "https://www.facebook.com/testvenue")
        self.assertEqual(results["X (Twitter)"], "https://x.com/testvenue")
        self.assertEqual(len(serp.requests), 1)

    def test_search_images_uses_wide_high_quality_google_images_params(self) -> None:
        serp = FakeSerpClient({"google_images": {"images_results": []}})

        module.search_images(
            serp=serp,
            query="test query",
            location="Austin, Texas, United States",
            country_iso2="US",
        )

        self.assertEqual(len(serp.requests), 1)
        namespace, params = serp.requests[0]
        self.assertEqual(namespace, "google_images")
        self.assertEqual(params["engine"], "google_images")
        self.assertEqual(params["google_domain"], "google.com")
        self.assertEqual(params["imgar"], "w")
        self.assertEqual(params["imgsz"], "4mp")
        self.assertEqual(params["image_type"], "photo")
        self.assertEqual(params["safe"], "active")
        self.assertEqual(params["filter"], "1")
        self.assertTrue(params["no_cache"])
        self.assertEqual(params["ijn"], 0)

    def test_search_images_retries_without_location_after_400(self) -> None:
        class RetrySerp:
            def __init__(self) -> None:
                self.calls: list[dict[str, object]] = []

            def request(self, namespace: str, params: dict[str, object]) -> dict[str, object]:
                self.calls.append(dict(params))
                if len(self.calls) == 1:
                    raise RuntimeError(
                        "SerpApi request failed for google_images: 400 Client Error: Bad Request"
                    )
                return {"images_results": [{"original": "https://cdn.example.com/fallback.jpg"}]}

        serp = RetrySerp()

        results = module.search_images(
            serp=serp,
            query="test query",
            location="AAL, Argentina",
            country_iso2="AR",
        )

        self.assertEqual(results[0]["original"], "https://cdn.example.com/fallback.jpg")
        self.assertEqual(len(serp.calls), 2)
        self.assertEqual(serp.calls[0]["location"], "AAL, Argentina")
        self.assertIsNone(serp.calls[1]["location"])

    def test_search_place_photos_uses_google_maps_data_id(self) -> None:
        serp = FakeSerpClient(
            {
                "google_maps_photos": {
                    "photos": [
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=1",
                        }
                    ]
                }
            }
        )

        results = module.search_place_photos(serp, {"data_id": "0xabc"})

        self.assertEqual(results[0]["image"], "https://lh5.googleusercontent.com/p/test-photo=w4032-h3024-k-no")
        self.assertEqual(len(serp.requests), 1)
        namespace, params = serp.requests[0]
        self.assertEqual(namespace, "google_maps_photos")
        self.assertEqual(params["engine"], "google_maps_photos")
        self.assertEqual(params["data_id"], "0xabc")
        self.assertEqual(params["hl"], "en")

    def test_create_place_enrichment_uses_place_photos_before_google_images(self) -> None:
        google_places = FakeGooglePlacesClient(
            {
                "title": "Test Venue from Google",
                "address": "123 Main St",
                "gps_coordinates": {"latitude": "30.2672", "longitude": "-97.7431"},
                "phone": "+1 512-555-0100",
                "hours": [{"monday": "9:00 AM - 5:00 PM"}],
                "website": "https://testvenue.example.com",
            }
        )
        serp = FakeSerpClient(
            {
                "google_maps_place": {
                    "place_results": {
                        "data_id": "0xabc",
                    }
                },
                "google_maps_photos": {
                    "photos": [
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-1=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=1",
                        },
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-2=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=2",
                        },
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-3=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=3",
                        },
                    ]
                },
                "google_social_bundle": {"organic_results": []},
            }
        )

        payload = module.create_place_enrichment(
            google_places=google_places,
            serp=serp,
            image_inspector=None,
            row={
                "GOOGLE_PLACE_ID": "abc123",
                "NAME": "Test Venue",
                "CITY": "Austin",
                "COUNTRY": "United States",
                "COUNTRY_ISO2": "US",
            },
            skip_images=False,
            gallery_mode="mirror-main",
        )

        self.assertEqual(
            payload["fields"]["Main Image URL"],
            "https://lh5.googleusercontent.com/p/test-photo-1=w4032-h3024-k-no",
        )
        self.assertEqual(payload["fields"]["Name"], "Test Venue from Google")
        self.assertEqual(payload["fields"]["Website"], "https://testvenue.example.com")
        self.assertIn(
            "https://lh5.googleusercontent.com/p/test-photo-2=w4032-h3024-k-no",
            payload["fields"]["Additional Image URL(s)"],
        )
        self.assertEqual(google_places.lookups, [("abc123", "US")])
        self.assertFalse(any(namespace == "google_images" for namespace, _params in serp.requests))
        self.assertIn("google_social_bundle", [namespace for namespace, _params in serp.requests])

    def test_create_place_enrichment_runs_google_images_when_place_photos_only_fill_one_gallery_slot(self) -> None:
        google_places = FakeGooglePlacesClient(
            {
                "title": "Test Venue from Google",
                "address": "123 Main St",
            }
        )
        serp = FakeSerpClient(
            {
                "google_maps_place": {
                    "place_results": {
                        "data_id": "0xabc",
                    }
                },
                "google_maps_photos": {
                    "photos": [
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-1=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=1",
                        },
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-2=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=2",
                        },
                    ]
                },
                "google_social_bundle": {"organic_results": []},
                "google_images": {
                    "images_results": [
                        make_image(
                            original="https://cdn.example.com/live-wide.jpg",
                            title="Wide live concert crowd inside venue stage room",
                        ),
                        make_image(
                            original="https://cdn.example.com/exterior.jpg",
                            title="Venue exterior facade entrance marquee",
                            position=2,
                        ),
                    ]
                },
            }
        )

        payload = module.create_place_enrichment(
            google_places=google_places,
            serp=serp,
            image_inspector=None,
            row={
                "GOOGLE_PLACE_ID": "abc123",
                "NAME": "Test Venue",
                "CITY": "Austin",
                "COUNTRY": "United States",
                "COUNTRY_ISO2": "US",
            },
            skip_images=False,
            gallery_mode="mirror-main",
        )

        self.assertEqual(payload["fields"]["Main Image URL"], "https://cdn.example.com/live-wide.jpg")
        image_requests = [params for namespace, params in serp.requests if namespace == "google_images"]
        self.assertEqual(len(image_requests), 1)

    def test_create_place_enrichment_falls_back_to_google_images_when_place_photos_are_insufficient(self) -> None:
        google_places = FakeGooglePlacesClient(
            {
                "title": "Test Venue from Google",
                "address": "123 Main St",
            }
        )
        serp = FakeSerpClient(
            {
                "google_maps_place": {
                    "place_results": {
                        "data_id": "0xabc",
                    }
                },
                "google_maps_photos": {
                    "photos": [
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-1=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=1",
                        }
                    ]
                },
                "google_social_bundle": {"organic_results": []},
                "google_images": {
                    "images_results": [
                        make_image(
                            original="https://cdn.example.com/live-wide.jpg",
                            title="Wide live concert crowd inside venue stage room",
                        ),
                        make_image(
                            original="https://cdn.example.com/exterior.jpg",
                            title="Venue exterior facade entrance marquee",
                            position=2,
                        ),
                    ]
                },
            }
        )

        payload = module.create_place_enrichment(
            google_places=google_places,
            serp=serp,
            image_inspector=None,
            row={
                "GOOGLE_PLACE_ID": "abc123",
                "NAME": "Test Venue",
                "CITY": "Austin",
                "COUNTRY": "United States",
                "COUNTRY_ISO2": "US",
            },
            skip_images=False,
            gallery_mode="mirror-main",
        )

        self.assertEqual(payload["fields"]["Main Image URL"], "https://cdn.example.com/live-wide.jpg")
        self.assertEqual(google_places.lookups, [("abc123", "US")])
        image_requests = [params for namespace, params in serp.requests if namespace == "google_images"]
        self.assertEqual(len(image_requests), 1)
        self.assertIn('"Test Venue from Google" austin united states', image_requests[0]["q"])
        self.assertEqual(image_requests[0]["location"], "austin, united states")

    def test_create_place_enrichment_skip_images_keeps_socials_without_image_serp_calls(self) -> None:
        google_places = FakeGooglePlacesClient(
            {
                "title": "Test Venue from Google",
                "address": "123 Main St",
                "gps_coordinates": {"latitude": "30.2672", "longitude": "-97.7431"},
                "phone": "+1 512-555-0100",
                "hours": [{"monday": "9:00 AM - 5:00 PM"}],
                "website": "https://testvenue.example.com",
            }
        )
        serp = FakeSerpClient(
            {
                "google_social_bundle": {
                    "organic_results": [
                        {"link": "https://www.instagram.com/testvenue/"},
                        {"link": "https://www.facebook.com/testvenue"},
                        {"link": "https://x.com/testvenue"},
                        {"link": "https://www.tiktok.com/@testvenue"},
                        {"link": "https://www.youtube.com/@testvenue"},
                    ]
                }
            }
        )

        payload = module.create_place_enrichment(
            google_places=google_places,
            serp=serp,
            image_inspector=None,
            row={
                "GOOGLE_PLACE_ID": "abc123",
                "NAME": "Test Venue",
                "CITY": "Austin",
                "COUNTRY": "United States",
                "COUNTRY_ISO2": "US",
            },
            skip_images=True,
            gallery_mode="mirror-main",
        )

        self.assertEqual(payload["fields"]["Name"], "Test Venue from Google")
        self.assertEqual(payload["fields"]["Website"], "https://testvenue.example.com")
        self.assertEqual(payload["fields"]["Instagram"], "https://www.instagram.com/testvenue/")
        self.assertEqual(payload["fields"]["Facebook"], "https://www.facebook.com/testvenue")
        self.assertEqual(payload["fields"]["X (Twitter)"], "https://x.com/testvenue")
        self.assertEqual(payload["fields"]["TikTok"], "https://www.tiktok.com/@testvenue")
        self.assertEqual(payload["fields"]["YouTube"], "https://www.youtube.com/@testvenue")
        self.assertEqual(payload["image_status"], "not_requested")
        self.assertEqual(google_places.lookups, [("abc123", "US")])
        self.assertEqual([namespace for namespace, _params in serp.requests], ["google_social_bundle"])

    def test_create_place_enrichment_reuses_serp_bundle_per_place_id(self) -> None:
        google_places = FakeGooglePlacesClient(
            {
                "title": "Test Venue from Google",
                "address": "123 Main St",
                "website": "https://testvenue.example.com",
            }
        )
        serp = FakeSerpClient(
            {
                "google_maps_place": {"place_results": {"data_id": "0xabc"}},
                "google_maps_photos": {
                    "photos": [
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-1=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=1",
                        },
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-2=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=2",
                        },
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-3=w4032-h3024-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=3",
                        },
                    ]
                },
                "google_social_bundle": {"organic_results": []},
            }
        )

        first = module.create_place_enrichment(
            google_places=google_places,
            serp=serp,
            image_inspector=None,
            row={
                "GOOGLE_PLACE_ID": "abc123",
                "NAME": "Test Venue",
                "CITY": " Austin ",
                "COUNTRY": "United States",
                "COUNTRY_ISO2": "US",
            },
            skip_images=False,
            gallery_mode="mirror-main",
        )
        first_request_count = len(serp.requests)

        second = module.create_place_enrichment(
            google_places=google_places,
            serp=serp,
            image_inspector=None,
            row={
                "GOOGLE_PLACE_ID": "abc123",
                "NAME": "Test Venue",
                "CITY": "austin",
                "COUNTRY": " united   states ",
                "COUNTRY_ISO2": "US",
            },
            skip_images=False,
            gallery_mode="mirror-main",
        )

        self.assertEqual(first["fields"]["Main Image URL"], second["fields"]["Main Image URL"])
        self.assertEqual(len(serp.requests), first_request_count)
        self.assertGreaterEqual(serp.cache_hits, 1)

    def test_get_place_serp_bundle_runs_one_unified_image_search(self) -> None:
        serp = FakeSerpClient(
            {
                "google_social_bundle": {"organic_results": []},
                "google_maps_place": {"place_results": {"data_id": "0xabc"}},
                "google_maps_photos": {
                    "photos": [
                        {
                            "image": "https://lh5.googleusercontent.com/p/test-photo-1=w1600-h900-k-no",
                            "photo_meta_serpapi_link": "https://serpapi.com/search.json?engine=google_maps_photo_meta&data_id=1",
                        }
                    ]
                },
                "google_images": {
                    "images_results": [
                        make_image(
                            original="https://cdn.example.com/live-wide.jpg",
                            title="Wide live concert crowd inside venue stage room",
                        )
                    ]
                },
            }
        )

        bundle = module.get_place_serp_bundle(
            place_id="abc123",
            row={
                "GOOGLE_PLACE_ID": "abc123",
                "NAME": " Test Venue ",
                "CITY": " Austin ",
                "COUNTRY": " United States ",
                "COUNTRY_ISO2": "US",
            },
            serp_client=serp,
            place_results={"title": "Test Venue from Google"},
            current_fields={},
            skip_images=False,
        )

        image_requests = [params for namespace, params in serp.requests if namespace == "google_images"]
        self.assertEqual(len(image_requests), 1)
        self.assertIn(
            '("concert" OR "crowd" OR "stage" OR "exterior" OR "facade" OR "entrance")',
            image_requests[0]["q"],
        )
        self.assertEqual(bundle["image_results"][0]["original"], "https://cdn.example.com/live-wide.jpg")

    def test_get_place_serp_bundle_coalesces_concurrent_requests(self) -> None:
        class BlockingSerp:
            def __init__(self) -> None:
                self.requests: list[tuple[str, dict[str, object]]] = []
                self.started = threading.Event()
                self.release = threading.Event()

            def request(self, namespace: str, params: dict[str, object]) -> dict[str, object]:
                self.requests.append((namespace, dict(params)))
                self.started.set()
                self.release.wait(timeout=2.0)
                return {"organic_results": []}

        serp = BlockingSerp()
        row = {
            "GOOGLE_PLACE_ID": "abc123",
            "NAME": "Test Venue",
            "CITY": "Austin",
            "COUNTRY": "United States",
            "COUNTRY_ISO2": "US",
        }
        results: list[dict[str, object]] = []
        errors: list[Exception] = []

        def worker() -> None:
            try:
                results.append(
                    module.get_place_serp_bundle(
                        place_id="abc123",
                        row=row,
                        serp_client=serp,
                        place_results={"title": "Test Venue"},
                        current_fields={},
                        skip_images=True,
                    )
                )
            except Exception as exc:  # noqa: BLE001
                errors.append(exc)

        first = threading.Thread(target=worker)
        second = threading.Thread(target=worker)
        first.start()
        self.assertTrue(serp.started.wait(timeout=1.0))
        second.start()
        time.sleep(0.05)
        self.assertEqual(len(serp.requests), 1)
        serp.release.set()
        first.join(timeout=1.0)
        second.join(timeout=1.0)

        self.assertEqual(errors, [])
        self.assertEqual(len(results), 2)
        self.assertEqual(len(serp.requests), 1)

    def test_gallery_mode_separate_main_omits_main_from_additional_images(self) -> None:
        concert_results = [
            make_image(
                original="https://cdn.example.com/main-shot.jpg",
                title="Wide live concert crowd inside venue stage room",
            ),
            make_image(
                original="https://cdn.example.com/interior-room.jpg",
                title="Interior audience room at the venue",
                position=2,
            ),
        ]
        exterior_results = [
            make_image(
                original="https://cdn.example.com/exterior-marquee.jpg",
                title="Venue exterior facade entrance marquee",
            )
        ]

        selected = module.select_image_fields(
            concert_results,
            exterior_results,
            place_results={},
            gallery_mode="separate-main",
        )

        self.assertFalse(selected["Additional Image URL(s)"].startswith(selected["Main Image URL"]))

    def test_build_final_row_values_overwrites_existing_image_fields(self) -> None:
        row = {
            "NAME": "Test Venue",
            "ADDRESS": "123 Main St",
            "GOOGLE_PLACE_ID": "abc123",
            "Main Image URL": "https://old.example.com/old.jpg",
            "Additional Image URL(s)": "https://old.example.com/old.jpg$https://old.example.com/old2.jpg",
        }
        place_payload = {
            "fields": {
                "Main Image URL": "https://cdn.example.com/new-main.jpg",
                "Additional Image URL(s)": "https://cdn.example.com/new-main.jpg$https://cdn.example.com/new-2.jpg",
            }
        }

        final_values = module.build_final_row_values(row, place_payload, descriptions={})

        self.assertEqual(final_values["Main Image URL"], "https://cdn.example.com/new-main.jpg")
        self.assertEqual(
            final_values["Additional Image URL(s)"],
            "https://cdn.example.com/new-main.jpg$https://cdn.example.com/new-2.jpg",
        )

    @mock.patch("enrich_mapotic_places.inspect_remote_image_url", side_effect=TimeoutError("slow"))
    def test_image_inspection_download_failure_rejects_candidate_without_crashing(
        self,
        mocked_inspection: mock.Mock,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            client = module.ImageInspectionClient(cache_dir=Path(tmpdir))
            payload = client.inspect_url("https://cdn.example.com/problem.jpg")

        self.assertIn("image_download_failed:TimeoutError", payload["rejection_reasons"])
        mocked_inspection.assert_called_once()

    @mock.patch("enrich_mapotic_places.subprocess.check_output")
    def test_get_secret_prefers_environment_variable(self, mocked_check_output: mock.Mock) -> None:
        with mock.patch.dict("os.environ", {"SERPAPI_API_KEY": "env-secret"}, clear=False):
            value = module.get_secret("SERPAPI_API_KEY")

        self.assertEqual(value, "env-secret")
        mocked_check_output.assert_not_called()

    @mock.patch("enrich_mapotic_places.create_place_enrichment")
    @mock.patch("enrich_mapotic_places.get_secret", return_value="secret")
    def test_enrich_saves_checkpoint_workbook_every_50_rows(
        self,
        mocked_get_secret: mock.Mock,
        mocked_create_place_enrichment: mock.Mock,
    ) -> None:
        def fake_place_enrichment(**kwargs: object) -> dict[str, object]:
            row = kwargs["row"]
            place_id = module.valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
            return {
                "place_id": place_id,
                "place_results": {},
                "fields": {"Website": f"https://example.com/{place_id}"},
                "description_context": {},
                "image_status": "no_images_found",
            }

        mocked_create_place_enrichment.side_effect = fake_place_enrichment

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_json = tmp_path / "summary.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["GOOGLE_PLACE_ID", "NAME", "ADDRESS"])
            for index in range(51):
                worksheet.append([f"place-{index}", f"Venue {index}", f"{index} Main Street"])
            workbook.save(input_xlsx)

            args = argparse.Namespace(
                input_xlsx=input_xlsx,
                output_xlsx=output_xlsx,
                summary_json=summary_json,
                cache_dir=tmp_path / "cache",
                limit=None,
                row_indices="",
                request_pause=0.0,
                heartbeat_seconds=999999,
                description_batch_size=12,
                openai_model="gpt-4o-mini",
                workers=1,
                show_cost=False,
                gallery_mode="mirror-main",
                skip_images=True,
                skip_openai=True,
            )

            save_call_paths: list[Path] = []
            original_save = Workbook.save

            def counting_save(self: Workbook, filename: str | Path) -> None:
                save_call_paths.append(Path(filename))
                original_save(self, filename)

            with (
                mock.patch("openpyxl.workbook.workbook.Workbook.save", new=counting_save),
                mock.patch.object(module, "DEFAULT_COST_LOG_DIR", tmp_path / "logs"),
            ):
                summary = module.enrich(args)

            self.assertTrue(output_xlsx.exists())
            self.assertTrue(summary_json.exists())

        self.assertEqual(summary["rows_selected"], 51)
        self.assertEqual(save_call_paths.count(output_xlsx), 3)
        self.assertEqual(mocked_create_place_enrichment.call_count, 51)
        self.assertEqual(mocked_get_secret.call_count, 3)

    @mock.patch("enrich_mapotic_places.create_place_enrichment")
    @mock.patch("enrich_mapotic_places.get_secret", return_value="secret")
    def test_enrich_worker_mode_preserves_row_order_and_survives_row_errors(
        self,
        mocked_get_secret: mock.Mock,
        mocked_create_place_enrichment: mock.Mock,
    ) -> None:
        def fake_place_enrichment(**kwargs: object) -> dict[str, object]:
            row = kwargs["row"]
            place_id = module.valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
            if place_id == "place-0":
                time.sleep(0.05)
            if place_id == "place-1":
                raise RuntimeError("boom")
            if place_id == "place-2":
                time.sleep(0.01)
            return {
                "place_id": place_id,
                "place_results": {},
                "fields": {"Website": f"https://example.com/{place_id}"},
                "description_context": {},
                "image_status": "no_images_found",
            }

        mocked_create_place_enrichment.side_effect = fake_place_enrichment

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_json = tmp_path / "summary.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["GOOGLE_PLACE_ID", "NAME", "ADDRESS"])
            for index in range(3):
                worksheet.append([f"place-{index}", f"Venue {index}", f"{index} Main Street"])
            workbook.save(input_xlsx)

            args = argparse.Namespace(
                input_xlsx=input_xlsx,
                output_xlsx=output_xlsx,
                summary_json=summary_json,
                cache_dir=tmp_path / "cache",
                limit=None,
                row_indices="",
                request_pause=0.0,
                heartbeat_seconds=999999,
                description_batch_size=12,
                openai_model="gpt-4o-mini",
                workers=3,
                show_cost=False,
                gallery_mode="mirror-main",
                skip_images=True,
                skip_openai=True,
            )

            with mock.patch.object(module, "DEFAULT_COST_LOG_DIR", tmp_path / "logs"):
                summary = module.enrich(args)
                cost_log_exists = Path(summary["cost_log_json"]).exists()
                result_workbook = module.load_workbook(output_xlsx)
                result_sheet = result_workbook.active
                headers = [
                    result_sheet.cell(1, column).value for column in range(1, result_sheet.max_column + 1)
                ]
                website_column = headers.index("Website") + 1

        self.assertEqual(summary["row_error_count"], 1)
        self.assertEqual(summary["skipped_rows"], 0)
        self.assertEqual(
            result_sheet.cell(2, website_column).value,
            "https://example.com/place-0",
        )
        self.assertEqual(result_sheet.cell(3, website_column).value, None)
        self.assertEqual(
            result_sheet.cell(4, website_column).value,
            "https://example.com/place-2",
        )
        self.assertTrue(cost_log_exists)
        self.assertEqual(mocked_create_place_enrichment.call_count, 3)
        self.assertEqual(mocked_get_secret.call_count, 3)

    @mock.patch("enrich_mapotic_places.create_place_enrichment")
    @mock.patch("enrich_mapotic_places.get_secret", return_value="secret")
    def test_enrich_resume_skips_already_checkpointed_rows(
        self,
        mocked_get_secret: mock.Mock,
        mocked_create_place_enrichment: mock.Mock,
    ) -> None:
        def fake_place_enrichment(**kwargs: object) -> dict[str, object]:
            row = kwargs["row"]
            place_id = module.valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
            return {
                "place_id": place_id,
                "place_results": {},
                "fields": {"Website": f"https://example.com/{place_id}"},
                "description_context": {},
                "image_status": "no_images_found",
            }

        mocked_create_place_enrichment.side_effect = fake_place_enrichment

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_one = tmp_path / "summary-one.json"
            summary_two = tmp_path / "summary-two.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["GOOGLE_PLACE_ID", "NAME", "ADDRESS"])
            for index in range(3):
                worksheet.append([f"place-{index}", f"Venue {index}", f"{index} Main Street"])
            workbook.save(input_xlsx)

            base_args = dict(
                input_xlsx=input_xlsx,
                output_xlsx=output_xlsx,
                cache_dir=tmp_path / "cache",
                row_indices="",
                request_pause=0.0,
                heartbeat_seconds=999999,
                description_batch_size=12,
                openai_model="gpt-4o-mini",
                workers=2,
                show_cost=False,
                gallery_mode="mirror-main",
                skip_images=True,
                skip_openai=True,
            )

            with mock.patch.object(module, "DEFAULT_COST_LOG_DIR", tmp_path / "logs"):
                first_summary = module.enrich(
                    argparse.Namespace(
                        summary_json=summary_one,
                        limit=2,
                        **base_args,
                    )
                )
                second_summary = module.enrich(
                    argparse.Namespace(
                        summary_json=summary_two,
                        limit=None,
                        **base_args,
                    )
                )
                result_workbook = module.load_workbook(output_xlsx)
                result_sheet = result_workbook.active
                headers = [
                    result_sheet.cell(1, column).value for column in range(1, result_sheet.max_column + 1)
                ]
                website_column = headers.index("Website") + 1

        self.assertEqual(first_summary["processed_rows_this_run"], 2)
        self.assertEqual(second_summary["skipped_rows"], 2)
        self.assertEqual(second_summary["processed_rows_this_run"], 1)
        self.assertEqual(second_summary["resumed_rows"], 1)
        self.assertEqual(mocked_create_place_enrichment.call_count, 3)
        self.assertEqual(
            result_sheet.cell(2, website_column).value,
            "https://example.com/place-0",
        )
        self.assertEqual(
            result_sheet.cell(3, website_column).value,
            "https://example.com/place-1",
        )
        self.assertEqual(
            result_sheet.cell(4, website_column).value,
            "https://example.com/place-2",
        )

    def test_verify_checkpoint_rows_reprocesses_missing_output_values(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(module.MAPOTIC_COLUMNS)
        worksheet.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        column_positions = {field: index + 1 for index, field in enumerate(module.MAPOTIC_COLUMNS)}

        verified = module.verify_checkpoint_rows(
            rows=[{"_rownum": 2}],
            processed_rownums={2},
            worksheet=worksheet,
            column_positions=column_positions,
        )

        self.assertEqual(verified, set())


if __name__ == "__main__":
    unittest.main()
