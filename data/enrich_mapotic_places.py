#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import concurrent.futures
import datetime as dt
import email.utils
import hashlib
import json
import os
import re
import subprocess
import sys
import threading
import time
from contextlib import nullcontext
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import pytesseract
import requests
from openpyxl import load_workbook
from PIL import Image
from rich.console import Console
from rich.live import Live
from rich.table import Table


SCRIPT_VERSION = "2026-04-16.1"
USER_AGENT = "MapoticPlaceEnricher/1.0 (+local-export)"
VAULT = "Employee"
GOOGLE_PLACES_URL = "https://places.googleapis.com/v1/places"
SERPAPI_URL = "https://serpapi.com/search.json"
OPENAI_RESPONSES_URL = "https://api.openai.com/v1/responses"

DEFAULT_INPUT_XLSX = Path("data/output/Jambase_Venues_Global_run_3_google_places_reviewed.xlsx")
DEFAULT_OUTPUT_XLSX = Path("data/output/Jambase_Venues_Global_run_3_mapotic_serp_openai.xlsx")
DEFAULT_SUMMARY_JSON = Path(
    "data/output/Jambase_Venues_Global_run_3_mapotic_serp_openai_summary.json"
)
DEFAULT_CACHE_DIR = Path("data/cache/mapotic_serp_openai")

DEFAULT_TIMEOUT = 60
DEFAULT_REQUEST_PAUSE = 0.15
DEFAULT_HEARTBEAT_SECONDS = 30
DEFAULT_DESCRIPTION_BATCH_SIZE = 12
DEFAULT_OPENAI_MODEL = "gpt-4o-mini"
DEFAULT_WORKERS = 4
CHECKPOINT_SAVE_EVERY_ROWS = 50
MAX_RETRIES = 5
DEFAULT_COST_LOG_DIR = Path("data/logs")
OPENAI_429_INITIAL_BACKOFF = 15.0
OPENAI_429_MAX_BACKOFF = 120.0
SERVICE_COSTS = {
    "google_places": 0.017,
    "serpapi": 0.01,
    "openai": 0.002,
}
SERVICE_LABELS = {
    "google_places": "Google Places",
    "serpapi": "SerpAPI",
    "openai": "OpenAI",
}
GOOGLE_PLACE_FIELD_MASK = ",".join(
    [
        "id",
        "displayName",
        "formattedAddress",
        "location",
        "internationalPhoneNumber",
        "nationalPhoneNumber",
        "websiteUri",
        "primaryType",
        "primaryTypeDisplayName",
        "businessStatus",
        "regularOpeningHours.weekdayDescriptions",
        "regularOpeningHours.openNow",
        "editorialSummary",
        "types",
    ]
)

MAPOTIC_COLUMNS = [
    "Name",
    "Address",
    "Latitude",
    "Longitude",
    "Description",
    "Hours of Operation",
    "Instagram",
    "Facebook",
    "X (Twitter)",
    "Email",
    "YouTube",
    "Phone",
    "Website",
    "TikTok",
    "PlacesID",
    "City",
    "State",
    "Zip Code",
    "Tickets Link",
    "Spotify URL",
    "Main Image URL",
    "Additional Image URL(s)",
]

BASE_COLUMN_SOURCES = {
    "Name": ["Name", "NAME"],
    "Address": ["Address", "ADDRESS"],
    "Latitude": ["Latitude", "LATITUDE"],
    "Longitude": ["Longitude", "LONGITUDE"],
    "Instagram": ["Instagram", "EXISTING_INSTAGRAM_CANDIDATE"],
    "Facebook": ["Facebook", "EXISTING_FACEBOOK_CANDIDATE"],
    "X (Twitter)": ["X (Twitter)", "EXISTING_X_CANDIDATE"],
    "Website": ["Website", "EXISTING_WEBSITE_CANDIDATE"],
    "TikTok": ["TikTok"],
    "City": ["City", "CITY"],
    "State": ["State", "STATE"],
    "Zip Code": ["Zip Code", "ZIP_CODE"],
    "Main Image URL": ["Main Image URL"],
    "Additional Image URL(s)": ["Additional Image URL(s)"],
    "Phone": ["Phone"],
    "Email": ["Email"],
    "YouTube": ["YouTube"],
    "Tickets Link": ["Tickets Link"],
    "Spotify URL": ["Spotify URL"],
    "Hours of Operation": ["Hours of Operation"],
    "Description": ["Description"],
    "PlacesID": ["PlacesID"],
}

DAY_ORDER = [
    "monday",
    "tuesday",
    "wednesday",
    "thursday",
    "friday",
    "saturday",
    "sunday",
]
DAY_LABELS = {
    "monday": "Mon",
    "tuesday": "Tue",
    "wednesday": "Wed",
    "thursday": "Thu",
    "friday": "Fri",
    "saturday": "Sat",
    "sunday": "Sun",
}
BANNED_DESCRIPTION_PHRASES = [
    "nestled in the heart of",
    "located in",
    "known for",
]
CONCERT_POSITIVE_TERMS = (
    "concert",
    "crowd",
    "stage",
    "live",
    "performance",
    "show",
    "gig",
    "festival",
    "audience",
    "band",
)
CONCERT_NEGATIVE_TERMS = (
    "poster",
    "flyer",
    "event card",
    "promo graphic",
    "promotional banner",
    "screenshot",
    "collage",
    "watermark",
    "logo",
    "text overlay",
    "ticket",
    "tickets",
    "map",
    "menu",
    "cover",
    "album",
    "setlist",
    "parking",
    "seating",
    "review",
)
EXTERIOR_POSITIVE_TERMS = (
    "exterior",
    "facade",
    "front",
    "entrance",
    "outside",
    "building",
    "venue",
)
EXTERIOR_NEGATIVE_TERMS = (
    "poster",
    "ticket",
    "logo",
    "watermark",
    "text overlay",
    "stage",
    "crowd",
    "performance",
    "concert",
)
SOCIAL_FIELD_ORDER = [
    "Instagram",
    "Facebook",
    "X (Twitter)",
    "TikTok",
    "YouTube",
    "Spotify URL",
    "Tickets Link",
    "Website",
]
TICKET_DOMAINS = (
    "ticketmaster.",
    "axs.",
    "livenation.",
    "eventbrite.",
    "ticketweb.",
    "dice.fm",
    "seetickets.",
    "bandsintown.",
    "residentadvisor.",
    "songkick.",
    "etix.",
    "ticketleap.",
    "stubhub.",
    "ticketek.",
)
SOCIAL_SEARCH_CONFIG = {
    "Instagram": {
        "query_prefix": "site:instagram.com",
        "disallowed_fragments": ("/p/", "/reel/", "/stories/", "/tv/"),
    },
    "Facebook": {
        "query_prefix": "site:facebook.com",
        "disallowed_fragments": ("/posts/", "/events/", "/photos/", "/share"),
    },
    "X (Twitter)": {
        "query_prefix": "site:x.com OR site:twitter.com",
        "disallowed_fragments": ("/status/", "/i/"),
    },
    "TikTok": {
        "query_prefix": "site:tiktok.com",
        "disallowed_fragments": ("/video/",),
    },
    "YouTube": {
        "query_prefix": "site:youtube.com OR site:youtu.be",
        "disallowed_fragments": ("/watch", "/shorts/"),
    },
}
IGNORED_URL_DOMAINS = (
    "google.com",
    "maps.google.com",
    "serpapi.com",
    "gstatic.com",
    "googleusercontent.com",
)
REPLACEABLE_WEBSITE_DOMAINS = (
    "jambase.com",
    "bandsintown.com",
    "songkick.com",
    "ticketmaster.com",
    "axs.com",
    "livenation.com",
    "eventbrite.com",
    "residentadvisor.net",
    "dice.fm",
    "stubhub.com",
)
IMAGE_FILE_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
    ".gif",
    ".bmp",
    ".avif",
    ".tif",
    ".tiff",
    ".heic",
    ".heif",
}
IMAGE_PAGE_EXTENSIONS = {
    ".html",
    ".htm",
    ".php",
    ".asp",
    ".aspx",
    ".jsp",
    ".pdf",
}
IMAGE_PAGE_HOSTS = (
    "facebook.com",
    "instagram.com",
    "pinterest.com",
    "tiktok.com",
    "x.com",
    "twitter.com",
    "youtube.com",
    "drive.google.com",
    "dropbox.com",
)
IMAGE_ASSET_PROXY_HOSTS = (
    "lookaside.fbsbx.com",
    "lookaside.instagram.com",
)
IMAGE_ASSET_PATH_HINTS = (
    "/images/",
    "/image/",
    "/media/",
    "/photo/",
    "/photos/",
    "/wp-content/uploads/",
    "/content/uploads/",
    "/uploads/",
    "/searches/",
)
IMAGE_QUERY_FORMAT_HINTS = ("fm", "format", "ext", "f")
IMAGE_SEARCH_MIN_WIDTH = 1200
IMAGE_SEARCH_MIN_HEIGHT = 700
IMAGE_SEARCH_MIN_ASPECT_RATIO = 1.35
IMAGE_SEARCH_RESULTS_PAGE = 0
IMAGE_SEARCH_NEGATIVE_TERMS = (
    "watermark",
    "logo",
    "poster",
    "flyer",
    "menu",
    "food",
    "drink",
    "cocktail",
    "beer",
    "wine",
    "restaurant",
    "brunch",
    "promo",
    "advertisement",
    "sponsor",
    "ticket",
)
IMAGE_SEARCH_NEGATIVE_CLAUSE = " ".join(f"-{term}" for term in IMAGE_SEARCH_NEGATIVE_TERMS)
IMAGE_REJECTION_TERMS = (
    "flyer",
    "poster",
    "event card",
    "event graphic",
    "promo graphic",
    "promotional banner",
    "ad creative",
    "advertisement",
    "screenshot",
    "screen shot",
    "collage",
    "composite",
    "social-media graphic",
    "social media graphic",
    "text overlay",
    "visible writing",
    "watermark",
    "photographer mark",
    "agency stamp",
    "copyright overlay",
    "logo bug",
    "food",
    "drink",
    "cocktail",
    "cocktails",
    "beer",
    "wine",
    "menu",
    "bottle service",
    "tabletop",
    "restaurant",
    "brunch",
    "promo",
    "sponsor",
    "brochure",
    "artist portrait",
    "press photo",
    "headshot",
    "portrait",
    "blurry",
    "distorted",
    "low-resolution",
    "low resolution",
    "low-res",
    "pixelated",
    "badly cropped",
    "cropped",
    "map",
    "seat map",
    "seating map",
)
TEXT_HEAVY_TERMS = (
    "text overlay",
    "writing",
    "caption",
    "lineup",
    "line-up",
    "date",
    "dates",
    "pricing",
    "price",
)
WATERMARK_TERMS = (
    "watermark",
    "copyright",
    "photographer mark",
    "agency stamp",
    "logo bug",
)
FOOD_DRINK_TERMS = (
    "food",
    "drink",
    "cocktail",
    "cocktails",
    "beer",
    "wine",
    "menu",
    "bottle service",
    "tabletop",
    "plate",
    "plates",
)
ARTIST_CENTRIC_TERMS = (
    "artist portrait",
    "press photo",
    "headshot",
    "portrait",
    "performer",
    "singer",
    "guitarist",
    "drummer",
    "dj",
    "rapper",
    "soloist",
    "frontman",
    "frontwoman",
)
VENUE_CONTEXT_TERMS = (
    "venue",
    "arena",
    "hall",
    "theater",
    "theatre",
    "club",
    "auditorium",
    "room",
    "crowd",
    "audience",
    "stage",
    "interior",
    "building",
    "facade",
    "entrance",
    "marquee",
    "balcony",
    "seating",
    "layout",
)
LIVE_PERFORMANCE_TERMS = (
    "live",
    "concert",
    "performance",
    "show",
    "gig",
    "festival",
    "crowd",
    "audience",
    "stage",
)
EXTERIOR_TERMS = (
    "exterior",
    "facade",
    "façade",
    "entrance",
    "marquee",
    "building",
    "front",
    "outside",
)
INTERIOR_TERMS = (
    "interior",
    "crowd",
    "room",
    "audience",
    "balcony",
    "seating",
)
STAGE_LAYOUT_TERMS = (
    "stage",
    "layout",
    "room",
    "audience",
    "crowd",
)
ARCHITECTURE_TERMS = (
    "architecture",
    "architectural",
    "facade",
    "marquee",
    "balcony",
    "ceiling",
    "building",
)
LOW_QUALITY_TERMS = (
    "blurry",
    "distorted",
    "low-resolution",
    "low resolution",
    "low-res",
    "pixelated",
    "badly cropped",
)
LOW_QUALITY_URL_HINTS = (
    "get_thumbnail=1",
    "thumbnail",
)
WATERMARK_KEYWORDS = (
    "getty",
    "shutterstock",
    "alamy",
    "istock",
    "dreamstime",
    "depositphotos",
    "watermark",
)
IMAGE_DOWNLOAD_MAX_BYTES = 8 * 1024 * 1024

# Cache the full Serp bundle per Google place id. The lock protects cache and
# in-flight bookkeeping only; all Serp network calls happen outside the lock.
place_serp_bundle_cache: dict[str, dict[str, Any]] = {}
place_serp_bundle_lock = threading.Lock()
place_serp_bundle_inflight: dict[str, threading.Event] = {}


def get_secret(name: str) -> str:
    env_value = os.environ.get(name)
    if clean_string(env_value):
        return clean_string(env_value)
    return subprocess.check_output(
        ["op", "read", f"op://{VAULT}/{name}/password"], text=True
    ).strip()


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_inline(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_string(value))


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_query_input(value: str) -> str:
    return normalize_spaces(clean_inline(value)).lower()


def get_serp_query_inputs(row: dict[str, Any], place_name: str) -> dict[str, str]:
    # Normalize the string inputs once so logically identical Serp queries map
    # to the same local cache key even when source rows vary in spacing/case.
    place_name = normalize_spaces(clean_inline(place_name))
    city = normalize_query_input(row.get("CITY") or row.get("City"))
    state = normalize_query_input(row.get("STATE") or row.get("State"))
    country = normalize_query_input(row.get("COUNTRY"))
    return {
        "place_name": place_name,
        "city": city,
        "state": state,
        "country": country,
    }


def is_blank(value: Any) -> bool:
    return clean_string(value) == ""


def valid_google_place_id(value: Any) -> str:
    text = clean_string(value)
    if not text:
        return ""
    if text.lower() in {"false", "none", "null", "nan"}:
        return ""
    return text


def sha1_text(value: str) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()


def parse_row_indices(raw: str | None) -> set[int]:
    if not raw:
        return set()
    values: set[int] = set()
    for piece in raw.split(","):
        piece = piece.strip()
        if not piece:
            continue
        values.add(int(piece))
    return values


def normalize_http_url(url: str) -> str:
    text = clean_string(url)
    if not text:
        return ""
    if text.startswith("x-raw-image://"):
        return ""
    if text.startswith("http://") or text.startswith("https://"):
        return text
    if text.startswith("www."):
        return f"https://{text}"
    return ""


def text_contains_any(text: str, phrases: Iterable[str]) -> bool:
    return any(phrase in text for phrase in phrases)


def parse_int(value: Any) -> int:
    try:
        return int(clean_string(value))
    except ValueError:
        return 0


def normalize_email(value: str) -> str:
    text = clean_string(value)
    if not text:
        return ""
    if text.startswith("mailto:"):
        text = text.split(":", 1)[1]
    match = re.search(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", text, re.I)
    return match.group(0) if match else ""


def canonical_domain(url: str) -> str:
    parsed = urlparse(url)
    netloc = parsed.netloc.lower().lstrip("www.")
    return netloc


def is_replaceable_website(url: str) -> bool:
    normalized = normalize_http_url(url)
    if not normalized:
        return False
    domain = canonical_domain(normalized)
    return any(domain == item or domain.endswith(f".{item}") for item in REPLACEABLE_WEBSITE_DOMAINS)


def classify_url(url: str, hinted_field: str = "") -> tuple[str, str]:
    if not url:
        return "", ""

    email = normalize_email(url)
    if email:
        return "Email", email

    normalized = normalize_http_url(url)
    if not normalized:
        return "", ""

    domain = canonical_domain(normalized)
    if any(domain.endswith(ignored) or domain == ignored for ignored in IGNORED_URL_DOMAINS):
        return "", ""

    field_hint = hinted_field.lower()
    if "ticket" in field_hint or "booking" in field_hint or domain.startswith(TICKET_DOMAINS):
        return "Tickets Link", normalized
    if "spotify" in field_hint or "spotify." in domain:
        return "Spotify URL", normalized
    if "youtube" in field_hint or domain in {"youtube.com", "m.youtube.com", "youtu.be"}:
        return "YouTube", normalized
    if "instagram" in field_hint or domain == "instagram.com":
        return "Instagram", normalized
    if "facebook" in field_hint or domain == "facebook.com":
        return "Facebook", normalized
    if "tiktok" in field_hint or domain == "tiktok.com":
        return "TikTok", normalized
    if domain in {"x.com", "twitter.com"}:
        return "X (Twitter)", normalized
    if field_hint in {"website", "officialsite", "official"}:
        return "Website", normalized
    if any(domain.startswith(prefix) for prefix in TICKET_DOMAINS):
        return "Tickets Link", normalized
    if domain == "open.spotify.com":
        return "Spotify URL", normalized
    if domain in {"instagram.com", "facebook.com", "x.com", "twitter.com", "tiktok.com"}:
        return classify_url(normalized, domain)
    return "Website", normalized


def merge_single_value(target: dict[str, str], field: str, value: str) -> None:
    if not field or not value:
        return
    if field == "Email":
        value = normalize_email(value)
    elif field != "Description":
        value = normalize_http_url(value) if field not in {"Name", "Address", "Phone", "PlacesID", "City", "State", "Zip Code", "Hours of Operation"} else clean_string(value)
    if not value:
        return
    if not target.get(field):
        target[field] = value


def get_first_value(record: dict[str, Any], column_names: list[str]) -> str:
    for name in column_names:
        if name in record and not is_blank(record[name]):
            return clean_string(record[name])
    return ""


def extract_urls_from_object(value: Any) -> list[str]:
    found: list[str] = []
    if isinstance(value, dict):
        for child in value.values():
            found.extend(extract_urls_from_object(child))
        return found
    if isinstance(value, list):
        for child in value:
            found.extend(extract_urls_from_object(child))
        return found
    if isinstance(value, str):
        for match in re.findall(r"https?://[^\s\"'<>]+", value):
            found.append(match.rstrip("),.]"))
    return found


def parse_same_as_json(raw: Any) -> dict[str, str]:
    payload = clean_string(raw)
    if not payload:
        return {}
    try:
        data = json.loads(payload)
    except json.JSONDecodeError:
        return {}

    result: dict[str, str] = {}
    if not isinstance(data, list):
        return result

    for item in data:
        if not isinstance(item, dict):
            continue
        identifier = clean_inline(item.get("identifier")).lower()
        url = clean_string(item.get("url"))
        field, normalized = classify_url(url, identifier)
        merge_single_value(result, field, normalized)
    return result


def format_hours(hours: Any) -> str:
    if not isinstance(hours, list):
        return ""

    values: dict[str, str] = {}
    extras: list[str] = []
    for item in hours:
        if not isinstance(item, dict):
            continue
        if len(item) != 1:
            extras.append(clean_inline(item))
            continue
        day, hours_value = next(iter(item.items()))
        day_name = clean_string(day).lower()
        formatted_hours = clean_inline(hours_value)
        if day_name in DAY_LABELS and formatted_hours:
            values[day_name] = formatted_hours
        elif formatted_hours:
            extras.append(f"{clean_inline(day)}: {formatted_hours}")

    lines = [f"{DAY_LABELS[day]}: {values[day]}" for day in DAY_ORDER if day in values]
    lines.extend(extras)
    return "\n".join(lines)


def normalize_google_hours(weekday_descriptions: Any) -> list[dict[str, str]]:
    if not isinstance(weekday_descriptions, list):
        return []

    normalized: list[dict[str, str]] = []
    for item in weekday_descriptions:
        text = clean_inline(item)
        if ":" not in text:
            continue
        day_name, hours_value = text.split(":", 1)
        day_key = clean_string(day_name).lower().replace(".", "")
        if day_key in DAY_LABELS and clean_inline(hours_value):
            normalized.append({day_key: clean_inline(hours_value)})
    return normalized


def normalize_google_place_details(payload: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}

    display_name = payload.get("displayName") or {}
    primary_type_display = payload.get("primaryTypeDisplayName") or {}
    editorial_summary = payload.get("editorialSummary") or {}
    regular_hours = payload.get("regularOpeningHours") or {}
    location = payload.get("location") or {}
    types = [
        clean_inline(value).replace("_", " ")
        for value in payload.get("types") or []
        if clean_inline(value)
    ]

    open_state = ""
    if regular_hours.get("openNow") is True:
        open_state = "Open"
    elif regular_hours.get("openNow") is False:
        open_state = "Closed"

    normalized: dict[str, Any] = {
        "title": clean_inline(display_name.get("text") or payload.get("displayName") or payload.get("name")),
        "address": clean_inline(payload.get("formattedAddress")),
        "phone": clean_inline(
            payload.get("internationalPhoneNumber") or payload.get("nationalPhoneNumber")
        ),
        "hours": normalize_google_hours(regular_hours.get("weekdayDescriptions")),
        "website": normalize_http_url(clean_string(payload.get("websiteUri"))),
        "type": clean_inline(primary_type_display.get("text") or payload.get("primaryType")),
        "type_id": clean_inline(payload.get("primaryType")),
        "description": clean_inline(editorial_summary.get("text")),
        "open_state": open_state,
        "business_status": clean_inline(payload.get("businessStatus")),
        "events": [],
        "links": [],
        "extensions": [{"types": types}] if types else [],
    }
    if location:
        normalized["gps_coordinates"] = {
            "latitude": clean_string(location.get("latitude")),
            "longitude": clean_string(location.get("longitude")),
        }
    return normalized


def style_hint_for(place_id: str) -> str:
    styles = [
        "scene-setting",
        "historical framing",
        "cultural significance",
        "first-impression tone",
    ]
    digest = int(sha1_text(place_id)[:8], 16)
    return styles[digest % len(styles)]


def word_count(text: str) -> int:
    return len(re.findall(r"\b[\w'.-]+\b", text))


def validate_description(text: str) -> list[str]:
    issues: list[str] = []
    normalized = clean_string(text)
    lowered = normalized.lower()
    if not normalized:
        issues.append("blank")
        return issues
    if normalized.count("musicroadtrip.com") != 1:
        issues.append("must_contain_exactly_one_musicroadtrip.com")
    count = word_count(normalized)
    if count < 60 or count > 120:
        issues.append(f"word_count_{count}")
    for phrase in BANNED_DESCRIPTION_PHRASES:
        if phrase in lowered:
            issues.append(f"contains_banned_phrase:{phrase}")
    return issues


def parse_output_text(payload: dict[str, Any]) -> str:
    texts: list[str] = []
    for item in payload.get("output") or []:
        if item.get("type") != "message":
            continue
        for content in item.get("content") or []:
            if content.get("type") == "output_text" and clean_string(content.get("text")):
                texts.append(content["text"])
    return "\n".join(texts).strip()


def extract_ocr_tokens(ocr_data: dict[str, list[Any]]) -> list[dict[str, int | str]]:
    texts = ocr_data.get("text") or []
    lefts = ocr_data.get("left") or []
    tops = ocr_data.get("top") or []
    widths = ocr_data.get("width") or []
    heights = ocr_data.get("height") or []

    tokens: list[dict[str, int | str]] = []
    for index, raw_text in enumerate(texts):
        text = re.sub(r"\s+", " ", clean_string(raw_text))
        if not text:
            continue
        tokens.append(
            {
                "text": text,
                "left": int(lefts[index]) if index < len(lefts) else 0,
                "top": int(tops[index]) if index < len(tops) else 0,
                "width": int(widths[index]) if index < len(widths) else 0,
                "height": int(heights[index]) if index < len(heights) else 0,
            }
        )
    return tokens


def analyze_ocr_tokens(
    tokens: list[dict[str, int | str]],
    image_width: int,
    image_height: int,
) -> list[str]:
    issues: list[str] = []
    combined_text = " ".join(clean_string(token["text"]).lower() for token in tokens)
    visible_words = [clean_string(token["text"]) for token in tokens if clean_string(token["text"])]
    visible_chars = len(re.sub(r"\s+", "", combined_text))

    if visible_chars >= 24 or len(visible_words) >= 5:
        issues.append("ocr_visible_text")
    if text_contains_any(combined_text, WATERMARK_KEYWORDS):
        issues.append("ocr_watermark_text")

    corner_hits = 0
    for token in tokens:
        left = int(token["left"])
        top = int(token["top"])
        width = int(token["width"])
        height = int(token["height"])
        right = left + width
        bottom = top + height
        near_left = left <= image_width * 0.18
        near_right = right >= image_width * 0.82
        near_top = top <= image_height * 0.18
        near_bottom = bottom >= image_height * 0.82
        if (near_left or near_right) and (near_top or near_bottom):
            if len(clean_string(token["text"])) >= 4:
                corner_hits += 1
    if corner_hits >= 2 and visible_chars >= 12:
        issues.append("ocr_corner_overlay")

    return dedupe_preserve_order(issues)


def inspect_image_bytes(image_bytes: bytes, content_type: str = "") -> dict[str, Any]:
    issues: list[str] = []
    try:
        image = Image.open(BytesIO(image_bytes))
        image.load()
    except Exception:
        return {
            "ok": False,
            "rejection_reasons": ["image_decode_failed"],
            "width": 0,
            "height": 0,
            "content_type": clean_string(content_type),
            "ocr_tokens": [],
        }

    width, height = image.size
    if clean_string(content_type) and not clean_string(content_type).lower().startswith("image/"):
        issues.append("non_image_content_type")
    if max(width, height) < 700:
        issues.append("binary_low_resolution")

    ocr_image = image.convert("L")
    if max(ocr_image.size) > 1600:
        ocr_image.thumbnail((1600, 1600))

    try:
        ocr_data = pytesseract.image_to_data(ocr_image, output_type=pytesseract.Output.DICT)
        ocr_tokens = extract_ocr_tokens(ocr_data)
    except Exception:
        ocr_tokens = []

    issues.extend(analyze_ocr_tokens(ocr_tokens, width, height))
    return {
        "ok": not issues,
        "rejection_reasons": dedupe_preserve_order(issues),
        "width": width,
        "height": height,
        "content_type": clean_string(content_type),
        "ocr_tokens": ocr_tokens,
    }


def inspect_remote_image_url(
    *,
    url: str,
    session: requests.Session,
    timeout: int,
) -> dict[str, Any]:
    response = session.get(url, timeout=timeout, stream=True)
    response.raise_for_status()
    content_type = clean_string(response.headers.get("Content-Type"))
    if content_type and not content_type.lower().startswith("image/"):
        return {
            "ok": False,
            "rejection_reasons": ["non_image_content_type"],
            "width": 0,
            "height": 0,
            "content_type": content_type,
            "ocr_tokens": [],
        }

    chunks: list[bytes] = []
    total = 0
    for chunk in response.iter_content(65536):
        if not chunk:
            continue
        total += len(chunk)
        if total > IMAGE_DOWNLOAD_MAX_BYTES:
            return {
                "ok": False,
                "rejection_reasons": ["image_too_large"],
                "width": 0,
                "height": 0,
                "content_type": content_type,
                "ocr_tokens": [],
            }
        chunks.append(chunk)
    return inspect_image_bytes(b"".join(chunks), content_type=content_type)


class CostTracker:
    def __init__(self, price_map: dict[str, float] | None = None) -> None:
        self.price_map = dict(price_map or SERVICE_COSTS)
        self.lock = threading.Lock()
        self.calls = {service: 0 for service in self.price_map}
        self.costs = {service: 0.0 for service in self.price_map}

    def record(self, service: str, units: int = 1) -> None:
        if service not in self.price_map or units <= 0:
            return
        with self.lock:
            self.calls[service] += units
            self.costs[service] += self.price_map[service] * units

    def snapshot(self) -> dict[str, Any]:
        with self.lock:
            by_service = {
                service: {
                    "calls": self.calls[service],
                    "cost": round(self.costs[service], 4),
                }
                for service in self.price_map
            }
        total = round(sum(item["cost"] for item in by_service.values()), 4)
        return {
            "by_service": by_service,
            "total": total,
        }


def build_cost_table(snapshot: dict[str, Any]) -> Table:
    table = Table(title="Estimated API Cost", expand=False)
    table.add_column("Service")
    table.add_column("Calls", justify="right")
    table.add_column("Cost ($)", justify="right")

    by_service = snapshot.get("by_service") or {}
    for service in ("google_places", "serpapi", "openai"):
        service_snapshot = by_service.get(service) or {}
        table.add_row(
            SERVICE_LABELS[service],
            str(service_snapshot.get("calls", 0)),
            f"{float(service_snapshot.get('cost', 0.0)):.2f}",
        )
    table.add_row("TOTAL", "-", f"{float(snapshot.get('total', 0.0)):.2f}")
    return table


class CostTableView:
    def __init__(self, tracker: CostTracker) -> None:
        self.tracker = tracker

    def __rich__(self) -> Table:
        return build_cost_table(self.tracker.snapshot())


def default_checkpoint_path(output_xlsx: Path) -> Path:
    return output_xlsx.with_suffix(".checkpoint.json")


def load_checkpoint(path: Path) -> set[int]:
    if not path.exists():
        return set()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return set()
    processed = payload.get("processed_rownums") or []
    return {int(value) for value in processed}


def save_checkpoint(path: Path, processed_rownums: set[int]) -> None:
    ensure_parent(path)
    payload = {
        "script_version": SCRIPT_VERSION,
        "saved_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "processed_rownums": sorted(processed_rownums),
    }
    temp_path = path.with_suffix(f"{path.suffix}.tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def read_existing_output_row_values(
    worksheet: Any,
    column_positions: dict[str, int],
    rownum: int,
) -> dict[str, str]:
    return {
        field: clean_string(worksheet.cell(rownum, column_positions[field]).value)
        for field in MAPOTIC_COLUMNS
    }


def row_has_written_output_values(final_values: dict[str, str]) -> bool:
    return any(clean_string(final_values.get(field)) for field in MAPOTIC_COLUMNS)


def verify_checkpoint_rows(
    rows: list[dict[str, Any]],
    processed_rownums: set[int],
    worksheet: Any,
    column_positions: dict[str, int],
    progress: "Progress" | None = None,
) -> set[int]:
    verified: set[int] = set()
    for row in rows:
        rownum = int(row["_rownum"])
        if rownum not in processed_rownums:
            continue
        existing_values = read_existing_output_row_values(worksheet, column_positions, rownum)
        if row_has_written_output_values(existing_values):
            verified.add(rownum)
            continue
        if progress is not None:
            progress.maybe_log(
                f"Checkpoint row {rownum} missing output values; reprocessing for safety.",
                force=True,
            )
    return verified


def cost_log_path(timestamp: dt.datetime) -> Path:
    return DEFAULT_COST_LOG_DIR / f"cost_log_{timestamp.strftime('%Y%m%dT%H%M%SZ')}.json"


def retry_delay_for_exception(
    exc: Exception,
    attempt: int,
    *,
    rate_limit_initial_backoff: float = OPENAI_429_INITIAL_BACKOFF,
    rate_limit_max_backoff: float = OPENAI_429_MAX_BACKOFF,
) -> float:
    default_delay = min(2 ** (attempt - 1), 8)
    response = exc.response if isinstance(exc, requests.HTTPError) else None
    if response is None or response.status_code != 429:
        return default_delay

    retry_after = clean_string(response.headers.get("Retry-After"))
    if retry_after:
        try:
            return max(float(retry_after), 0.0)
        except ValueError:
            retry_at = email.utils.parsedate_to_datetime(retry_after)
            if retry_at is not None:
                now = dt.datetime.now(retry_at.tzinfo or dt.timezone.utc)
                return max((retry_at - now).total_seconds(), 0.0)

    return min(rate_limit_initial_backoff * (2 ** (attempt - 1)), rate_limit_max_backoff)


class Progress:
    def __init__(self, heartbeat_seconds: int, console: Console | None = None) -> None:
        self.heartbeat_seconds = heartbeat_seconds
        self.last_log = time.monotonic()
        self.console = console

    def maybe_log(self, message: str, force: bool = False) -> None:
        now = time.monotonic()
        if force or (now - self.last_log) >= self.heartbeat_seconds:
            if self.console is not None:
                self.console.print(message)
            else:
                print(message, flush=True)
            self.last_log = now


class SerpApiClient:
    def __init__(
        self,
        api_key: str,
        cache_dir: Path,
        timeout: int = DEFAULT_TIMEOUT,
        request_pause: float = DEFAULT_REQUEST_PAUSE,
        cost_tracker: CostTracker | None = None,
    ) -> None:
        self.api_key = api_key
        self.cache_dir = cache_dir / "serp"
        self.timeout = timeout
        self.request_pause = request_pause
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})
        self.cache_hits = 0
        self.network_calls = 0
        self.cost_tracker = cost_tracker
        self.lock = threading.Lock()

    def _cache_path(self, namespace: str, params: dict[str, Any]) -> Path:
        normalized = {
            key: clean_string(value)
            for key, value in sorted(params.items())
            if value is not None and clean_string(value) != ""
        }
        digest = sha1_text(json.dumps(normalized, ensure_ascii=True, sort_keys=True))[:20]
        return self.cache_dir / namespace / f"{digest}.json"

    def request(self, namespace: str, params: dict[str, Any]) -> dict[str, Any]:
        with self.lock:
            cache_path = self._cache_path(namespace, params)
            if cache_path.exists():
                self.cache_hits += 1
                return json.loads(cache_path.read_text(encoding="utf-8"))

            request_params = {
                key: value
                for key, value in params.items()
                if value is not None and clean_string(value) != ""
            }
            request_params["api_key"] = self.api_key

            last_error: Exception | None = None
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    response = self.session.get(
                        SERPAPI_URL,
                        params=request_params,
                        timeout=self.timeout,
                    )
                    if response.status_code in {429, 500, 502, 503, 504}:
                        response.raise_for_status()
                    response.raise_for_status()
                    payload = response.json()
                    ensure_parent(cache_path)
                    cache_path.write_text(
                        json.dumps(payload, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )
                    self.network_calls += 1
                    if self.cost_tracker is not None:
                        self.cost_tracker.record("serpapi")
                    time.sleep(self.request_pause)
                    return payload
                except Exception as exc:  # noqa: BLE001
                    last_error = exc
                    if attempt == MAX_RETRIES:
                        break
                    time.sleep(min(2 ** (attempt - 1), 8))

        detail = str(last_error) if last_error else "unknown_error"
        raise RuntimeError(f"SerpApi request failed for {namespace}: {detail}") from last_error


class GooglePlacesClient:
    def __init__(
        self,
        api_key: str,
        cache_dir: Path,
        timeout: int = DEFAULT_TIMEOUT,
        request_pause: float = DEFAULT_REQUEST_PAUSE,
        cost_tracker: CostTracker | None = None,
    ) -> None:
        self.api_key = api_key
        self.cache_dir = cache_dir / "google_places"
        self.timeout = timeout
        self.request_pause = request_pause
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})
        self.cache_hits = 0
        self.network_calls = 0
        self.cost_tracker = cost_tracker
        self.lock = threading.Lock()

    def _cache_path(self, place_id: str, country_iso2: str) -> Path:
        digest = sha1_text(
            json.dumps(
                {
                    "place_id": clean_string(place_id),
                    "country_iso2": clean_string(country_iso2).upper(),
                    "field_mask": GOOGLE_PLACE_FIELD_MASK,
                },
                ensure_ascii=True,
                sort_keys=True,
            )
        )[:20]
        return self.cache_dir / f"{digest}.json"

    def lookup_place(self, place_id: str, country_iso2: str = "") -> dict[str, Any]:
        with self.lock:
            cache_path = self._cache_path(place_id, country_iso2)
            if cache_path.exists():
                self.cache_hits += 1
                return json.loads(cache_path.read_text(encoding="utf-8"))

            headers = {
                "X-Goog-Api-Key": self.api_key,
                "X-Goog-FieldMask": GOOGLE_PLACE_FIELD_MASK,
            }
            params = {
                "languageCode": "en",
                "regionCode": clean_string(country_iso2).upper() or None,
            }

            last_error: Exception | None = None
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    response = self.session.get(
                        f"{GOOGLE_PLACES_URL}/{place_id}",
                        params={key: value for key, value in params.items() if value is not None},
                        headers=headers,
                        timeout=self.timeout,
                    )
                    if response.status_code in {429, 500, 502, 503, 504}:
                        response.raise_for_status()
                    response.raise_for_status()
                    payload = normalize_google_place_details(response.json())
                    ensure_parent(cache_path)
                    cache_path.write_text(
                        json.dumps(payload, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )
                    self.network_calls += 1
                    if self.cost_tracker is not None:
                        self.cost_tracker.record("google_places")
                    time.sleep(self.request_pause)
                    return payload
                except Exception as exc:  # noqa: BLE001
                    last_error = exc
                    if attempt == MAX_RETRIES:
                        break
                    time.sleep(min(2 ** (attempt - 1), 8))

        detail = str(last_error) if last_error else "unknown_error"
        raise RuntimeError(f"Google Places lookup failed for {place_id}: {detail}") from last_error


class ImageInspectionClient:
    def __init__(
        self,
        cache_dir: Path,
        timeout: int = DEFAULT_TIMEOUT,
        request_pause: float = DEFAULT_REQUEST_PAUSE,
    ) -> None:
        self.cache_dir = cache_dir / "image_inspection"
        self.timeout = timeout
        self.request_pause = request_pause
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})
        self.cache_hits = 0
        self.network_calls = 0
        self.lock = threading.Lock()

    def _cache_path(self, url: str) -> Path:
        return self.cache_dir / f"{sha1_text(url)[:20]}.json"

    def inspect_url(self, url: str) -> dict[str, Any]:
        with self.lock:
            cache_path = self._cache_path(url)
            if cache_path.exists():
                self.cache_hits += 1
                return json.loads(cache_path.read_text(encoding="utf-8"))

            try:
                payload = inspect_remote_image_url(
                    url=url,
                    session=self.session,
                    timeout=self.timeout,
                )
            except Exception as exc:  # noqa: BLE001
                payload = {
                    "ok": False,
                    "rejection_reasons": [f"image_download_failed:{type(exc).__name__}"],
                    "width": 0,
                    "height": 0,
                    "content_type": "",
                    "ocr_tokens": [],
                }
            ensure_parent(cache_path)
            cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            self.network_calls += 1
            time.sleep(self.request_pause)
            return payload


class OpenAIClient:
    def __init__(
        self,
        api_key: str,
        cache_dir: Path,
        model: str = DEFAULT_OPENAI_MODEL,
        timeout: int = DEFAULT_TIMEOUT,
        request_pause: float = DEFAULT_REQUEST_PAUSE,
        cost_tracker: CostTracker | None = None,
    ) -> None:
        self.api_key = api_key
        self.cache_dir = cache_dir / "openai"
        self.model = model
        self.timeout = timeout
        self.request_pause = request_pause
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            }
        )
        self.cache_hits = 0
        self.network_calls = 0
        self.cost_tracker = cost_tracker
        self.lock = threading.Lock()

    def _description_cache_path(self, place_id: str) -> Path:
        return self.cache_dir / "descriptions" / f"{sha1_text(place_id)[:20]}.json"

    def load_cached_description(self, place_id: str) -> str:
        cache_path = self._description_cache_path(place_id)
        if not cache_path.exists():
            return ""
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
        if payload.get("version") != SCRIPT_VERSION:
            return ""
        self.cache_hits += 1
        return clean_string(payload.get("description"))

    def save_cached_description(self, place_id: str, description: str) -> None:
        cache_path = self._description_cache_path(place_id)
        payload = {
            "version": SCRIPT_VERSION,
            "model": self.model,
            "place_id": place_id,
            "generated_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
            "description": description,
        }
        ensure_parent(cache_path)
        cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def describe_batch(self, items: list[dict[str, Any]]) -> dict[str, str]:
        schema = {
            "type": "object",
            "properties": {
                "descriptions": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "place_id": {"type": "string"},
                            "description": {"type": "string"},
                        },
                        "required": ["place_id", "description"],
                        "additionalProperties": False,
                    },
                }
            },
            "required": ["descriptions"],
            "additionalProperties": False,
        }

        system_text = (
            "Return JSON only. You are writing Mapotic venue descriptions in the voice of a "
            "senior Rolling Stone Magazine travel writer. Every description must be 60-120 words, "
            "must contain exactly one mention of musicroadtrip.com, must feel specific to the venue, "
            "must include at least one sensory, experiential, or cultural detail, and must avoid the "
            "banned openings 'Nestled in the heart of', 'Located in', and 'Known for'. Use only the "
            "facts supplied in the venue payload. Do not invent named festivals, artists, historical "
            "claims, capacities, renovations, awards, or specific events that are not explicitly "
            "present. Respect the requested style_hint so openings and structure vary."
        )
        user_text = json.dumps({"venues": items}, ensure_ascii=False)
        payload = {
            "model": self.model,
            "max_output_tokens": 5000,
            "input": [
                {
                    "role": "system",
                    "content": [{"type": "input_text", "text": system_text}],
                },
                {
                    "role": "user",
                    "content": [{"type": "input_text", "text": user_text}],
                },
            ],
            "text": {
                "format": {
                    "type": "json_schema",
                    "name": "venue_descriptions",
                    "strict": True,
                    "schema": schema,
                }
            },
        }

        with self.lock:
            last_error: Exception | None = None
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    response = self.session.post(
                        OPENAI_RESPONSES_URL,
                        json=payload,
                        timeout=self.timeout,
                    )
                    if response.status_code in {429, 500, 502, 503, 504}:
                        response.raise_for_status()
                    response.raise_for_status()
                    self.network_calls += 1
                    if self.cost_tracker is not None:
                        self.cost_tracker.record("openai")
                    response_payload = response.json()
                    raw_text = parse_output_text(response_payload)
                    parsed = json.loads(raw_text)
                    result = {
                        clean_string(item.get("place_id")): clean_string(item.get("description"))
                        for item in parsed.get("descriptions") or []
                        if clean_string(item.get("place_id")) and clean_string(item.get("description"))
                    }
                    if result:
                        time.sleep(self.request_pause)
                        return result
                    raise RuntimeError("OpenAI returned no descriptions.")
                except Exception as exc:  # noqa: BLE001
                    last_error = exc
                    if attempt == MAX_RETRIES:
                        break
                    time.sleep(retry_delay_for_exception(exc, attempt))

        detail = str(last_error) if last_error else "unknown_error"
        raise RuntimeError(f"OpenAI description batch failed: {detail}") from last_error


def build_row_contexts(
    input_xlsx: Path,
    limit: int | None = None,
    row_indices: set[int] | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(input_xlsx, read_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean_string(value) for value in next(rows)]
    contexts: list[dict[str, Any]] = []
    selected = 0

    for rownum, row in enumerate(rows, start=2):
        if row_indices and rownum not in row_indices:
            continue
        record = {headers[index]: row[index] for index in range(len(headers))}
        record["_rownum"] = rownum
        contexts.append(record)
        selected += 1
        if limit is not None and selected >= limit:
            break

    return headers, contexts


def extract_place_urls(place_results: dict[str, Any]) -> dict[str, str]:
    merged: dict[str, str] = {}
    if not isinstance(place_results, dict):
        return merged

    merge_single_value(merged, "Website", clean_string(place_results.get("website")))
    merge_single_value(merged, "Phone", clean_string(place_results.get("phone")))
    merge_single_value(merged, "Tickets Link", clean_string(place_results.get("booking_link")))

    for raw_url in extract_urls_from_object(place_results.get("links")):
        field, normalized = classify_url(raw_url)
        merge_single_value(merged, field, normalized)

    for event in place_results.get("events") or []:
        if not isinstance(event, dict):
            continue
        for key, value in event.items():
            hinted_field = key if "ticket" in key.lower() or "booking" in key.lower() else ""
            if isinstance(value, str):
                field, normalized = classify_url(value, hinted_field)
                merge_single_value(merged, field, normalized)
            else:
                for raw_url in extract_urls_from_object(value):
                    field, normalized = classify_url(raw_url, hinted_field)
                    merge_single_value(merged, field, normalized)

    return merged


def is_valid_social_candidate(url: str, field: str) -> bool:
    normalized = normalize_http_url(url)
    if not normalized:
        return False
    lowered = normalized.lower()
    config = SOCIAL_SEARCH_CONFIG.get(field)
    if not config:
        return False
    return not any(fragment in lowered for fragment in config["disallowed_fragments"])


def build_social_search_query(field: str, row: dict[str, Any], place_name: str) -> str:
    del field
    query_inputs = get_serp_query_inputs(row, place_name)
    parts = [f'"{query_inputs["place_name"]}"'] if query_inputs["place_name"] else []
    if query_inputs["city"]:
        parts.append(query_inputs["city"])
    if query_inputs["country"]:
        parts.append(query_inputs["country"])
    parts.append(
        "site:instagram.com OR site:facebook.com OR site:tiktok.com OR site:x.com OR "
        "site:twitter.com OR site:youtube.com"
    )
    return " ".join(parts)


def search_missing_socials(
    serp: SerpApiClient,
    row: dict[str, Any],
    place_name: str,
    current_fields: dict[str, str],
) -> dict[str, str]:
    missing_fields = [
        field
        for field in ("Instagram", "Facebook", "X (Twitter)", "TikTok", "YouTube")
        if not current_fields.get(field)
    ]
    if not missing_fields:
        return {}

    payload = serp.request(
        "google_social_bundle",
        {
            "engine": "google",
            "q": build_social_search_query("", row, place_name),
            "hl": "en",
            "gl": clean_string(row.get("COUNTRY_ISO2")).lower() or None,
            "num": 10,
        },
    )

    results: dict[str, str] = {}
    # Classify the single response back into every canonical Mapotic social
    # column so the cached place bundle stays reusable across duplicate rows.
    for organic in payload.get("organic_results") or []:
        for raw_url in extract_urls_from_object(organic):
            classified_field, normalized = classify_url(raw_url)
            if (
                classified_field in SOCIAL_SEARCH_CONFIG
                and normalized
                and is_valid_social_candidate(normalized, classified_field)
                and classified_field not in results
            ):
                results[classified_field] = normalized
        if len(results) >= len(SOCIAL_SEARCH_CONFIG):
            break
    return results


def has_image_extension(path: str) -> bool:
    lower = path.lower()
    return any(lower.endswith(extension) for extension in IMAGE_FILE_EXTENSIONS)


def has_page_extension(path: str) -> bool:
    lower = path.lower()
    return any(lower.endswith(extension) for extension in IMAGE_PAGE_EXTENSIONS)


def infer_image_size_from_url(url: str) -> tuple[int, int]:
    normalized = normalize_http_url(url)
    if not normalized:
        return (0, 0)
    width_match = re.search(r"[?&]width=(\d+)", normalized, re.I)
    height_match = re.search(r"[?&]height=(\d+)", normalized, re.I)
    if width_match and height_match:
        return (int(width_match.group(1)), int(height_match.group(1)))

    compact_match = re.search(r"[=/_-]w(\d{2,5})[-_]?h(\d{2,5})\b", normalized, re.I)
    if compact_match:
        return (int(compact_match.group(1)), int(compact_match.group(2)))
    return (0, 0)


def is_probable_direct_image_asset_url(url: str) -> bool:
    normalized = normalize_http_url(url)
    if not normalized:
        return False

    parsed = urlparse(normalized)
    host = parsed.netloc.lower()
    path = parsed.path.lower()

    if any(host == blocked or host.endswith(f".{blocked}") for blocked in IMAGE_PAGE_HOSTS):
        return False
    if host in IMAGE_ASSET_PROXY_HOSTS:
        return False
    if has_page_extension(path):
        return False
    if any(marker in normalized.lower() for marker in ("/posts/", "/post/", "/status/", "/reel/", "/pin/")):
        return False
    if has_image_extension(path):
        return True
    if host.endswith("googleusercontent.com") or host.endswith("gstatic.com"):
        return True
    if any(hint in path for hint in IMAGE_ASSET_PATH_HINTS):
        return True
    if any(f"{key}=" in normalized.lower() for key in IMAGE_QUERY_FORMAT_HINTS):
        return True
    if any(token in host for token in ("cdn", "img", "image", "images", "media", "static")):
        return True
    return False


def normalize_image_url_for_dedupe(url: str) -> str:
    normalized = normalize_http_url(url)
    if not normalized:
        return ""
    parsed = urlparse(normalized)
    host = parsed.netloc.lower()
    path = parsed.path.lower().rstrip("/")
    path = re.sub(r"([_-])(\d{2,5}x\d{2,5}|w\d+|h\d+|thumb|thumbnail|small|medium|large|wide|resized|cropped|copy)(?=\.|$)", "", path)
    return f"{host}{path}"


def build_image_candidate(item: dict[str, Any], query_type: str) -> dict[str, Any]:
    url = normalize_http_url(clean_string(item.get("original"))) or normalize_http_url(
        clean_string(item.get("image"))
    )
    title = clean_inline(item.get("title"))
    source = clean_inline(item.get("source"))
    link = normalize_http_url(clean_string(item.get("link")))
    snippet = f"{title} {source} {clean_inline(link)}".lower()
    width = parse_int(item.get("original_width"))
    height = parse_int(item.get("original_height"))
    if width == 0 or height == 0:
        width, height = infer_image_size_from_url(url)
    position = parse_int(item.get("position")) or 999

    rejection_reasons: list[str] = []
    if not is_probable_direct_image_asset_url(url):
        rejection_reasons.append("not_direct_public_image_asset")
    if text_contains_any(snippet, IMAGE_REJECTION_TERMS):
        rejection_reasons.append("disallowed_image_type_or_content")
    if text_contains_any(snippet, TEXT_HEAVY_TERMS):
        rejection_reasons.append("visible_text_or_writing")
    if text_contains_any(snippet, WATERMARK_TERMS):
        rejection_reasons.append("watermark_or_logo_overlay")
    if text_contains_any(snippet, FOOD_DRINK_TERMS):
        rejection_reasons.append("food_or_drink_focus")
    if text_contains_any(snippet, LOW_QUALITY_TERMS):
        rejection_reasons.append("low_quality")
    if text_contains_any(url.lower(), LOW_QUALITY_URL_HINTS):
        rejection_reasons.append("thumbnail_or_low_quality_url")
    if item.get("is_product"):
        rejection_reasons.append("product_image")

    is_landscape = width > 0 and height > 0 and width >= int(height * 1.15)
    is_portrait = width > 0 and height > 0 and height >= int(width * 1.15)
    aspect_ratio = (width / height) if width > 0 and height > 0 else 0.0
    if width > 0 and width < IMAGE_SEARCH_MIN_WIDTH:
        rejection_reasons.append("low_resolution")
    if height > 0 and height < IMAGE_SEARCH_MIN_HEIGHT:
        rejection_reasons.append("low_resolution")
    if width > 0 and height > 0 and aspect_ratio < IMAGE_SEARCH_MIN_ASPECT_RATIO:
        rejection_reasons.append("narrow_or_vertical_framing")
    if width > 0 and height > 0 and max(width, height) < 700:
        rejection_reasons.append("low_resolution")

    categories: set[str] = set()
    if text_contains_any(snippet, VENUE_CONTEXT_TERMS):
        categories.add("venue")
    if text_contains_any(snippet, LIVE_PERFORMANCE_TERMS):
        categories.add("live")
    if text_contains_any(snippet, EXTERIOR_TERMS):
        categories.add("exterior")
    if text_contains_any(snippet, INTERIOR_TERMS):
        categories.add("interior")
    if text_contains_any(snippet, STAGE_LAYOUT_TERMS):
        categories.add("stage")
    if text_contains_any(snippet, ARCHITECTURE_TERMS):
        categories.add("architecture")

    if query_type == "live":
        categories.add("live_query")
    if query_type == "exterior":
        categories.add("exterior_query")
    if query_type == "unified":
        categories.add("unified_query")
    if query_type == "fallback":
        categories.add("fallback")

    artist_centric = text_contains_any(snippet, ARTIST_CENTRIC_TERMS)
    if artist_centric and "venue" not in categories:
        rejection_reasons.append("artist_centric_without_venue_context")

    score = 1000 - (position * 5)
    if query_type in {"live", "unified"} and "live" in categories and "venue" in categories:
        score += 180
    elif query_type in {"live", "unified"} and "live" in categories:
        score += 90
    if query_type in {"exterior", "unified"} and "exterior" in categories:
        score += 140
    if "interior" in categories:
        score += 50
    if "stage" in categories:
        score += 40
    if "architecture" in categories:
        score += 30
    if is_landscape:
        score += 35
    if is_portrait:
        score -= 20
    if aspect_ratio >= 1.6:
        score += 20
    elif aspect_ratio >= IMAGE_SEARCH_MIN_ASPECT_RATIO:
        score += 10
    if width >= 1600:
        score += 20
    elif width >= 1200:
        score += 10
    if artist_centric:
        score -= 80
    if query_type == "live":
        positive_terms = CONCERT_POSITIVE_TERMS
        negative_terms = CONCERT_NEGATIVE_TERMS
    elif query_type == "unified":
        positive_terms = tuple(
            dedupe_preserve_order(list(CONCERT_POSITIVE_TERMS) + list(EXTERIOR_POSITIVE_TERMS))
        )
        negative_terms = tuple(
            dedupe_preserve_order(list(CONCERT_NEGATIVE_TERMS) + list(EXTERIOR_NEGATIVE_TERMS))
        )
    else:
        positive_terms = EXTERIOR_POSITIVE_TERMS
        negative_terms = EXTERIOR_NEGATIVE_TERMS

    for term in positive_terms:
        if term in snippet:
            score += 18
    for term in negative_terms:
        if term in snippet:
            score -= 35
    if canonical_domain(url) in {"pinterest.com"}:
        score -= 100

    return {
        "url": url,
        "title": title,
        "source": source,
        "link": link,
        "query_type": query_type,
        "width": width,
        "height": height,
        "position": position,
        "categories": categories,
        "artist_centric": artist_centric,
        "is_landscape": is_landscape,
        "score": score,
        "fingerprint": normalize_image_url_for_dedupe(url),
        "rejection_reasons": dedupe_preserve_order(rejection_reasons),
    }


def build_place_thumbnail_candidate(place_results: dict[str, Any]) -> dict[str, Any] | None:
    thumbnail = extract_place_thumbnail(place_results)
    if not thumbnail:
        return None
    candidate = build_image_candidate(
        {
            "original": thumbnail,
            "title": clean_string(place_results.get("title")),
            "source": "google maps place",
            "position": 999,
        },
        query_type="fallback",
    )
    return candidate


def build_place_photo_candidates(
    place_results: dict[str, Any],
    place_photo_results: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    place_name = clean_string(place_results.get("title")) or "venue"
    candidates: list[dict[str, Any]] = []
    for position, photo in enumerate(place_photo_results, start=1):
        if not isinstance(photo, dict):
            continue
        candidate = build_image_candidate(
            {
                "original": clean_string(photo.get("image")) or clean_string(photo.get("thumbnail")),
                "title": f"{place_name} venue place photo",
                "source": "google place photos",
                "link": clean_string(photo.get("image")),
                "position": position,
            },
            query_type="fallback",
        )
        if candidate["width"] > candidate["height"] and (candidate["width"] / max(candidate["height"], 1)) >= 1.25:
            candidate["rejection_reasons"] = [
                reason for reason in candidate["rejection_reasons"] if reason != "narrow_or_vertical_framing"
            ]
        candidate["categories"].add("venue")
        candidate["categories"].add("place_photo")
        candidate["score"] += 45
        candidates.append(candidate)
    return candidates


def accepted_place_photo_candidates(
    place_results: dict[str, Any],
    place_photo_results: list[dict[str, Any]],
    image_inspector: ImageInspectionClient | None = None,
) -> list[dict[str, Any]]:
    place_photo_candidates = build_place_photo_candidates(place_results, place_photo_results)
    place_photo_candidates = [
        candidate for candidate in place_photo_candidates if not candidate["rejection_reasons"] and candidate["url"]
    ]
    place_photo_candidates.sort(key=lambda candidate: candidate["score"], reverse=True)
    if image_inspector is None:
        return place_photo_candidates

    inspected_place_photo_candidates: list[dict[str, Any]] = []
    for candidate in place_photo_candidates[:8]:
        inspection = image_inspector.inspect_url(candidate["url"])
        candidate["binary_inspection"] = inspection
        candidate["rejection_reasons"] = dedupe_preserve_order(
            candidate["rejection_reasons"] + inspection.get("rejection_reasons", [])
        )
        if not candidate["rejection_reasons"]:
            inspected_place_photo_candidates.append(candidate)
        if len(inspected_place_photo_candidates) >= 6:
            break
    return inspected_place_photo_candidates


def accepted_image_candidates(
    images_results: list[dict[str, Any]],
    query_type: str,
    image_inspector: ImageInspectionClient | None = None,
) -> list[dict[str, Any]]:
    candidates = []
    for item in images_results:
        if not isinstance(item, dict):
            continue
        candidate = build_image_candidate(item, query_type=query_type)
        candidates.append(candidate)

    accepted = [candidate for candidate in candidates if not candidate["rejection_reasons"] and candidate["url"]]
    accepted.sort(key=lambda candidate: candidate["score"], reverse=True)
    if image_inspector is None:
        return accepted

    inspected_accepted: list[dict[str, Any]] = []
    for candidate in accepted[:8]:
        inspection = image_inspector.inspect_url(candidate["url"])
        candidate["binary_inspection"] = inspection
        candidate["rejection_reasons"] = dedupe_preserve_order(
            candidate["rejection_reasons"] + inspection.get("rejection_reasons", [])
        )
        if not candidate["rejection_reasons"]:
            inspected_accepted.append(candidate)
        if len(inspected_accepted) >= 6:
            break
    return inspected_accepted


def pick_candidate(
    candidates: list[dict[str, Any]],
    *,
    require_categories: set[str] | None = None,
    exclude_fingerprints: set[str] | None = None,
) -> dict[str, Any] | None:
    required = require_categories or set()
    excluded = exclude_fingerprints or set()
    for candidate in candidates:
        if candidate["fingerprint"] in excluded:
            continue
        if required and not required.issubset(candidate["categories"]):
            continue
        return candidate
    return None


def format_additional_image_value(urls: list[str], use_delete_sentinel: bool = False) -> str:
    if use_delete_sentinel:
        return "$DELETE"
    cleaned = dedupe_preserve_order([normalize_http_url(url) for url in urls if normalize_http_url(url)])
    return "$".join(cleaned)


def validate_additional_image_value(value: str) -> list[str]:
    text = clean_string(value)
    if not text:
        return []
    if text == "$DELETE":
        return []

    issues: list[str] = []
    if "$DELETE" in text:
        issues.append("delete_sentinel_mixed_with_urls")
    if text.startswith("$") or text.endswith("$"):
        issues.append("leading_or_trailing_separator")
    parts = text.split("$")
    if any(not clean_string(part) for part in parts):
        issues.append("empty_gallery_entry")
    for part in parts:
        if not is_probable_direct_image_asset_url(part):
            issues.append("gallery_contains_non_direct_asset_url")
            break
    return dedupe_preserve_order(issues)


def select_image_fields(
    concert_results: list[dict[str, Any]],
    exterior_results: list[dict[str, Any]],
    place_results: dict[str, Any],
    place_photo_results: list[dict[str, Any]] | None = None,
    image_inspector: ImageInspectionClient | None = None,
    gallery_mode: str = "mirror-main",
) -> dict[str, str]:
    place_photo_candidates = accepted_place_photo_candidates(
        place_results,
        place_photo_results or [],
        image_inspector=image_inspector,
    )

    # Unified Serp image search is classified once into live/exterior/interior/
    # stage buckets so selection can preserve the old priorities without
    # issuing separate network searches for each image type.
    unified_candidates = accepted_image_candidates(
        [*concert_results, *exterior_results],
        query_type="unified",
        image_inspector=image_inspector,
    )
    fallback_candidates = []
    thumbnail_candidate = build_place_thumbnail_candidate(place_results)
    if thumbnail_candidate and image_inspector is not None and not thumbnail_candidate["rejection_reasons"]:
        inspection = image_inspector.inspect_url(thumbnail_candidate["url"])
        thumbnail_candidate["binary_inspection"] = inspection
        thumbnail_candidate["rejection_reasons"] = dedupe_preserve_order(
            thumbnail_candidate["rejection_reasons"] + inspection.get("rejection_reasons", [])
        )
    if thumbnail_candidate and not thumbnail_candidate["rejection_reasons"]:
        fallback_candidates.append(thumbnail_candidate)

    main_candidate = (
        pick_candidate(place_photo_candidates, require_categories={"live", "venue"})
        or pick_candidate(unified_candidates, require_categories={"live", "venue"})
        or pick_candidate(place_photo_candidates, require_categories={"exterior"})
        or pick_candidate(unified_candidates, require_categories={"exterior"})
        or pick_candidate(place_photo_candidates)
        or pick_candidate(unified_candidates)
        or pick_candidate(fallback_candidates)
    )
    if not main_candidate:
        return {
            "Main Image URL": "",
            "Additional Image URL(s)": "",
            "image_status": "no_images_found",
        }

    selected_candidates = [main_candidate]
    seen_fingerprints = {main_candidate["fingerprint"]}

    for required in ({"exterior"}, {"interior"}, {"stage"}, {"architecture"}):
        candidate = (
            pick_candidate(place_photo_candidates, require_categories=required, exclude_fingerprints=seen_fingerprints)
            or pick_candidate(unified_candidates, require_categories=required, exclude_fingerprints=seen_fingerprints)
            or pick_candidate(fallback_candidates, require_categories=required, exclude_fingerprints=seen_fingerprints)
        )
        if candidate:
            selected_candidates.append(candidate)
            seen_fingerprints.add(candidate["fingerprint"])

    for pool in (place_photo_candidates, unified_candidates, fallback_candidates):
        candidate = pick_candidate(pool, exclude_fingerprints=seen_fingerprints)
        if candidate:
            selected_candidates.append(candidate)
            seen_fingerprints.add(candidate["fingerprint"])
        if len(selected_candidates) >= 4:
            break

    additional_urls = [candidate["url"] for candidate in selected_candidates if candidate["url"]]
    if gallery_mode == "separate-main" and additional_urls:
        additional_urls = additional_urls[1:]
    additional_value = format_additional_image_value(additional_urls)
    if validate_additional_image_value(additional_value):
        additional_value = ""

    if "live" in main_candidate["categories"] and "venue" in main_candidate["categories"]:
        image_status = "main_live_performance_with_ordered_gallery"
    elif "exterior" in main_candidate["categories"]:
        image_status = "main_exterior_fallback_with_ordered_gallery"
    else:
        image_status = "main_general_venue_image_with_ordered_gallery"

    return {
        "Main Image URL": main_candidate["url"],
        "Additional Image URL(s)": additional_value,
        "image_status": image_status,
    }


def dedupe_preserve_order(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = clean_string(value)
        if cleaned and cleaned not in seen:
            result.append(cleaned)
            seen.add(cleaned)
    return result


def build_search_location(row: dict[str, Any]) -> str:
    query_inputs = get_serp_query_inputs(row, clean_inline(row.get("NAME")))
    parts = [part for part in [query_inputs["city"], query_inputs["state"], query_inputs["country"]] if part]
    return ", ".join(parts)


def build_unified_image_query(row: dict[str, Any], place_name: str) -> str:
    query_inputs = get_serp_query_inputs(row, place_name)
    parts = [f'"{query_inputs["place_name"]}"'] if query_inputs["place_name"] else []
    if query_inputs["city"]:
        parts.append(query_inputs["city"])
    if query_inputs["country"]:
        parts.append(query_inputs["country"])
    parts.extend(
        [
            '("live music venue" OR "concert venue")',
            '("concert" OR "crowd" OR "stage" OR "exterior" OR "facade" OR "entrance")',
            '("wide shot" OR "establishing shot")',
            IMAGE_SEARCH_NEGATIVE_CLAUSE,
        ]
    )
    return " ".join(parts)


def search_images(
    serp: SerpApiClient,
    query: str,
    location: str,
    country_iso2: str,
) -> list[dict[str, Any]]:
    base_params = {
        "engine": "google_images",
        "q": query,
        "google_domain": "google.com",
        "gl": clean_string(country_iso2).lower() or None,
        "hl": "en",
        "imgar": "w",
        "image_type": "photo",
        "imgsz": "4mp",
        "safe": "active",
        "filter": "1",
        "no_cache": True,
        "ijn": IMAGE_SEARCH_RESULTS_PAGE,
    }
    last_error: Exception | None = None
    for location_value in [location or None, None]:
        params = dict(base_params)
        params["location"] = location_value
        try:
            payload = serp.request("google_images", params)
            return payload.get("images_results") or []
        except RuntimeError as exc:
            last_error = exc
            if location_value and "400 Client Error" in str(exc):
                continue
            raise
    if last_error:
        raise last_error
    return []


def search_place_photo_seed(
    serp: SerpApiClient,
    place_id: str,
    country_iso2: str,
) -> dict[str, Any]:
    payload = serp.request(
        "google_maps_place",
        {
            "engine": "google_maps",
            "place_id": place_id,
            "hl": "en",
            "gl": clean_string(country_iso2).lower() or None,
        },
    )
    place_results = payload.get("place_results") or {}
    return place_results if isinstance(place_results, dict) else {}


def search_place_photos(serp: SerpApiClient, place_results: dict[str, Any]) -> list[dict[str, Any]]:
    data_id = clean_string(place_results.get("data_id"))
    if not data_id:
        return []
    payload = serp.request(
        "google_maps_photos",
        {
            "engine": "google_maps_photos",
            "data_id": data_id,
            "hl": "en",
        },
    )
    photos = payload.get("photos") or []
    return photos if isinstance(photos, list) else []


def extract_place_thumbnail(place_results: dict[str, Any]) -> str:
    thumbnail = normalize_http_url(clean_string(place_results.get("thumbnail")))
    if thumbnail:
        return thumbnail
    images = place_results.get("images") or []
    for item in images:
        if not isinstance(item, dict):
            continue
        thumbnail = normalize_http_url(clean_string(item.get("thumbnail")))
        if thumbnail:
            return thumbnail
    return ""


def merge_place_photo_seed(
    place_results: dict[str, Any],
    place_photo_seed: dict[str, Any],
) -> dict[str, Any]:
    merged = dict(place_results)
    if clean_string(place_photo_seed.get("data_id")):
        merged["data_id"] = clean_string(place_photo_seed.get("data_id"))
    if normalize_http_url(clean_string(place_photo_seed.get("thumbnail"))):
        merged["thumbnail"] = normalize_http_url(clean_string(place_photo_seed.get("thumbnail")))
    if isinstance(place_photo_seed.get("images"), list) and place_photo_seed.get("images"):
        merged["images"] = place_photo_seed.get("images")
    return merged


def build_place_context(row: dict[str, Any], place_results: dict[str, Any]) -> dict[str, Any]:
    title = clean_inline(place_results.get("title")) or clean_inline(row.get("NAME"))
    description = clean_inline(place_results.get("description"))
    type_name = clean_inline(place_results.get("type") or place_results.get("type_id"))
    open_state = clean_inline(place_results.get("open_state"))
    highlights: list[str] = []
    for extension in place_results.get("extensions") or []:
        if not isinstance(extension, dict):
            continue
        for values in extension.values():
            if isinstance(values, list):
                highlights.extend(clean_inline(value) for value in values if clean_inline(value))
    event_titles: list[str] = []
    for event in place_results.get("events") or []:
        if isinstance(event, dict) and clean_inline(event.get("title")):
            event_titles.append(clean_inline(event.get("title")))
    return {
        "place_id": valid_google_place_id(row.get("GOOGLE_PLACE_ID")),
        "style_hint": style_hint_for(valid_google_place_id(row.get("GOOGLE_PLACE_ID"))),
        "name": title,
        "city": clean_inline(row.get("CITY") or row.get("City")),
        "state": clean_inline(row.get("STATE") or row.get("State")),
        "country": clean_inline(row.get("COUNTRY")),
        "address": clean_inline(place_results.get("address") or row.get("ADDRESS")),
        "place_type": type_name,
        "source_description": description,
        "open_state": open_state,
        "highlights": highlights[:5],
        "event_titles": event_titles[:3],
    }


def get_place_serp_bundle(
    place_id: str,
    row: dict[str, Any],
    serp_client: SerpApiClient,
    *,
    place_results: dict[str, Any] | None = None,
    current_fields: dict[str, str] | None = None,
    image_inspector: ImageInspectionClient | None = None,
    skip_images: bool = False,
    gallery_mode: str = "mirror-main",
) -> dict[str, Any]:
    normalized_place_id = valid_google_place_id(place_id)
    if not normalized_place_id:
        return {
            "place_photo_seed": {},
            "place_photos": [],
            "image_results": [],
            "concert_images": [],
            "exterior_images": [],
            "socials": {},
            "images_requested": not skip_images,
            "cache_units": 0,
        }

    should_include_images = not skip_images
    while True:
        with place_serp_bundle_lock:
            cached_bundle = place_serp_bundle_cache.get(normalized_place_id)
            if cached_bundle and (not should_include_images or cached_bundle.get("images_requested")):
                if hasattr(serp_client, "cache_hits"):
                    serp_client.cache_hits += max(1, int(cached_bundle.get("cache_units", 1)))
                return copy.deepcopy(cached_bundle)

            inflight = place_serp_bundle_inflight.get(normalized_place_id)
            if inflight is None:
                inflight = threading.Event()
                place_serp_bundle_inflight[normalized_place_id] = inflight
                break
        inflight.wait()

    try:
        resolved_place_results = dict(place_results or {})
        resolved_current_fields = dict(current_fields or {})
        cache_units = 0
        country_iso2 = clean_string(row.get("COUNTRY_ISO2"))
        place_photo_seed: dict[str, Any] = {}
        place_photo_results: list[dict[str, Any]] = []
        image_results: list[dict[str, Any]] = []
        place_name = clean_string(resolved_place_results.get("title")) or clean_inline(row.get("NAME"))

        missing_social_fields = [
            field
            for field in ("Instagram", "Facebook", "X (Twitter)", "TikTok", "YouTube")
            if not resolved_current_fields.get(field)
        ]
        if missing_social_fields:
            socials = search_missing_socials(
                serp=serp_client,
                row=row,
                place_name=place_name,
                current_fields=resolved_current_fields,
            )
            cache_units += 1
        else:
            socials = {}

        if should_include_images:
            place_photo_seed = search_place_photo_seed(
                serp=serp_client,
                place_id=normalized_place_id,
                country_iso2=country_iso2,
            )
            cache_units += 1
            resolved_place_results = merge_place_photo_seed(resolved_place_results, place_photo_seed)

            if clean_string(resolved_place_results.get("data_id")):
                place_photo_results = search_place_photos(
                    serp=serp_client,
                    place_results=resolved_place_results,
                )
                cache_units += 1

            # Only spend the extra image search when place-photo candidates do
            # not already cover the main image plus enough gallery depth.
            place_photo_candidates = accepted_place_photo_candidates(
                resolved_place_results,
                place_photo_results,
                image_inspector=image_inspector,
            )
            if len(place_photo_candidates) < 3:
                place_name = clean_string(resolved_place_results.get("title")) or clean_inline(row.get("NAME"))
                location = build_search_location(row)
                image_results = search_images(
                    serp=serp_client,
                    query=build_unified_image_query(row, place_name),
                    location=location,
                    country_iso2=country_iso2,
                )
                cache_units += 1

        bundle = {
            "place_photo_seed": place_photo_seed,
            "place_photos": place_photo_results,
            "image_results": image_results,
            "concert_images": image_results,
            "exterior_images": image_results,
            "socials": socials,
            "images_requested": should_include_images,
            "cache_units": cache_units,
        }
        with place_serp_bundle_lock:
            place_serp_bundle_cache[normalized_place_id] = copy.deepcopy(bundle)
        return bundle
    finally:
        with place_serp_bundle_lock:
            inflight = place_serp_bundle_inflight.pop(normalized_place_id, None)
            if inflight is not None:
                inflight.set()


def create_place_enrichment(
    google_places: GooglePlacesClient,
    serp: SerpApiClient,
    image_inspector: ImageInspectionClient | None,
    row: dict[str, Any],
    skip_images: bool,
    gallery_mode: str,
) -> dict[str, Any]:
    place_id = valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
    if not place_id:
        return {
            "place_id": "",
            "place_results": {},
            "fields": {},
            "description_context": {},
            "image_status": "skipped_no_google_place_id",
        }

    country_iso2 = clean_string(row.get("COUNTRY_ISO2"))
    place_results = google_places.lookup_place(place_id, country_iso2=country_iso2)

    fields: dict[str, str] = {}
    merge_single_value(fields, "Name", clean_string(place_results.get("title")))
    merge_single_value(fields, "Address", clean_string(place_results.get("address")))
    coordinates = place_results.get("gps_coordinates") or {}
    if coordinates:
        merge_single_value(fields, "Latitude", clean_string(coordinates.get("latitude")))
        merge_single_value(fields, "Longitude", clean_string(coordinates.get("longitude")))
    merge_single_value(fields, "Phone", clean_string(place_results.get("phone")))
    merge_single_value(fields, "Hours of Operation", format_hours(place_results.get("hours")))
    merge_single_value(fields, "PlacesID", place_id)

    for field, value in extract_place_urls(place_results).items():
        merge_single_value(fields, field, value)

    serp_bundle = get_place_serp_bundle(
        place_id=place_id,
        row=row,
        serp_client=serp,
        place_results=place_results,
        current_fields=fields,
        image_inspector=image_inspector,
        skip_images=skip_images,
        gallery_mode=gallery_mode,
    )

    if not skip_images:
        place_results = merge_place_photo_seed(
            place_results,
            serp_bundle.get("place_photo_seed") or {},
        )

    social_links = serp_bundle["socials"]
    for field, value in social_links.items():
        merge_single_value(fields, field, value)

    image_status = "not_requested"

    if not skip_images:
        place_photo_results = serp_bundle["place_photos"]
        image_results = serp_bundle.get("image_results")
        if not isinstance(image_results, list):
            image_results = list(serp_bundle.get("concert_images") or []) + list(
                serp_bundle.get("exterior_images") or []
            )
        image_fields = select_image_fields(
            place_photo_results=place_photo_results,
            concert_results=image_results,
            exterior_results=[],
            place_results=place_results,
            image_inspector=image_inspector,
            gallery_mode=gallery_mode,
        )

        merge_single_value(fields, "Main Image URL", image_fields["Main Image URL"])
        merge_single_value(fields, "Additional Image URL(s)", image_fields["Additional Image URL(s)"])
        image_status = image_fields["image_status"]

    return {
        "place_id": place_id,
        "place_results": place_results,
        "fields": fields,
        "description_context": build_place_context(row, place_results),
        "image_status": image_status,
    }


def build_base_field_values(row: dict[str, Any]) -> dict[str, str]:
    fields = {column: "" for column in MAPOTIC_COLUMNS}
    for field, sources in BASE_COLUMN_SOURCES.items():
        fields[field] = get_first_value(row, sources)

    if is_replaceable_website(fields["Website"]):
        fields["Website"] = ""

    same_as_urls = parse_same_as_json(row.get("SAME_AS_JSON"))
    for field, value in same_as_urls.items():
        merge_single_value(fields, field, value)

    google_place_id = valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
    if google_place_id and not fields["PlacesID"]:
        fields["PlacesID"] = google_place_id
    if not fields["PlacesID"]:
        fields["PlacesID"] = clean_string(row.get("SOURCE_IDENTIFIER"))

    return fields


def build_description_inputs(
    unique_places: dict[str, dict[str, Any]],
    openai_client: OpenAIClient,
) -> tuple[dict[str, str], list[dict[str, Any]]]:
    cached: dict[str, str] = {}
    missing: list[dict[str, Any]] = []
    for place_id, place_payload in unique_places.items():
        cached_description = openai_client.load_cached_description(place_id)
        if cached_description:
            cached[place_id] = cached_description
            continue
        context = place_payload.get("description_context") or {}
        if not context:
            continue
        missing.append(context)
    return cached, missing


def generate_descriptions(
    openai_client: OpenAIClient,
    unique_places: dict[str, dict[str, Any]],
    batch_size: int,
    heartbeat_seconds: int,
    console: Console | None = None,
) -> dict[str, str]:
    descriptions, pending = build_description_inputs(unique_places, openai_client)
    if not pending:
        return descriptions

    progress = Progress(heartbeat_seconds, console=console)
    for start in range(0, len(pending), batch_size):
        batch = pending[start : start + batch_size]
        progress.maybe_log(
            f"OpenAI descriptions {start + 1}-{start + len(batch)} of {len(pending)}",
            force=start == 0,
        )
        generated = openai_client.describe_batch(batch)

        invalid: list[dict[str, Any]] = []
        for item in batch:
            place_id = item["place_id"]
            description = clean_string(generated.get(place_id))
            issues = validate_description(description)
            if issues:
                invalid.append(item)
                continue
            descriptions[place_id] = description
            openai_client.save_cached_description(place_id, description)

        for item in invalid:
            repaired = openai_client.describe_batch([item]).get(item["place_id"], "")
            repaired = clean_string(repaired)
            issues = validate_description(repaired)
            if not issues:
                descriptions[item["place_id"]] = repaired
                openai_client.save_cached_description(item["place_id"], repaired)

    progress.maybe_log("OpenAI descriptions complete.", force=True)
    return descriptions


def apply_output_columns(headers: list[str], worksheet: Any) -> dict[str, int]:
    existing_positions = {
        clean_string(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
    }
    template_cell = worksheet.cell(1, 1)

    for column_name in MAPOTIC_COLUMNS:
        if column_name in existing_positions:
            continue
        next_column = worksheet.max_column + 1
        cell = worksheet.cell(1, next_column)
        cell.value = column_name
        cell._style = copy.copy(template_cell._style)
        cell.number_format = template_cell.number_format
        cell.alignment = copy.copy(template_cell.alignment)
        cell.font = copy.copy(template_cell.font)
        cell.fill = copy.copy(template_cell.fill)
        cell.border = copy.copy(template_cell.border)
        cell.protection = copy.copy(template_cell.protection)
        existing_positions[column_name] = next_column

    return existing_positions


def build_final_row_values(
    row: dict[str, Any],
    place_payload: dict[str, Any],
    descriptions: dict[str, str],
) -> dict[str, str]:
    final_values = build_base_field_values(row)
    image_fields = {"Main Image URL", "Additional Image URL(s)"}
    for field, value in (place_payload.get("fields") or {}).items():
        if field in image_fields and clean_string(value):
            final_values[field] = value
        elif field in final_values and not final_values[field]:
            final_values[field] = value

    place_id = valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
    if place_id and descriptions.get(place_id) and not final_values["Description"]:
        final_values["Description"] = descriptions[place_id]

    return final_values


def write_final_values_to_worksheet(
    worksheet: Any,
    column_positions: dict[str, int],
    rownum: int,
    final_values: dict[str, str],
) -> None:
    for field, value in final_values.items():
        column = column_positions[field]
        cell = worksheet.cell(rownum, column)
        if field in {"Main Image URL", "Additional Image URL(s)"} and clean_string(value):
            cell.value = value
        elif is_blank(cell.value) and clean_string(value):
            cell.value = value


def summarize_field_counts(row_results: list[dict[str, str]]) -> dict[str, int]:
    counts = {column: 0 for column in MAPOTIC_COLUMNS}
    for row in row_results:
        for column in MAPOTIC_COLUMNS:
            if clean_string(row.get(column)):
                counts[column] += 1
    return counts


def build_failed_place_payload(row: dict[str, Any], error_message: str) -> dict[str, Any]:
    return {
        "place_id": valid_google_place_id(row.get("GOOGLE_PLACE_ID")),
        "place_results": {},
        "fields": {},
        "description_context": {},
        "image_status": "row_error",
        "error": clean_string(error_message),
    }


def enrich_row_worker(
    google_places: GooglePlacesClient,
    serp: SerpApiClient,
    image_inspector: ImageInspectionClient | None,
    row: dict[str, Any],
    skip_images: bool,
    gallery_mode: str,
) -> dict[str, Any]:
    place_id = valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
    if not place_id:
        return {
            "rownum": int(row["_rownum"]),
            "place_id": "",
            "place_payload": {
                "place_id": "",
                "place_results": {},
                "fields": {},
                "description_context": {},
                "image_status": "skipped_no_google_place_id",
            },
            "error": "",
        }

    try:
        payload = create_place_enrichment(
            google_places=google_places,
            serp=serp,
            image_inspector=image_inspector,
            row=row,
            skip_images=skip_images,
            gallery_mode=gallery_mode,
        )
        return {
            "rownum": int(row["_rownum"]),
            "place_id": place_id,
            "place_payload": payload,
            "error": "",
        }
    except Exception as exc:  # noqa: BLE001
        error_message = f"{type(exc).__name__}: {exc}"
        return {
            "rownum": int(row["_rownum"]),
            "place_id": place_id,
            "place_payload": build_failed_place_payload(row, error_message),
            "error": error_message,
        }


def enrich(args: argparse.Namespace) -> dict[str, Any]:
    row_indices = parse_row_indices(args.row_indices)
    headers, rows = build_row_contexts(
        input_xlsx=args.input_xlsx,
        limit=args.limit,
        row_indices=row_indices,
    )
    if not rows:
        raise RuntimeError("No rows selected from input workbook.")

    workers = max(1, int(getattr(args, "workers", DEFAULT_WORKERS)))
    show_cost = bool(getattr(args, "show_cost", True))
    valid_rows = sum(1 for row in rows if valid_google_place_id(row.get("GOOGLE_PLACE_ID")))
    selected_unique_place_ids = len(
        {
            valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
            for row in rows
            if valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
        }
    )
    checkpoint_json = default_checkpoint_path(args.output_xlsx)
    cost_tracker = CostTracker()
    cost_console = Console(stderr=True) if show_cost and sys.stderr.isatty() else None
    progress = Progress(args.heartbeat_seconds, console=cost_console)

    google_places_client = GooglePlacesClient(
        api_key=get_secret("GOOGLE_PLACES_API_KEY"),
        cache_dir=args.cache_dir,
        request_pause=args.request_pause,
        cost_tracker=cost_tracker,
    )
    serp_client = SerpApiClient(
        api_key=get_secret("SERPAPI_API_KEY"),
        cache_dir=args.cache_dir,
        request_pause=args.request_pause,
        cost_tracker=cost_tracker,
    )
    image_inspector = ImageInspectionClient(
        cache_dir=args.cache_dir,
        request_pause=args.request_pause,
    )
    openai_client = OpenAIClient(
        api_key=get_secret("OPENAI_API_KEY"),
        cache_dir=args.cache_dir,
        model=args.openai_model,
        request_pause=args.request_pause,
        cost_tracker=cost_tracker,
    )
    resume_requested = checkpoint_json.exists() and args.output_xlsx.exists()
    if checkpoint_json.exists() and not args.output_xlsx.exists():
        progress.maybe_log(
            f"Checkpoint {checkpoint_json.name} found without {args.output_xlsx.name}; starting fresh for safety.",
            force=True,
        )
    workbook_source = args.output_xlsx if resume_requested else args.input_xlsx
    workbook = load_workbook(workbook_source)
    worksheet = workbook.active
    column_positions = apply_output_columns(headers, worksheet)
    ensure_parent(args.output_xlsx)
    workbook.save(args.output_xlsx)

    processed_rownums: set[int] = set()
    if resume_requested:
        processed_rownums = verify_checkpoint_rows(
            rows=rows,
            processed_rownums=load_checkpoint(checkpoint_json),
            worksheet=worksheet,
            column_positions=column_positions,
            progress=progress,
        )
        if processed_rownums:
            progress.maybe_log(
                f"Resuming from checkpoint: {len(processed_rownums)} rows already processed",
                force=True,
            )

    skipped_rows = len(processed_rownums)
    rows_to_process = [row for row in rows if int(row["_rownum"]) not in processed_rownums]
    resumed_rows = len(rows_to_process) if skipped_rows else 0

    unique_places: dict[str, dict[str, Any]] = {}
    descriptions: dict[str, str] = {}
    row_results_by_rownum: dict[int, dict[str, str]] = {
        int(row["_rownum"]): read_existing_output_row_values(
            worksheet=worksheet,
            column_positions=column_positions,
            rownum=int(row["_rownum"]),
        )
        for row in rows
        if int(row["_rownum"]) in processed_rownums
    }
    row_errors: list[dict[str, Any]] = []
    processed_rows_this_run = 0

    def handle_row_result(result: dict[str, Any]) -> None:
        nonlocal processed_rows_this_run
        place_id = clean_string(result.get("place_id"))
        if place_id and (
            place_id not in unique_places
            or (
                clean_string(unique_places[place_id].get("error"))
                and not clean_string(result["place_payload"].get("error"))
            )
        ):
            unique_places[place_id] = result["place_payload"]
        if clean_string(result.get("error")):
            row_errors.append(
                {
                    "rownum": int(result["rownum"]),
                    "place_id": place_id,
                    "error": clean_string(result["error"]),
                }
            )
        processed_rows_this_run += 1
        completed_rows = len(processed_rownums) + processed_rows_this_run
        progress.maybe_log(
            f"Google place + Serp image enrichment {completed_rows}/{len(rows)} rows, {len(unique_places)} unique place ids",
            force=completed_rows == max(skipped_rows + 1, 1),
        )

    live_context: Any = nullcontext(None)
    if cost_console is not None:
        live_context = Live(
            CostTableView(cost_tracker),
            console=cost_console,
            refresh_per_second=4,
            transient=False,
        )

    with live_context:
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            for start in range(0, len(rows_to_process), CHECKPOINT_SAVE_EVERY_ROWS):
                batch_rows = rows_to_process[start : start + CHECKPOINT_SAVE_EVERY_ROWS]
                if not batch_rows:
                    continue

                batch_results: dict[int, dict[str, Any]] = {}
                if workers == 1:
                    for row in batch_rows:
                        result = enrich_row_worker(
                            google_places=google_places_client,
                            serp=serp_client,
                            image_inspector=image_inspector,
                            row=row,
                            skip_images=args.skip_images,
                            gallery_mode=args.gallery_mode,
                        )
                        batch_results[int(row["_rownum"])] = result
                        handle_row_result(result)
                else:
                    future_to_rownum = {
                        executor.submit(
                            enrich_row_worker,
                            google_places_client,
                            serp_client,
                            image_inspector,
                            row,
                            args.skip_images,
                            args.gallery_mode,
                        ): int(row["_rownum"])
                        for row in batch_rows
                    }
                    for future in concurrent.futures.as_completed(future_to_rownum):
                        result = future.result()
                        batch_results[int(result["rownum"])] = result
                        handle_row_result(result)

                if not args.skip_openai:
                    pending_place_ids = dedupe_preserve_order(
                        [
                            clean_string(batch_results[int(row["_rownum"])]["place_id"])
                            for row in batch_rows
                            if clean_string(batch_results[int(row["_rownum"])]["place_id"])
                            and clean_string(batch_results[int(row["_rownum"])]["place_id"]) not in descriptions
                        ]
                    )
                    pending_places = {
                        place_id: unique_places[place_id]
                        for place_id in pending_place_ids
                        if place_id in unique_places and place_id not in descriptions
                    }
                    if pending_places:
                        descriptions.update(
                            generate_descriptions(
                                openai_client=openai_client,
                                unique_places=pending_places,
                                batch_size=args.description_batch_size,
                                heartbeat_seconds=args.heartbeat_seconds,
                                console=cost_console,
                            )
                        )

                for row in batch_rows:
                    result = batch_results[int(row["_rownum"])]
                    final_values = build_final_row_values(
                        row=row,
                        place_payload=result["place_payload"],
                        descriptions=descriptions,
                    )
                    row_results_by_rownum[int(row["_rownum"])] = final_values
                    write_final_values_to_worksheet(
                        worksheet=worksheet,
                        column_positions=column_positions,
                        rownum=int(row["_rownum"]),
                        final_values=final_values,
                    )

                workbook.save(args.output_xlsx)
                processed_rownums.update(int(row["_rownum"]) for row in batch_rows)
                save_checkpoint(checkpoint_json, processed_rownums)
                progress.maybe_log(
                    f"Checkpoint saved {len(processed_rownums)}/{len(rows)} rows to {args.output_xlsx.name}",
                    force=True,
                )

    row_results = [
        row_results_by_rownum.get(int(row["_rownum"]))
        or build_final_row_values(row=row, place_payload={}, descriptions=descriptions)
        for row in rows
    ]
    field_counts = summarize_field_counts(row_results)
    image_status_counts: dict[str, int] = {}
    for payload in unique_places.values():
        status = clean_string(payload.get("image_status")) or "unknown"
        image_status_counts[status] = image_status_counts.get(status, 0) + 1

    generated_at = dt.datetime.now(dt.timezone.utc)
    cost_snapshot = cost_tracker.snapshot()
    cost_log_json = cost_log_path(generated_at)
    ensure_parent(cost_log_json)
    cost_log_json.write_text(
        json.dumps(
            {
                "total_cost": cost_snapshot["total"],
                "by_service": cost_snapshot["by_service"],
                "total_rows_processed": len(rows_to_process),
                "skipped_rows": skipped_rows,
                "timestamp": generated_at.isoformat(),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    summary = {
        "generated_at_utc": generated_at.isoformat(),
        "script_version": SCRIPT_VERSION,
        "input_xlsx": str(args.input_xlsx.resolve()),
        "output_xlsx": str(args.output_xlsx.resolve()),
        "cache_dir": str(args.cache_dir.resolve()),
        "checkpoint_json": str(checkpoint_json.resolve()),
        "cost_log_json": str(cost_log_json.resolve()),
        "rows_selected": len(rows),
        "processed_rows_this_run": len(rows_to_process),
        "skipped_rows": skipped_rows,
        "resumed_rows": resumed_rows,
        "rows_with_valid_google_place_id": valid_rows,
        "rows_without_valid_google_place_id": len(rows) - valid_rows,
        "unique_google_place_ids": selected_unique_place_ids,
        "unique_google_place_ids_processed_this_run": len(unique_places),
        "workers": workers,
        "show_cost": show_cost,
        "skip_images": bool(args.skip_images),
        "skip_openai": bool(args.skip_openai),
        "openai_model": args.openai_model,
        "field_population_counts": field_counts,
        "rows_with_description": field_counts["Description"],
        "rows_with_main_image": field_counts["Main Image URL"],
        "rows_with_additional_images": field_counts["Additional Image URL(s)"],
        "image_status_counts": image_status_counts,
        "row_error_count": len(row_errors),
        "row_errors": row_errors,
        "costs": cost_snapshot,
        "google_places": {
            "network_calls": google_places_client.network_calls,
            "cache_hits": google_places_client.cache_hits,
        },
        "serpapi": {
            "network_calls": serp_client.network_calls,
            "cache_hits": serp_client.cache_hits,
        },
        "image_inspection": {
            "network_calls": image_inspector.network_calls,
            "cache_hits": image_inspector.cache_hits,
        },
        "openai": {
            "network_calls": openai_client.network_calls,
            "cache_hits": openai_client.cache_hits,
            "descriptions_cached_or_generated": len(descriptions),
        },
    }
    ensure_parent(args.summary_json)
    args.summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Use Google Places for place-id metadata, SerpApi for image/social enrichment, "
            "and OpenAI for descriptions in a Mapotic-ready workbook."
        )
    )
    parser.add_argument(
        "--input-xlsx",
        type=Path,
        default=DEFAULT_INPUT_XLSX,
        help="Reviewed workbook containing GOOGLE_PLACE_ID values.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=DEFAULT_OUTPUT_XLSX,
        help="Destination workbook path.",
    )
    parser.add_argument(
        "--summary-json",
        type=Path,
        default=DEFAULT_SUMMARY_JSON,
        help="Summary JSON output path.",
    )
    parser.add_argument(
        "--cache-dir",
        type=Path,
        default=DEFAULT_CACHE_DIR,
        help="Cache directory for SerpApi and OpenAI responses.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Only process the first N selected data rows.",
    )
    parser.add_argument(
        "--row-indices",
        type=str,
        default="",
        help="Comma-separated Excel row numbers to process.",
    )
    parser.add_argument(
        "--request-pause",
        type=float,
        default=DEFAULT_REQUEST_PAUSE,
        help="Pause between network requests.",
    )
    parser.add_argument(
        "--heartbeat-seconds",
        type=int,
        default=DEFAULT_HEARTBEAT_SECONDS,
        help="Seconds between progress logs.",
    )
    parser.add_argument(
        "--description-batch-size",
        type=int,
        default=DEFAULT_DESCRIPTION_BATCH_SIZE,
        help="Number of venues per OpenAI description batch.",
    )
    parser.add_argument(
        "--openai-model",
        type=str,
        default=DEFAULT_OPENAI_MODEL,
        help="OpenAI model used for description generation.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=DEFAULT_WORKERS,
        help="Number of worker threads used for row enrichment.",
    )
    parser.add_argument(
        "--show-cost",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Show a live estimated cost table when attached to an interactive terminal.",
    )
    parser.add_argument(
        "--gallery-mode",
        choices=["mirror-main", "separate-main"],
        default="mirror-main",
        help="Whether Additional Image URL(s) should repeat the main image first or exclude it.",
    )
    parser.add_argument(
        "--skip-images",
        action="store_true",
        help="Skip SerpApi image lookups.",
    )
    parser.add_argument(
        "--skip-openai",
        action="store_true",
        help="Skip OpenAI description generation.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        summary = enrich(args)
    except KeyboardInterrupt:
        print("Enrichment interrupted.", file=sys.stderr)
        return 130
    except Exception as exc:  # noqa: BLE001
        print(f"Enrichment failed: {exc}", file=sys.stderr)
        return 1

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
