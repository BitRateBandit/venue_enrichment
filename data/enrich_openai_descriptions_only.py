#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import datetime as dt
import json
import re
import sys
from contextlib import nullcontext
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rich.console import Console
from rich.live import Live
from rich.table import Table


THIS_DIR = Path(__file__).resolve().parent
if str(THIS_DIR) not in sys.path:
    sys.path.insert(0, str(THIS_DIR))

import enrich_mapotic_places as shared


SCRIPT_VERSION = "2026-04-17.1"
DEFAULT_INPUT_XLSX = Path("data/input/Jambase_Venues_Global_input.xlsx")
DEFAULT_OUTPUT_XLSX = Path("data/output/Jambase_Venues_Global_input_openai_descriptions.xlsx")
DEFAULT_SUMMARY_JSON = Path("data/output/Jambase_Venues_Global_input_openai_descriptions_summary.json")
DEFAULT_CACHE_DIR = Path("data/cache/openai_descriptions_only")
CHECKPOINT_SAVE_EVERY_ROWS = shared.CHECKPOINT_SAVE_EVERY_ROWS
DESCRIPTION_COLUMN = "Description"


def select_worksheet(workbook: Any, sheet_name: str = "") -> Any:
    if not sheet_name:
        return workbook.active
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise RuntimeError(f"Worksheet '{sheet_name}' not found. Available sheets: {available}")
    return workbook[sheet_name]


def build_row_contexts(
    input_xlsx: Path,
    *,
    sheet: str = "",
    limit: int | None = None,
    row_indices: set[int] | None = None,
) -> tuple[str, list[str], list[dict[str, Any]]]:
    workbook = load_workbook(input_xlsx, read_only=True)
    worksheet = select_worksheet(workbook, sheet)
    rows = worksheet.iter_rows(values_only=True)
    headers = [shared.clean_string(value) for value in next(rows)]
    contexts: list[dict[str, Any]] = []
    selected = 0

    for rownum, row in enumerate(rows, start=2):
        if row_indices and rownum not in row_indices:
            continue
        record = {headers[index]: row[index] for index in range(len(headers))}
        record["_rownum"] = rownum
        contexts.append(record)
        selected += 1
        if limit is not None and selected >= limit:
            break

    return worksheet.title, headers, contexts


def apply_description_output_column(worksheet: Any) -> int:
    existing_positions = {
        shared.clean_string(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
    }
    if DESCRIPTION_COLUMN in existing_positions:
        return existing_positions[DESCRIPTION_COLUMN]

    template_cell = worksheet.cell(1, 1)
    next_column = worksheet.max_column + 1
    cell = worksheet.cell(1, next_column)
    cell.value = DESCRIPTION_COLUMN
    cell._style = copy.copy(template_cell._style)
    cell.number_format = template_cell.number_format
    cell.alignment = copy.copy(template_cell.alignment)
    cell.font = copy.copy(template_cell.font)
    cell.fill = copy.copy(template_cell.fill)
    cell.border = copy.copy(template_cell.border)
    cell.protection = copy.copy(template_cell.protection)
    return next_column


def description_is_complete(value: Any) -> bool:
    text = shared.clean_string(value)
    return bool(text) and not shared.validate_description(text)


def read_existing_description(worksheet: Any, description_column: int, rownum: int) -> str:
    return shared.clean_string(worksheet.cell(rownum, description_column).value)


def verify_checkpoint_rows(
    rows: list[dict[str, Any]],
    processed_rownums: set[int],
    worksheet: Any,
    description_column: int,
    progress: shared.Progress | None = None,
) -> set[int]:
    verified: set[int] = set()
    for row in rows:
        rownum = int(row["_rownum"])
        if rownum not in processed_rownums:
            continue
        if description_is_complete(read_existing_description(worksheet, description_column, rownum)):
            verified.add(rownum)
            continue
        if progress is not None:
            progress.maybe_log(
                f"Checkpoint row {rownum} missing a valid description; reprocessing for safety.",
                force=True,
            )
    return verified


def row_place_id(row: dict[str, Any]) -> str:
    return (
        shared.valid_google_place_id(row.get("GOOGLE_PLACE_ID"))
        or shared.clean_string(row.get("PlacesID"))
        or shared.clean_string(row.get("SOURCE_IDENTIFIER"))
        or shared.clean_string(row.get("SOURCE_ID"))
        or f"row-{int(row['_rownum'])}"
    )


def parse_google_types(raw: Any) -> list[str]:
    text = shared.clean_string(raw)
    if not text:
        return []
    if text.startswith("["):
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            payload = None
        if isinstance(payload, list):
            return shared.dedupe_preserve_order(
                [shared.clean_inline(item) for item in payload if shared.clean_inline(item)]
            )
    return shared.dedupe_preserve_order(
        [shared.clean_inline(part) for part in re.split(r"[|,]", text) if shared.clean_inline(part)]
    )


def build_row_place_results(row: dict[str, Any]) -> dict[str, Any]:
    type_name = shared.clean_inline(row.get("GOOGLE_PRIMARY_TYPE") or row.get("RECORD_TYPE"))
    google_types = parse_google_types(row.get("GOOGLE_TYPES"))
    return {
        "title": shared.get_first_value(row, ["Name", "NAME"]),
        "description": shared.get_first_value(row, ["Description"]),
        "type": type_name,
        "type_id": type_name,
        "open_state": shared.clean_inline(row.get("GOOGLE_BUSINESS_STATUS")),
        "address": shared.get_first_value(row, ["Address", "ADDRESS"]),
        "extensions": [{"types": google_types}] if google_types else [],
        "events": [],
    }


def description_context_has_content(context: dict[str, Any]) -> bool:
    scalar_fields = (
        "name",
        "city",
        "state",
        "country",
        "address",
        "place_type",
        "source_description",
        "open_state",
    )
    if any(shared.clean_string(context.get(field)) for field in scalar_fields):
        return True
    return bool(context.get("highlights") or context.get("event_titles"))


def build_description_context(row: dict[str, Any]) -> dict[str, Any]:
    place_id = row_place_id(row)
    if not place_id:
        return {}
    context = shared.build_place_context(row, build_row_place_results(row))
    context["place_id"] = place_id
    context["style_hint"] = shared.style_hint_for(place_id)
    if not description_context_has_content(context):
        return {}
    return context


def write_description_to_worksheet(
    worksheet: Any,
    description_column: int,
    rownum: int,
    description: str,
) -> None:
    worksheet.cell(rownum, description_column).value = description


def build_openai_cost_table(snapshot: dict[str, Any]) -> Table:
    table = Table(title="Estimated OpenAI Cost", expand=False)
    table.add_column("Service")
    table.add_column("Calls", justify="right")
    table.add_column("Cost ($)", justify="right")

    openai_snapshot = (snapshot.get("by_service") or {}).get("openai") or {}
    table.add_row(
        "OpenAI",
        str(openai_snapshot.get("calls", 0)),
        f"{float(openai_snapshot.get('cost', 0.0)):.2f}",
    )
    table.add_row("TOTAL", "-", f"{float(snapshot.get('total', 0.0)):.2f}")
    return table


class OpenAICostTableView:
    def __init__(self, tracker: shared.CostTracker) -> None:
        self.tracker = tracker

    def __rich__(self) -> Table:
        return build_openai_cost_table(self.tracker.snapshot())


def enrich(args: argparse.Namespace) -> dict[str, Any]:
    row_indices = shared.parse_row_indices(getattr(args, "row_indices", ""))
    checkpoint_json = shared.default_checkpoint_path(args.output_xlsx)
    show_cost = bool(getattr(args, "show_cost", True))
    workers = max(1, int(getattr(args, "workers", 1)))
    output_exists = args.output_xlsx.exists()
    resume_from_checkpoint = checkpoint_json.exists() and output_exists
    workbook_source = args.output_xlsx if output_exists else args.input_xlsx

    sheet_name, headers, rows = build_row_contexts(
        workbook_source,
        sheet=getattr(args, "sheet", ""),
        limit=getattr(args, "limit", None),
        row_indices=row_indices,
    )
    if not rows:
        raise RuntimeError("No rows selected from input workbook.")

    progress_console = Console(stderr=True) if show_cost and sys.stderr.isatty() else None
    progress = shared.Progress(args.heartbeat_seconds, console=progress_console)
    if checkpoint_json.exists() and not output_exists:
        progress.maybe_log(
            f"Checkpoint {checkpoint_json.name} found without {args.output_xlsx.name}; starting fresh for safety.",
            force=True,
        )
    elif output_exists and not checkpoint_json.exists() and args.output_xlsx != args.input_xlsx:
        progress.maybe_log(
            f"Output workbook {args.output_xlsx.name} already exists without a checkpoint; "
            "using existing descriptions to avoid duplicate work.",
            force=True,
        )

    workbook = load_workbook(workbook_source)
    worksheet = select_worksheet(workbook, getattr(args, "sheet", ""))
    description_column = apply_description_output_column(worksheet)
    shared.ensure_parent(args.output_xlsx)
    workbook.save(args.output_xlsx)

    checkpoint_rownums: set[int] = set()
    if resume_from_checkpoint:
        checkpoint_rownums = verify_checkpoint_rows(
            rows=rows,
            processed_rownums=shared.load_checkpoint(checkpoint_json),
            worksheet=worksheet,
            description_column=description_column,
            progress=progress,
        )
        if checkpoint_rownums:
            progress.maybe_log(
                f"Resuming from checkpoint: {len(checkpoint_rownums)} rows already processed",
                force=True,
            )

    rows_with_valid_description = 0
    rows_missing_context = 0
    unique_places: dict[str, dict[str, Any]] = {}
    rows_by_place_id: dict[str, list[dict[str, Any]]] = {}

    for row in rows:
        if description_is_complete(row.get(DESCRIPTION_COLUMN)):
            rows_with_valid_description += 1
            continue
        context = build_description_context(row)
        if not context:
            rows_missing_context += 1
            continue
        place_id = context["place_id"]
        unique_places.setdefault(place_id, {"description_context": context})
        rows_by_place_id.setdefault(place_id, []).append(row)

    rows_eligible = sum(len(item) for item in rows_by_place_id.values())
    rows_skipped = len(rows) - rows_eligible
    cost_tracker = shared.CostTracker(price_map={"openai": shared.SERVICE_COSTS["openai"]})
    openai_client = shared.OpenAIClient(
        api_key=shared.get_secret("OPENAI_API_KEY"),
        cache_dir=args.cache_dir,
        model=args.openai_model,
        request_pause=args.request_pause,
        cost_tracker=cost_tracker,
    )
    row_errors: list[dict[str, Any]] = []
    completed_rownums: set[int] = set()
    dirty_completed_rownums = 0

    def record_row_failure(place_id: str, error_message: str) -> None:
        for row in rows_by_place_id.get(place_id, []):
            row_errors.append(
                {
                    "rownum": int(row["_rownum"]),
                    "place_id": place_id,
                    "error": error_message,
                }
            )

    def persist_progress(*, force: bool = False) -> None:
        nonlocal dirty_completed_rownums
        if not checkpoint_rownums:
            return
        if dirty_completed_rownums == 0 and checkpoint_json.exists():
            return
        if not force and dirty_completed_rownums < CHECKPOINT_SAVE_EVERY_ROWS:
            return
        workbook.save(args.output_xlsx)
        shared.save_checkpoint(checkpoint_json, checkpoint_rownums)
        progress.maybe_log(
            f"Checkpoint saved {len(checkpoint_rownums)}/{len(rows)} rows to {args.output_xlsx.name}",
            force=True,
        )
        dirty_completed_rownums = 0

    def record_description_success(place_id: str, description: str) -> None:
        nonlocal dirty_completed_rownums
        for row in rows_by_place_id.get(place_id, []):
            rownum = int(row["_rownum"])
            if rownum in completed_rownums:
                continue
            write_description_to_worksheet(
                worksheet=worksheet,
                description_column=description_column,
                rownum=rownum,
                description=description,
            )
            checkpoint_rownums.add(rownum)
            completed_rownums.add(rownum)
            dirty_completed_rownums += 1

    cached_descriptions, pending_contexts = shared.build_description_inputs(unique_places, openai_client)
    for place_id, description in cached_descriptions.items():
        record_description_success(place_id, description)
    if cached_descriptions:
        progress.maybe_log(
            f"Loaded {len(cached_descriptions)} cached OpenAI descriptions.",
            force=True,
        )
        persist_progress(force=True)

    live_context: Any = nullcontext(None)
    if progress_console is not None:
        live_context = Live(
            OpenAICostTableView(cost_tracker),
            console=progress_console,
            refresh_per_second=4,
            transient=False,
        )

    with live_context:
        for start in range(0, len(pending_contexts), args.description_batch_size):
            batch = pending_contexts[start : start + args.description_batch_size]
            if not batch:
                continue
            progress.maybe_log(
                f"OpenAI descriptions {start + 1}-{start + len(batch)} of {len(pending_contexts)}",
                force=start == 0,
            )
            try:
                generated = openai_client.describe_batch(batch)
            except KeyboardInterrupt:
                persist_progress(force=True)
                raise
            except Exception as exc:  # noqa: BLE001
                error_message = f"{type(exc).__name__}: {exc}"
                for item in batch:
                    record_row_failure(item["place_id"], error_message)
                progress.maybe_log(
                    f"OpenAI batch failed for {len(batch)} place ids: {error_message}",
                    force=True,
                )
                continue

            invalid: list[dict[str, Any]] = []
            for item in batch:
                place_id = item["place_id"]
                description = shared.clean_string(generated.get(place_id))
                issues = shared.validate_description(description)
                if issues:
                    invalid.append(item)
                    continue
                openai_client.save_cached_description(place_id, description)
                record_description_success(place_id, description)

            for item in invalid:
                place_id = item["place_id"]
                try:
                    repaired = openai_client.describe_batch([item]).get(place_id, "")
                except KeyboardInterrupt:
                    persist_progress(force=True)
                    raise
                except Exception as exc:  # noqa: BLE001
                    record_row_failure(place_id, f"{type(exc).__name__}: {exc}")
                    continue

                repaired = shared.clean_string(repaired)
                issues = shared.validate_description(repaired)
                if issues:
                    record_row_failure(place_id, f"validation_failed:{','.join(issues)}")
                    continue
                openai_client.save_cached_description(place_id, repaired)
                record_description_success(place_id, repaired)

            persist_progress(force=True)

    persist_progress(force=True)
    progress.maybe_log("OpenAI descriptions complete.", force=True)

    generated_at = dt.datetime.now(dt.timezone.utc)
    cost_snapshot = cost_tracker.snapshot()
    summary = {
        "generated_at_utc": generated_at.isoformat(),
        "script_version": SCRIPT_VERSION,
        "input_xlsx": str(args.input_xlsx.resolve()),
        "output_xlsx": str(args.output_xlsx.resolve()),
        "summary_json": str(args.summary_json.resolve()),
        "sheet": sheet_name,
        "cache_dir": str(args.cache_dir.resolve()),
        "checkpoint_json": str(checkpoint_json.resolve()),
        "total_rows_scanned": len(rows),
        "rows_eligible": rows_eligible,
        "rows_completed": len(completed_rownums),
        "processed_rows_this_run": len(completed_rownums),
        "rows_skipped": rows_skipped,
        "rows_skipped_existing_valid_description": rows_with_valid_description,
        "rows_skipped_missing_context": rows_missing_context,
        "resumed_rows": len(checkpoint_rownums - completed_rownums),
        "failed_rows": len(row_errors),
        "row_errors": row_errors,
        "workers_requested": workers,
        "workers_used": 1,
        "show_cost": show_cost,
        "openai_model": args.openai_model,
        "description_batch_size": args.description_batch_size,
        "openai_call_count": openai_client.network_calls,
        "estimated_openai_cost": cost_snapshot["total"],
        "total_cost": cost_snapshot["total"],
        "costs": cost_snapshot,
        "openai": {
            "network_calls": openai_client.network_calls,
            "cache_hits": openai_client.cache_hits,
            "descriptions_cached_or_generated": len(cached_descriptions) + len(completed_rownums),
        },
    }
    shared.ensure_parent(args.summary_json)
    args.summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate Mapotic venue descriptions from an import workbook using OpenAI only."
    )
    parser.add_argument(
        "--input-xlsx",
        type=Path,
        default=DEFAULT_INPUT_XLSX,
        help="Source workbook containing the import sheet.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=DEFAULT_OUTPUT_XLSX,
        help="Destination workbook path.",
    )
    parser.add_argument(
        "--summary-json",
        type=Path,
        default=DEFAULT_SUMMARY_JSON,
        help="Summary JSON output path.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default="",
        help="Workbook sheet name to process. Defaults to the active sheet.",
    )
    parser.add_argument(
        "--cache-dir",
        type=Path,
        default=DEFAULT_CACHE_DIR,
        help="Cache directory for OpenAI responses.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Only process the first N selected data rows.",
    )
    parser.add_argument(
        "--row-indices",
        type=str,
        default="",
        help="Comma-separated Excel row numbers to process.",
    )
    parser.add_argument(
        "--request-pause",
        type=float,
        default=shared.DEFAULT_REQUEST_PAUSE,
        help="Pause between network requests.",
    )
    parser.add_argument(
        "--heartbeat-seconds",
        type=int,
        default=shared.DEFAULT_HEARTBEAT_SECONDS,
        help="Seconds between progress logs.",
    )
    parser.add_argument(
        "--description-batch-size",
        type=int,
        default=shared.DEFAULT_DESCRIPTION_BATCH_SIZE,
        help="Number of venues per OpenAI description batch.",
    )
    parser.add_argument(
        "--openai-model",
        type=str,
        default=shared.DEFAULT_OPENAI_MODEL,
        help="OpenAI model used for description generation.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=1,
        help="Kept for CLI compatibility; descriptions are processed sequentially.",
    )
    parser.add_argument(
        "--show-cost",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Show a live estimated OpenAI cost table when attached to an interactive terminal.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        summary = enrich(args)
    except KeyboardInterrupt:
        print("Description enrichment interrupted.", file=sys.stderr)
        return 130
    except Exception as exc:  # noqa: BLE001
        print(f"Description enrichment failed: {exc}", file=sys.stderr)
        return 1

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
