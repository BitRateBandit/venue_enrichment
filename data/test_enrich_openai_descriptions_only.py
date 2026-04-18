#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
import tempfile
import unittest
from pathlib import Path
from typing import Any
from unittest import mock

import requests
from openpyxl import Workbook, load_workbook

THIS_DIR = Path(__file__).resolve().parent
if str(THIS_DIR) not in sys.path:
    sys.path.insert(0, str(THIS_DIR))

import enrich_openai_descriptions_only as module


def make_valid_description(name: str) -> str:
    return (
        f"{name} hums with that pre-show tension you feel in the floorboards before the first "
        "chord lands. Warm light, worn surfaces, and close crowd energy give the room a lived-in "
        "character that reads as real rather than staged. The setting feels tuned for a night "
        "that can move from quiet anticipation to full-throated release without losing its sense "
        "of place. For route-planning context and playlists that match the mood, musicroadtrip.com "
        "makes a strong companion stop."
    )


def build_success_response(items: list[dict[str, object]]) -> mock.Mock:
    response = mock.Mock()
    response.status_code = 200
    response.raise_for_status.return_value = None
    response.json.return_value = {
        "output": [
            {
                "type": "message",
                "content": [
                    {
                        "type": "output_text",
                        "text": json.dumps(
                            {
                                "descriptions": [
                                    {
                                        "place_id": str(item["place_id"]),
                                        "description": make_valid_description(str(item["name"])),
                                    }
                                    for item in items
                                ]
                            }
                        ),
                    }
                ],
            }
        ]
    }
    return response


def build_args(
    *,
    input_xlsx: Path,
    output_xlsx: Path,
    summary_json: Path,
    cache_dir: Path,
    sheet: str = "",
    limit: int | None = None,
    description_batch_size: int = 2,
) -> argparse.Namespace:
    return argparse.Namespace(
        input_xlsx=input_xlsx,
        output_xlsx=output_xlsx,
        summary_json=summary_json,
        sheet=sheet,
        cache_dir=cache_dir,
        limit=limit,
        row_indices="",
        request_pause=0.0,
        heartbeat_seconds=999999,
        description_batch_size=description_batch_size,
        openai_model="gpt-4o-mini",
        workers=1,
        show_cost=False,
    )


def description_column_index(worksheet: Any) -> int:
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    return headers.index("Description") + 1


class DescriptionOnlyTests(unittest.TestCase):
    @mock.patch.object(module.shared, "get_secret", return_value="openai-secret")
    @mock.patch("requests.sessions.Session.post", autospec=True)
    def test_enrich_reads_requested_sheet_skips_valid_rows_and_writes_summary(
        self,
        mocked_post: mock.Mock,
        mocked_get_secret: mock.Mock,
    ) -> None:
        def fake_post(_session: requests.Session, _url: str, **kwargs: object) -> mock.Mock:
            payload = json.loads(kwargs["json"]["input"][1]["content"][0]["text"])
            return build_success_response(payload["venues"])

        mocked_post.side_effect = fake_post

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_json = tmp_path / "summary.json"

            workbook = Workbook()
            default_sheet = workbook.active
            default_sheet.title = "Ignore Me"
            default_sheet.append(["SOURCE_IDENTIFIER", "NAME", "ADDRESS", "CITY", "COUNTRY", "Description"])
            default_sheet.append(["ignore-1", "Ignore Venue", "1 Nowhere St", "Elsewhere", "US", ""])

            target_sheet = workbook.create_sheet("Import")
            target_sheet.append(
                ["SOURCE_IDENTIFIER", "NAME", "ADDRESS", "CITY", "COUNTRY", "GOOGLE_PRIMARY_TYPE", "Description"]
            )
            target_sheet.append(["source-1", "Venue One", "1 Main St", "Austin", "US", "music_venue", ""])
            target_sheet.append(
                [
                    "source-2",
                    "Venue Two",
                    "2 Main St",
                    "Austin",
                    "US",
                    "music_venue",
                    make_valid_description("Venue Two"),
                ]
            )
            target_sheet.append(["source-3", "Venue Three", "3 Main St", "Austin", "US", "music_venue", ""])
            workbook.save(input_xlsx)

            summary = module.enrich(
                build_args(
                    input_xlsx=input_xlsx,
                    output_xlsx=output_xlsx,
                    summary_json=summary_json,
                    cache_dir=tmp_path / "cache",
                    sheet="Import",
                    description_batch_size=2,
                )
            )

            result_workbook = load_workbook(output_xlsx)
            result_sheet = result_workbook["Import"]
            ignore_sheet = result_workbook["Ignore Me"]
            description_column = description_column_index(result_sheet)
            summary_exists = summary_json.exists()
            written_summary = json.loads(summary_json.read_text(encoding="utf-8"))

        self.assertEqual(mocked_post.call_count, 1)
        self.assertEqual(mocked_get_secret.call_count, 1)
        self.assertEqual(summary["sheet"], "Import")
        self.assertEqual(summary["total_rows_scanned"], 3)
        self.assertEqual(summary["rows_eligible"], 2)
        self.assertEqual(summary["rows_completed"], 2)
        self.assertEqual(summary["rows_skipped"], 1)
        self.assertEqual(summary["resumed_rows"], 0)
        self.assertEqual(summary["failed_rows"], 0)
        self.assertEqual(summary["openai_call_count"], 1)
        self.assertEqual(summary["estimated_openai_cost"], 0.002)
        self.assertEqual(summary["total_cost"], 0.002)
        self.assertEqual(summary["costs"]["by_service"]["openai"]["calls"], 1)
        self.assertEqual(result_sheet.cell(2, description_column).value, make_valid_description("Venue One"))
        self.assertEqual(result_sheet.cell(3, description_column).value, make_valid_description("Venue Two"))
        self.assertEqual(result_sheet.cell(4, description_column).value, make_valid_description("Venue Three"))
        self.assertEqual(ignore_sheet.max_row, 2)
        self.assertTrue(summary_exists)
        for key in (
            "total_rows_scanned",
            "rows_eligible",
            "rows_completed",
            "rows_skipped",
            "resumed_rows",
            "failed_rows",
            "openai_call_count",
            "estimated_openai_cost",
            "output_xlsx",
            "checkpoint_json",
            "generated_at_utc",
        ):
            self.assertIn(key, written_summary)

    @mock.patch.object(module.shared, "get_secret", return_value="openai-secret")
    @mock.patch("requests.sessions.Session.post", autospec=True)
    def test_enrich_resume_skips_completed_rows_without_duplicate_calls(
        self,
        mocked_post: mock.Mock,
        mocked_get_secret: mock.Mock,
    ) -> None:
        def fake_post(_session: requests.Session, _url: str, **kwargs: object) -> mock.Mock:
            payload = json.loads(kwargs["json"]["input"][1]["content"][0]["text"])
            return build_success_response(payload["venues"])

        mocked_post.side_effect = fake_post

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_one = tmp_path / "summary-one.json"
            summary_two = tmp_path / "summary-two.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["SOURCE_IDENTIFIER", "NAME", "ADDRESS", "CITY", "COUNTRY", "Description"])
            worksheet.append(["source-1", "Venue One", "1 Main St", "Austin", "US", ""])
            worksheet.append(["source-2", "Venue Two", "2 Main St", "Austin", "US", ""])
            worksheet.append(["source-3", "Venue Three", "3 Main St", "Austin", "US", ""])
            workbook.save(input_xlsx)

            base_args = dict(
                input_xlsx=input_xlsx,
                output_xlsx=output_xlsx,
                cache_dir=tmp_path / "cache",
                sheet="",
                description_batch_size=1,
            )

            first_summary = module.enrich(
                build_args(summary_json=summary_one, limit=2, **base_args)
            )
            second_summary = module.enrich(
                build_args(summary_json=summary_two, limit=None, **base_args)
            )

            result_workbook = load_workbook(output_xlsx)
            result_sheet = result_workbook.active
            description_column = description_column_index(result_sheet)

        self.assertEqual(mocked_post.call_count, 3)
        self.assertEqual(mocked_get_secret.call_count, 2)
        self.assertEqual(first_summary["rows_completed"], 2)
        self.assertEqual(first_summary["openai_call_count"], 2)
        self.assertEqual(first_summary["estimated_openai_cost"], 0.004)
        self.assertEqual(second_summary["rows_eligible"], 1)
        self.assertEqual(second_summary["rows_completed"], 1)
        self.assertEqual(second_summary["rows_skipped"], 2)
        self.assertEqual(second_summary["resumed_rows"], 2)
        self.assertEqual(second_summary["openai_call_count"], 1)
        self.assertEqual(second_summary["estimated_openai_cost"], 0.002)
        self.assertEqual(result_sheet.cell(2, description_column).value, make_valid_description("Venue One"))
        self.assertEqual(result_sheet.cell(3, description_column).value, make_valid_description("Venue Two"))
        self.assertEqual(result_sheet.cell(4, description_column).value, make_valid_description("Venue Three"))

    @mock.patch.object(module.shared, "get_secret", return_value="openai-secret")
    @mock.patch("requests.sessions.Session.post", autospec=True)
    def test_enrich_uses_cached_description_without_network(
        self,
        mocked_post: mock.Mock,
        mocked_get_secret: mock.Mock,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_json = tmp_path / "summary.json"
            cache_dir = tmp_path / "cache"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["SOURCE_IDENTIFIER", "NAME", "ADDRESS", "CITY", "COUNTRY", "Description"])
            worksheet.append(["source-1", "Venue One", "1 Main St", "Austin", "US", ""])
            workbook.save(input_xlsx)

            place_id = module.row_place_id({"SOURCE_IDENTIFIER": "source-1", "_rownum": 2})
            client = module.shared.OpenAIClient(
                api_key="openai-secret",
                cache_dir=cache_dir,
                model="gpt-4o-mini",
                request_pause=0.0,
            )
            client.save_cached_description(place_id, make_valid_description("Venue One"))

            summary = module.enrich(
                build_args(
                    input_xlsx=input_xlsx,
                    output_xlsx=output_xlsx,
                    summary_json=summary_json,
                    cache_dir=cache_dir,
                )
            )
            result_workbook = load_workbook(output_xlsx)
            result_sheet = result_workbook.active
            description_column = description_column_index(result_sheet)

        mocked_post.assert_not_called()
        mocked_get_secret.assert_called_once()
        self.assertEqual(summary["rows_completed"], 1)
        self.assertEqual(summary["openai_call_count"], 0)
        self.assertEqual(summary["estimated_openai_cost"], 0.0)
        self.assertEqual(summary["openai"]["cache_hits"], 1)
        self.assertEqual(result_sheet.cell(2, description_column).value, make_valid_description("Venue One"))

    @mock.patch.object(module.shared, "get_secret", return_value="openai-secret")
    @mock.patch.object(module.shared.time, "sleep", return_value=None)
    @mock.patch("requests.sessions.Session.post", autospec=True)
    def test_enrich_retries_after_429_and_counts_cost_once(
        self,
        mocked_post: mock.Mock,
        _mocked_sleep: mock.Mock,
        mocked_get_secret: mock.Mock,
    ) -> None:
        throttled_response = mock.Mock()
        throttled_response.status_code = 429
        throttled_response.headers = {"Retry-After": "0"}
        throttled_response.raise_for_status.side_effect = requests.HTTPError(
            "rate limited",
            response=throttled_response,
        )

        success_response = build_success_response(
            [{"place_id": "source-1", "name": "Venue One"}]
        )
        mocked_post.side_effect = [throttled_response, success_response]

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_json = tmp_path / "summary.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["SOURCE_IDENTIFIER", "NAME", "ADDRESS", "CITY", "COUNTRY", "Description"])
            worksheet.append(["source-1", "Venue One", "1 Main St", "Austin", "US", ""])
            workbook.save(input_xlsx)

            summary = module.enrich(
                build_args(
                    input_xlsx=input_xlsx,
                    output_xlsx=output_xlsx,
                    summary_json=summary_json,
                    cache_dir=tmp_path / "cache",
                    description_batch_size=1,
                )
            )

        self.assertEqual(mocked_post.call_count, 2)
        mocked_get_secret.assert_called_once()
        self.assertEqual(summary["rows_completed"], 1)
        self.assertEqual(summary["openai_call_count"], 1)
        self.assertEqual(summary["estimated_openai_cost"], 0.002)

    @mock.patch.object(module.shared, "get_secret", return_value="openai-secret")
    @mock.patch.object(module.shared.OpenAIClient, "describe_batch")
    def test_enrich_records_batch_failures_and_continues(
        self,
        mocked_describe_batch: mock.Mock,
        mocked_get_secret: mock.Mock,
    ) -> None:
        def fake_describe_batch(*args: object, **_kwargs: object) -> dict[str, str]:
            items = args[-1]
            if len(items) == 2:
                raise RuntimeError("429 too many requests")
            return {str(items[0]["place_id"]): make_valid_description(str(items[0]["name"]))}

        mocked_describe_batch.side_effect = fake_describe_batch

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_json = tmp_path / "summary.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["SOURCE_IDENTIFIER", "NAME", "ADDRESS", "CITY", "COUNTRY", "Description"])
            worksheet.append(["source-1", "Venue One", "1 Main St", "Austin", "US", ""])
            worksheet.append(["source-2", "Venue Two", "2 Main St", "Austin", "US", ""])
            worksheet.append(["source-3", "Venue Three", "3 Main St", "Austin", "US", ""])
            workbook.save(input_xlsx)

            summary = module.enrich(
                build_args(
                    input_xlsx=input_xlsx,
                    output_xlsx=output_xlsx,
                    summary_json=summary_json,
                    cache_dir=tmp_path / "cache",
                    description_batch_size=2,
                )
            )

            result_workbook = load_workbook(output_xlsx)
            result_sheet = result_workbook.active
            description_column = description_column_index(result_sheet)

        mocked_get_secret.assert_called_once()
        self.assertEqual(summary["rows_completed"], 1)
        self.assertEqual(summary["failed_rows"], 2)
        self.assertEqual(len(summary["row_errors"]), 2)
        self.assertEqual(result_sheet.cell(2, description_column).value, None)
        self.assertEqual(result_sheet.cell(3, description_column).value, None)
        self.assertEqual(result_sheet.cell(4, description_column).value, make_valid_description("Venue Three"))

    @mock.patch.object(module.shared, "get_secret", return_value="openai-secret")
    @mock.patch.object(module.shared.OpenAIClient, "describe_batch")
    def test_enrich_persists_completed_batches_before_interrupt(
        self,
        mocked_describe_batch: mock.Mock,
        mocked_get_secret: mock.Mock,
    ) -> None:
        def fake_describe_batch(*args: object, **_kwargs: object) -> dict[str, str]:
            items = args[-1]
            place_id = str(items[0]["place_id"])
            if place_id == "source-2":
                raise KeyboardInterrupt()
            return {place_id: make_valid_description(str(items[0]["name"]))}

        mocked_describe_batch.side_effect = fake_describe_batch

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_xlsx = tmp_path / "input.xlsx"
            output_xlsx = tmp_path / "output.xlsx"
            summary_json = tmp_path / "summary.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["SOURCE_IDENTIFIER", "NAME", "ADDRESS", "CITY", "COUNTRY", "Description"])
            worksheet.append(["source-1", "Venue One", "1 Main St", "Austin", "US", ""])
            worksheet.append(["source-2", "Venue Two", "2 Main St", "Austin", "US", ""])
            workbook.save(input_xlsx)

            with self.assertRaises(KeyboardInterrupt):
                module.enrich(
                    build_args(
                        input_xlsx=input_xlsx,
                        output_xlsx=output_xlsx,
                        summary_json=summary_json,
                        cache_dir=tmp_path / "cache",
                        description_batch_size=1,
                    )
                )

            result_workbook = load_workbook(output_xlsx)
            result_sheet = result_workbook.active
            description_column = description_column_index(result_sheet)
            checkpoint = module.shared.load_checkpoint(module.shared.default_checkpoint_path(output_xlsx))
            summary_exists = summary_json.exists()

        mocked_get_secret.assert_called_once()
        self.assertEqual(result_sheet.cell(2, description_column).value, make_valid_description("Venue One"))
        self.assertEqual(result_sheet.cell(3, description_column).value, None)
        self.assertEqual(checkpoint, {2})
        self.assertFalse(summary_exists)


if __name__ == "__main__":
    unittest.main()
