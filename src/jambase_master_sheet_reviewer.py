#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import subprocess
import threading
import time
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


BASE_URL = "https://www.jambase.com/jb-api/v1"
USER_AGENT = "MusicVenueMasterReviewer/1.0 (+local-export)"
VAULT = "Employee"
MAX_RETRIES = 5
DEFAULT_TIMEOUT = 45
DEFAULT_WORKERS = 4

EXPORT_MATCH_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
QUESTIONABLE_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
UNKNOWN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
BOLD_FONT = Font(bold=True)

REVIEW_COLUMNS = [
    "CODEX_DISCOVERED_IN_EXPORT",
    "CODEX_EXPORT_QUALIFYING_EVENT_COUNT",
    "CODEX_EXPORT_SOURCE_CITIES",
    "CODEX_MUSIC_SIGNAL_STATUS",
    "CODEX_QUESTIONABLE",
    "CODEX_MUSIC_SIGNAL_REASON",
    "CODEX_API_CONCERT_EVENTS_FOUND",
    "CODEX_API_QUALIFYING_EVENTS_FOUND",
    "CODEX_API_QUALIFYING_EVENT_IDS",
    "CODEX_REVIEWED_AT_UTC",
]


def get_secret(name: str) -> str:
    return subprocess.check_output(
        ["op", "read", f"op://{VAULT}/{name}/password"], text=True
    ).strip()


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_int(value: Any) -> int:
    text = clean_string(value)
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def serialize_param(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


class JamBaseClient:
    def __init__(self, api_key: str, cache_dir: Path, timeout: int = DEFAULT_TIMEOUT) -> None:
        self.api_key = api_key
        self.cache_dir = cache_dir
        self.timeout = timeout
        self.local = threading.local()

    def _session(self) -> requests.Session:
        session = getattr(self.local, "session", None)
        if session is None:
            session = requests.Session()
            session.headers.update({"User-Agent": USER_AGENT})
            self.local.session = session
        return session

    def _cache_path(self, endpoint: str, params: dict[str, Any]) -> Path:
        normalized = {key: serialize_param(value) for key, value in sorted(params.items())}
        digest = hashlib.sha1(
            json.dumps({"endpoint": endpoint, "params": normalized}, sort_keys=True).encode(
                "utf-8"
            )
        ).hexdigest()[:16]
        endpoint_slug = endpoint.strip("/").replace("/", "_") or "root"
        return self.cache_dir / endpoint_slug / f"{digest}.json"

    def get(self, endpoint: str, params: dict[str, Any]) -> dict[str, Any]:
        cache_path = self._cache_path(endpoint, params)
        if cache_path.exists():
            return json.loads(cache_path.read_text(encoding="utf-8"))

        request_params = {
            key: serialize_param(value) for key, value in params.items() if value is not None
        }
        request_params["apikey"] = self.api_key

        last_error: Exception | None = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                response = self._session().get(
                    f"{BASE_URL}{endpoint}",
                    params=request_params,
                    timeout=self.timeout,
                )
                if response.status_code == 429:
                    response.raise_for_status()
                if response.status_code in {500, 502, 503, 504}:
                    response.raise_for_status()
                response.raise_for_status()
                payload = response.json()
                ensure_parent(cache_path)
                cache_path.write_text(
                    json.dumps(payload, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                return payload
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                if isinstance(exc, requests.HTTPError) and exc.response is not None:
                    if exc.response.status_code == 429:
                        break
                if attempt == MAX_RETRIES:
                    break
                time.sleep(min(2 ** (attempt - 1), 8))

        detail = str(last_error) if last_error else "unknown_error"
        if isinstance(last_error, requests.HTTPError) and last_error.response is not None:
            response = last_error.response
            body = response.text[:400].replace("\n", " ").strip()
            detail = f"HTTP {response.status_code}: {body}"

        raise RuntimeError(
            f"JamBase request failed for {endpoint} with params {params}: {detail}"
        ) from last_error


def load_export_index(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True)
    worksheet = workbook["audit"]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    positions = {name: index for index, name in enumerate(header)}

    export_index: dict[str, dict[str, Any]] = {}
    for row in rows:
        venue_id = clean_string(row[positions["JamBase Venue ID"]])
        if not venue_id:
            continue
        export_index[venue_id] = {
            "qualifying_event_count": row[positions["Qualifying Event Count"]],
            "source_cities": clean_string(row[positions["Source Cities"]]),
        }
    return export_index


def event_has_music_signal(event: dict[str, Any]) -> tuple[bool, str]:
    offers = event.get("offers") or []
    performers = event.get("performer") or []

    has_ticket = any(
        clean_string(offer.get("url")) and "ticket" in clean_string(offer.get("category")).lower()
        or (clean_string(offer.get("url")) and not clean_string(offer.get("category")))
        for offer in offers
        if isinstance(offer, dict)
    )
    has_performer = any(
        clean_string(performer.get("name"))
        for performer in performers
        if isinstance(performer, dict)
    )

    if has_ticket and has_performer:
        return True, ""
    if not has_ticket and not has_performer:
        return False, "no_ticket_url_and_no_performer"
    if not has_ticket:
        return False, "no_ticket_url"
    return False, "no_performer"


def merge_event_into_bucket(bucket: dict[str, Any], event: dict[str, Any]) -> None:
    bucket["concert_events_found"] += 1
    qualifies, reason = event_has_music_signal(event)
    if qualifies:
        event_id = clean_string(event.get("identifier"))
        if event_id and event_id not in bucket["qualifying_event_id_set"]:
            bucket["qualifying_event_id_set"].add(event_id)
            bucket["qualifying_event_ids"].append(event_id)
            bucket["qualifying_events_found"] += 1
    elif reason:
        bucket["reasons"].append(reason)


def summarize_bucket(bucket: dict[str, Any]) -> dict[str, Any]:
    concert_events_found = int(bucket["concert_events_found"])
    qualifying_events_found = int(bucket["qualifying_events_found"])
    unique_reasons = sorted(set(reason for reason in bucket["reasons"] if reason))

    if concert_events_found < 2:
        status = "questionable"
        question_reason = "fewer_than_2_upcoming_concert_events"
        questionable = True
    elif qualifying_events_found < 2:
        status = "questionable"
        detail = ",".join(unique_reasons) if unique_reasons else "insufficient_ticket_or_artist_signal"
        question_reason = f"fewer_than_2_qualifying_events:{detail}"
        questionable = True
    else:
        status = "strong"
        question_reason = ""
        questionable = False

    return {
        "status": status,
        "questionable": "YES" if questionable else "NO",
        "reason": question_reason,
        "concert_events_found": concert_events_found,
        "qualifying_events_found": qualifying_events_found,
        "qualifying_event_ids": "$".join(bucket["qualifying_event_ids"]),
    }


def fetch_country_event_index(client: JamBaseClient, country_iso2: str) -> tuple[dict[str, dict[str, Any]], int]:
    page = 1
    total_events = 0
    venue_buckets: dict[str, dict[str, Any]] = {}

    while True:
        payload = client.get(
            "/events",
            {
                "geoCountryIso2": country_iso2,
                "eventType": "concert",
                "eventDateFrom": dt.date.today().isoformat(),
                "page": page,
                "perPage": 100,
            },
        )
        events = payload.get("events") or []
        total_events += len(events)

        for event in events:
            venue_identifier = clean_string((event.get("location") or {}).get("identifier"))
            if not venue_identifier:
                continue
            bucket = venue_buckets.setdefault(
                venue_identifier,
                {
                    "concert_events_found": 0,
                    "qualifying_events_found": 0,
                    "qualifying_event_ids": [],
                    "qualifying_event_id_set": set(),
                    "reasons": [],
                },
            )
            merge_event_into_bucket(bucket, event)

        pagination = payload.get("pagination") or {}
        total_pages = pagination.get("totalPages")
        next_page = pagination.get("nextPage")
        if total_pages is not None and page >= total_pages:
            break
        if total_pages is None and not next_page:
            break
        if not events:
            break
        page += 1

    summarized = {
        venue_identifier: summarize_bucket(bucket)
        for venue_identifier, bucket in venue_buckets.items()
    }
    return summarized, total_events


def add_review_columns(worksheet) -> dict[str, int]:
    start_column = worksheet.max_column + 1
    column_map: dict[str, int] = {}
    for offset, name in enumerate(REVIEW_COLUMNS):
        column_index = start_column + offset
        worksheet.cell(row=1, column=column_index, value=name)
        worksheet.cell(row=1, column=column_index).font = BOLD_FONT
        worksheet.cell(row=1, column=column_index).fill = HEADER_FILL
        column_map[name] = column_index
    return column_map


def annotate_row(worksheet, row_number: int, column_map: dict[str, int], review: dict[str, Any]) -> None:
    for key, value in review.items():
        if key not in column_map:
            continue
        worksheet.cell(row=row_number, column=column_map[key], value=value)

    status_cell = worksheet.cell(row=row_number, column=column_map["CODEX_MUSIC_SIGNAL_STATUS"])
    questionable_cell = worksheet.cell(row=row_number, column=column_map["CODEX_QUESTIONABLE"])
    export_cell = worksheet.cell(row=row_number, column=column_map["CODEX_DISCOVERED_IN_EXPORT"])

    if clean_string(review.get("CODEX_DISCOVERED_IN_EXPORT")) == "YES":
        export_cell.fill = EXPORT_MATCH_FILL
    if clean_string(review.get("CODEX_QUESTIONABLE")) == "YES":
        status_cell.fill = QUESTIONABLE_FILL
        questionable_cell.fill = QUESTIONABLE_FILL
    elif clean_string(review.get("CODEX_MUSIC_SIGNAL_STATUS")) == "strong":
        status_cell.fill = EXPORT_MATCH_FILL
    elif clean_string(review.get("CODEX_MUSIC_SIGNAL_STATUS")) == "unknown":
        status_cell.fill = UNKNOWN_FILL
        questionable_cell.fill = UNKNOWN_FILL


def autosize_review_columns(worksheet, column_map: dict[str, int]) -> None:
    for name, column_index in column_map.items():
        width = min(max(len(name), 14), 60)
        worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width = width


def build_summary_sheet(workbook, summary: dict[str, Any]) -> None:
    if "Codex Review Summary" in workbook.sheetnames:
        del workbook["Codex Review Summary"]
    worksheet = workbook.create_sheet("Codex Review Summary")
    worksheet["A1"] = "Metric"
    worksheet["B1"] = "Value"
    worksheet["A1"].font = BOLD_FONT
    worksheet["B1"].font = BOLD_FONT
    worksheet["A1"].fill = HEADER_FILL
    worksheet["B1"].fill = HEADER_FILL

    row = 2
    for key, value in summary.items():
        worksheet.cell(row=row, column=1, value=key)
        worksheet.cell(row=row, column=2, value=value)
        row += 1

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 24


def review_workbook(args: argparse.Namespace) -> dict[str, Any]:
    export_index = load_export_index(args.export_xlsx)
    reviewed_at = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()

    read_workbook = load_workbook(args.master_xlsx, read_only=True)
    read_sheet = read_workbook[read_workbook.sheetnames[0]]
    rows = read_sheet.iter_rows(values_only=True)
    header = next(rows)
    positions = {name: index for index, name in enumerate(header)}

    row_payloads: list[dict[str, Any]] = []
    countries_to_query: set[str] = set()
    for excel_row_number, row in enumerate(rows, start=2):
        venue_identifier = clean_string(row[positions["SOURCE_IDENTIFIER"]])
        venue_name = clean_string(row[positions["NAME"]])
        city = clean_string(row[positions["CITY"]])
        country = clean_string(row[positions["COUNTRY"]])
        country_iso2 = clean_string(row[positions["COUNTRY_ISO2"]])
        num_upcoming_events = parse_int(row[positions["NUM_UPCOMING_EVENTS"]])
        export_match = export_index.get(venue_identifier)

        payload = {
            "row_number": excel_row_number,
            "venue_identifier": venue_identifier,
            "venue_name": venue_name,
            "city": city,
            "country": country,
            "country_iso2": country_iso2,
            "num_upcoming_events": num_upcoming_events,
            "export_match": export_match,
        }
        row_payloads.append(payload)

        if not export_match and num_upcoming_events >= 2 and country_iso2:
            countries_to_query.add(country_iso2)

        if args.max_rows is not None and len(row_payloads) >= args.max_rows:
            break

    api_results: dict[str, dict[str, Any]] = {}
    country_errors: dict[str, str] = {}
    total_events_seen = 0

    sorted_countries = sorted(countries_to_query)
    if sorted_countries:
        api_key = get_secret("JAMBASE_API_KEY")
        client = JamBaseClient(api_key=api_key, cache_dir=args.cache_dir)
        for index, country_iso2 in enumerate(sorted_countries, start=1):
            try:
                country_results, country_event_count = fetch_country_event_index(client, country_iso2)
                api_results.update(country_results)
                total_events_seen += country_event_count
                print(
                    f"Reviewed country {index}/{len(sorted_countries)}: {country_iso2} -> {country_event_count} events, {len(country_results)} venues",
                    flush=True,
                )
            except Exception as exc:  # noqa: BLE001
                country_errors[country_iso2] = str(exc)
                print(
                    f"Country error {index}/{len(sorted_countries)}: {country_iso2} -> {exc}",
                    flush=True,
                )

    workbook = load_workbook(args.master_xlsx)
    worksheet = workbook[workbook.sheetnames[0]]
    column_map = add_review_columns(worksheet)
    autosize_review_columns(worksheet, column_map)

    summary = {
        "generated_at_utc": reviewed_at,
        "master_rows_reviewed": len(row_payloads),
        "export_matches": 0,
        "strong_music_signal": 0,
        "questionable_music_signal": 0,
        "unknown_api_errors": len(country_errors),
        "api_countries_queried": len(sorted_countries),
        "api_events_seen": total_events_seen,
        "api_venues_indexed": len(api_results),
        "cache_dir": str(args.cache_dir),
    }

    for payload in row_payloads:
        venue_identifier = payload["venue_identifier"]
        export_match = payload["export_match"]
        country_iso2 = payload["country_iso2"]
        num_upcoming_events = payload["num_upcoming_events"]
        if export_match:
            review = {
                "CODEX_DISCOVERED_IN_EXPORT": "YES",
                "CODEX_EXPORT_QUALIFYING_EVENT_COUNT": export_match["qualifying_event_count"],
                "CODEX_EXPORT_SOURCE_CITIES": export_match["source_cities"],
                "CODEX_MUSIC_SIGNAL_STATUS": "strong",
                "CODEX_QUESTIONABLE": "NO",
                "CODEX_MUSIC_SIGNAL_REASON": "",
                "CODEX_API_CONCERT_EVENTS_FOUND": export_match["qualifying_event_count"],
                "CODEX_API_QUALIFYING_EVENTS_FOUND": export_match["qualifying_event_count"],
                "CODEX_API_QUALIFYING_EVENT_IDS": "",
                "CODEX_REVIEWED_AT_UTC": reviewed_at,
            }
            summary["export_matches"] += 1
            summary["strong_music_signal"] += 1
        elif num_upcoming_events < 2:
            review = {
                "CODEX_DISCOVERED_IN_EXPORT": "NO",
                "CODEX_EXPORT_QUALIFYING_EVENT_COUNT": "",
                "CODEX_EXPORT_SOURCE_CITIES": "",
                "CODEX_MUSIC_SIGNAL_STATUS": "questionable",
                "CODEX_QUESTIONABLE": "YES",
                "CODEX_MUSIC_SIGNAL_REASON": f"master_sheet_num_upcoming_events_lt_2:{num_upcoming_events}",
                "CODEX_API_CONCERT_EVENTS_FOUND": num_upcoming_events,
                "CODEX_API_QUALIFYING_EVENTS_FOUND": 0,
                "CODEX_API_QUALIFYING_EVENT_IDS": "",
                "CODEX_REVIEWED_AT_UTC": reviewed_at,
            }
            summary["questionable_music_signal"] += 1
        elif venue_identifier in api_results:
            result = api_results[venue_identifier]
            review = {
                "CODEX_DISCOVERED_IN_EXPORT": "NO",
                "CODEX_EXPORT_QUALIFYING_EVENT_COUNT": "",
                "CODEX_EXPORT_SOURCE_CITIES": "",
                "CODEX_MUSIC_SIGNAL_STATUS": result["status"],
                "CODEX_QUESTIONABLE": result["questionable"],
                "CODEX_MUSIC_SIGNAL_REASON": result["reason"],
                "CODEX_API_CONCERT_EVENTS_FOUND": result["concert_events_found"],
                "CODEX_API_QUALIFYING_EVENTS_FOUND": result["qualifying_events_found"],
                "CODEX_API_QUALIFYING_EVENT_IDS": result["qualifying_event_ids"],
                "CODEX_REVIEWED_AT_UTC": reviewed_at,
            }
            if result["questionable"] == "YES":
                summary["questionable_music_signal"] += 1
            else:
                summary["strong_music_signal"] += 1
        elif country_iso2 and country_iso2 not in country_errors:
            review = {
                "CODEX_DISCOVERED_IN_EXPORT": "NO",
                "CODEX_EXPORT_QUALIFYING_EVENT_COUNT": "",
                "CODEX_EXPORT_SOURCE_CITIES": "",
                "CODEX_MUSIC_SIGNAL_STATUS": "questionable",
                "CODEX_QUESTIONABLE": "YES",
                "CODEX_MUSIC_SIGNAL_REASON": "no_matching_upcoming_concert_events_found_in_country_scan",
                "CODEX_API_CONCERT_EVENTS_FOUND": 0,
                "CODEX_API_QUALIFYING_EVENTS_FOUND": 0,
                "CODEX_API_QUALIFYING_EVENT_IDS": "",
                "CODEX_REVIEWED_AT_UTC": reviewed_at,
            }
            summary["questionable_music_signal"] += 1
        else:
            review = {
                "CODEX_DISCOVERED_IN_EXPORT": "NO",
                "CODEX_EXPORT_QUALIFYING_EVENT_COUNT": "",
                "CODEX_EXPORT_SOURCE_CITIES": "",
                "CODEX_MUSIC_SIGNAL_STATUS": "unknown",
                "CODEX_QUESTIONABLE": "",
                "CODEX_MUSIC_SIGNAL_REASON": country_errors.get(country_iso2, "missing_api_result"),
                "CODEX_API_CONCERT_EVENTS_FOUND": "",
                "CODEX_API_QUALIFYING_EVENTS_FOUND": "",
                "CODEX_API_QUALIFYING_EVENT_IDS": "",
                "CODEX_REVIEWED_AT_UTC": reviewed_at,
            }

        annotate_row(worksheet, payload["row_number"], column_map, review)

    build_summary_sheet(workbook, summary)
    ensure_parent(args.output_xlsx)
    workbook.save(args.output_xlsx)

    ensure_parent(args.summary_json)
    args.summary_json.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return summary


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Review JamBase master sheet for music-signal quality.")
    parser.add_argument(
        "--master-xlsx",
        type=Path,
        default=Path("/Users/saugat/Downloads/Jambase_Venues_Global run 3.xlsx"),
        help="Master workbook path.",
    )
    parser.add_argument(
        "--export-xlsx",
        type=Path,
        default=Path("data/output/jambase_music_venues.xlsx"),
        help="Previously exported filtered JamBase workbook.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("data/output/Jambase_Venues_Global_run_3_reviewed.xlsx"),
        help="Annotated workbook output path.",
    )
    parser.add_argument(
        "--summary-json",
        type=Path,
        default=Path("data/output/Jambase_Venues_Global_run_3_review_summary.json"),
        help="Summary JSON path.",
    )
    parser.add_argument(
        "--cache-dir",
        type=Path,
        default=Path("data/cache/jambase_master_review"),
        help="JamBase API cache directory.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=DEFAULT_WORKERS,
        help="Unused compatibility flag; country aggregation now runs sequentially.",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="Review only the first N data rows for smoke testing.",
    )
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    try:
        summary = review_workbook(args)
    except KeyboardInterrupt:
        print("Review interrupted.")
        return 130
    except Exception as exc:  # noqa: BLE001
        print(f"Review failed: {exc}")
        return 1

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
