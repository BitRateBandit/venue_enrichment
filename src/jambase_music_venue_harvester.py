#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import re
import subprocess
import sys
import time
import unicodedata
from collections import OrderedDict
from pathlib import Path
from typing import Any

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.jambase.com/jb-api/v1"
DEFAULT_INPUT = Path("data/eu_countries_&_cities.yaml")
DEFAULT_OUTPUT = Path("data/output/jambase_music_venues.xlsx")
DEFAULT_SUMMARY = Path("data/output/jambase_music_venues_summary.json")
DEFAULT_UNRESOLVED = Path("data/output/jambase_unresolved_cities.csv")
DEFAULT_CACHE_DIR = Path("data/cache/jambase")
USER_AGENT = "MusicVenueHarvester/1.0 (+local-export)"
VAULT = "Employee"
PAGE_SIZE = 100
MAX_RETRIES = 5
DEFAULT_REQUEST_PAUSE = 0.15
DEFAULT_TIMEOUT = 45

MANUAL_COUNTRY_ALIASES = {
    "Czech Republic": ["Czechia"],
    "North Macedonia": ["Macedonia", "Republic of North Macedonia"],
    "Russia": ["Russian Federation"],
    "Turkey": ["Turkiye", "Türkiye", "Republic of Turkiye"],
    "Vatican City": ["Holy See", "Vatican City State"],
}

MANUAL_CITY_ALIASES = {
    "Ibiza Town": ["Ibiza"],
    "Luxembourg City": ["Luxembourg"],
    "Newcastle upon Tyne": ["Newcastle"],
    "Quebec City": ["Quebec", "Québec"],
    "St. Julian's": ["Saint Julian's", "St Julians"],
}

MAIN_SHEET_COLUMNS = [
    "Venue Name",
    "Address",
    "City",
    "State",
    "Zip Code",
    "Country",
    "Latitude",
    "Longitude",
    "Website",
    "Facebook",
    "Instagram",
    "X (Twitter)",
    "YouTube",
    "Spotify URL",
    "Other Links",
    "Image URLs",
    "Description",
]

AUDIT_SHEET_COLUMNS = [
    "JamBase Venue ID",
    "JamBase URL",
    "JamBase Upcoming Events",
    "Qualifying Event Count",
    "Qualifying Event IDs",
    "Ticket URLs",
    "Artists",
    "Source Countries",
    "Source Cities",
    "Resolved Country ISO2",
    "Resolved Country ISO3",
    "Resolved City IDs",
]


def get_secret(name: str) -> str:
    return subprocess.check_output(
        ["op", "read", f"op://{VAULT}/{name}/password"], text=True
    ).strip()


def normalize_text(value: str) -> str:
    if not value:
        return ""
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.replace("&", " and ")
    value = re.sub(r"[’'`]", "", value)
    value = re.sub(r"[^0-9A-Za-z]+", " ", value)
    return re.sub(r"\s+", " ", value).strip().lower()


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def slugify(value: str) -> str:
    normalized = normalize_text(value)
    return re.sub(r"[^0-9a-z]+", "-", normalized).strip("-") or "item"


def serialize_param(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def join_unique(values: list[str] | set[str], sep: str = "$") -> str:
    cleaned = sorted(
        {clean_string(value) for value in values if clean_string(value)}
    )
    return sep.join(cleaned)


def flatten_region(value: Any) -> str:
    if isinstance(value, dict):
        return clean_string(
            value.get("name")
            or value.get("identifier")
            or value.get("alternateName")
            or ""
        )
    return clean_string(value)


def flatten_country(value: Any) -> tuple[str, str, str]:
    if isinstance(value, dict):
        return (
            clean_string(value.get("name") or ""),
            clean_string(value.get("identifier") or ""),
            clean_string(value.get("alternateName") or ""),
        )
    return (clean_string(value), "", "")


def build_combined_address(address: dict[str, Any]) -> str:
    parts = [
        clean_string(address.get("streetAddress")),
        clean_string(address.get("x-streetAddress2")),
        clean_string(address.get("addressLocality")),
        flatten_region(address.get("addressRegion")),
        clean_string(address.get("postalCode")),
    ]
    country_name, country_iso2, country_iso3 = flatten_country(
        address.get("addressCountry")
    )
    parts.append(country_name or country_iso2 or country_iso3)
    return ", ".join(part for part in parts if clean_string(part))


def country_rank(country: dict[str, Any]) -> tuple[int, int]:
    return (
        int(country.get("x-numUpcomingEvents") or 0),
        1 if clean_string(country.get("identifier")) else 0,
    )


def city_query_candidates(input_city_name: str) -> list[str]:
    queries = [input_city_name]
    queries.extend(MANUAL_CITY_ALIASES.get(input_city_name, []))

    if input_city_name.endswith(" City"):
        queries.append(input_city_name[: -len(" City")])
    if "St." in input_city_name:
        queries.append(input_city_name.replace("St.", "Saint"))
        queries.append(input_city_name.replace("St.", "St"))
    if " upon " in input_city_name:
        queries.append(input_city_name.split(" upon ", 1)[0])

    unique_queries = []
    seen = set()
    for query in queries:
        cleaned = clean_string(query)
        key = normalize_text(cleaned)
        if cleaned and key not in seen:
            unique_queries.append(cleaned)
            seen.add(key)
    return unique_queries


def sheet_row_width(value: Any) -> int:
    if value is None:
        return 0
    return len(str(value))


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


class JamBaseClient:
    def __init__(
        self,
        api_key: str,
        cache_dir: Path,
        refresh_cache: bool = False,
        request_pause: float = DEFAULT_REQUEST_PAUSE,
        timeout: int = DEFAULT_TIMEOUT,
    ) -> None:
        self.api_key = api_key
        self.cache_dir = cache_dir
        self.refresh_cache = refresh_cache
        self.request_pause = request_pause
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})

    def _cache_path(self, endpoint: str, params: dict[str, Any]) -> Path:
        normalized_params = {
            key: serialize_param(value) for key, value in sorted(params.items())
        }
        digest_input = json.dumps(
            {"endpoint": endpoint, "params": normalized_params},
            sort_keys=True,
            ensure_ascii=True,
        ).encode("utf-8")
        digest = hashlib.sha1(digest_input).hexdigest()[:16]

        endpoint_slug = endpoint.strip("/").replace("/", "_") or "root"
        readable_bits = [
            normalized_params.get("geoCountryIso2", ""),
            normalized_params.get("geoCityId", ""),
            normalized_params.get("geoCityName", ""),
            normalized_params.get("eventType", ""),
            normalized_params.get("page", ""),
        ]
        readable = "-".join(slugify(bit) for bit in readable_bits if bit)
        filename = f"{digest}-{readable}.json" if readable else f"{digest}.json"
        return self.cache_dir / endpoint_slug / filename

    def get(self, endpoint: str, params: dict[str, Any]) -> dict[str, Any]:
        cache_path = self._cache_path(endpoint, params)
        if cache_path.exists() and not self.refresh_cache:
            return json.loads(cache_path.read_text(encoding="utf-8"))

        request_params = {
            key: serialize_param(value) for key, value in params.items() if value is not None
        }
        request_params["apikey"] = self.api_key

        last_error: Exception | None = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                response = self.session.get(
                    f"{BASE_URL}{endpoint}",
                    params=request_params,
                    timeout=self.timeout,
                )
                if response.status_code in {429, 500, 502, 503, 504}:
                    response.raise_for_status()
                response.raise_for_status()
                payload = response.json()
                ensure_parent(cache_path)
                cache_path.write_text(
                    json.dumps(payload, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                if self.request_pause:
                    time.sleep(self.request_pause)
                return payload
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                if attempt == MAX_RETRIES:
                    break
                time.sleep(min(2 ** (attempt - 1), 8))

        raise RuntimeError(
            f"JamBase request failed for {endpoint} with params {params}"
        ) from last_error


def paginate(
    client: JamBaseClient,
    endpoint: str,
    params: dict[str, Any],
    items_key: str,
) -> list[dict[str, Any]]:
    page = 1
    items: list[dict[str, Any]] = []
    while True:
        payload = client.get(endpoint, {**params, "page": page, "perPage": PAGE_SIZE})
        batch = payload.get(items_key) or []
        items.extend(batch)

        pagination = payload.get("pagination") or {}
        total_pages = pagination.get("totalPages")
        next_page = pagination.get("nextPage")

        if total_pages is not None and page >= total_pages:
            break
        if total_pages is None and not next_page:
            break
        if not batch:
            break
        page += 1
    return items


def load_city_plan(path: Path) -> list[tuple[str, str]]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8"))
    city_plan: list[tuple[str, str]] = []
    for country_name, section in payload.items():
        for key in ("major_cities", "expanded_cities"):
            for city_name in section.get(key, []) or []:
                city_plan.append((country_name, city_name))
    return city_plan


def resolve_country(
    input_country_name: str,
    countries: list[dict[str, Any]],
) -> dict[str, Any] | None:
    target = normalize_text(input_country_name)

    alias_pool = {
        normalize_text(alias)
        for alias in MANUAL_COUNTRY_ALIASES.get(input_country_name, [])
    }

    exact_matches: list[dict[str, Any]] = []
    for country in countries:
        candidate_keys = {
            normalize_text(country.get("name", "")),
            normalize_text(country.get("identifier", "")),
            normalize_text(country.get("alternateName", "")),
        }
        if target in candidate_keys or bool(alias_pool.intersection(candidate_keys)):
            exact_matches.append(country)

    if exact_matches:
        return max(exact_matches, key=country_rank)

    fuzzy_matches: list[dict[str, Any]] = []
    for country in countries:
        candidate_keys = {
            normalize_text(country.get("name", "")),
            normalize_text(country.get("identifier", "")),
            normalize_text(country.get("alternateName", "")),
        }
        if target and any(target in key or key in target for key in candidate_keys if key):
            fuzzy_matches.append(country)

    if fuzzy_matches:
        return max(fuzzy_matches, key=country_rank)

    return None


def resolve_city(
    client: JamBaseClient,
    country_iso2: str,
    input_city_name: str,
) -> dict[str, Any] | None:
    target = normalize_text(input_city_name)
    for city_query in city_query_candidates(input_city_name):
        payload = client.get(
            "/geographies/cities",
            {
                "geoCityName": city_query,
                "geoCountryIso2": country_iso2,
                "cityHasUpcomingEvents": True,
                "page": 1,
                "perPage": 25,
            },
        )
        results = payload.get("cities") or []
        if not results:
            continue

        exact = [city for city in results if normalize_text(city.get("name", "")) == target]
        if exact:
            return exact[0]

        close = [
            city
            for city in results
            if target in normalize_text(city.get("name", ""))
            or normalize_text(city.get("name", "")) in target
        ]
        if close:
            return close[0]

        if len(results) == 1:
            return results[0]

    return None


def extract_same_as_links(entries: Any) -> tuple[str, dict[str, set[str]]]:
    website = ""
    buckets: dict[str, set[str]] = {}

    for entry in entries or []:
        if isinstance(entry, str):
            continue
        if not isinstance(entry, dict):
            continue
        url = clean_string(entry.get("url"))
        if not url:
            continue
        link_type = clean_string(entry.get("identifier"))
        if link_type == "officialSite":
            website = website or url
            continue
        buckets.setdefault(link_type or "other", set()).add(url)

    return website, buckets


def event_has_music_signal(event: dict[str, Any]) -> tuple[bool, list[str], list[str]]:
    offers = event.get("offers") or []
    performers = event.get("performer") or []

    ticket_urls = []
    for offer in offers:
        if not isinstance(offer, dict):
            continue
        url = clean_string(offer.get("url"))
        category = clean_string(offer.get("category")).lower()
        if url and ("ticket" in category or not category):
            ticket_urls.append(url)

    artist_names = []
    for performer in performers:
        if not isinstance(performer, dict):
            continue
        name = clean_string(performer.get("name"))
        if name:
            artist_names.append(name)

    return (bool(ticket_urls) and bool(artist_names), ticket_urls, artist_names)


def collect_venue_records(
    events: list[dict[str, Any]],
    source_country_name: str,
    source_city_name: str,
    resolved_country_iso2: str,
    resolved_country_iso3: str,
    resolved_city_id: str,
    registry: OrderedDict[str, dict[str, Any]],
) -> None:
    for event in events:
        venue = event.get("location") or {}
        venue_id = venue.get("identifier")
        if not venue_id:
            continue

        if venue.get("x-isPermanentlyClosed") is True:
            continue

        address = venue.get("address") or {}
        country_name, country_iso2, country_iso3 = flatten_country(
            address.get("addressCountry")
        )

        website, social_links = extract_same_as_links(venue.get("sameAs"))
        image_urls = set()
        if venue.get("image"):
            image_urls.add(venue["image"])

        record = registry.setdefault(
            venue_id,
            {
                "venue_name": clean_string(venue.get("name")),
                "address": build_combined_address(address),
                "city": clean_string(address.get("addressLocality")),
                "state": flatten_region(address.get("addressRegion")),
                "zip_code": clean_string(address.get("postalCode")),
                "country": country_name or country_iso2 or country_iso3,
                "country_iso2": country_iso2 or resolved_country_iso2,
                "country_iso3": country_iso3 or resolved_country_iso3,
                "latitude": (venue.get("geo") or {}).get("latitude") or "",
                "longitude": (venue.get("geo") or {}).get("longitude") or "",
                "website": website,
                "social_links": {},
                "image_urls": image_urls,
                "description": "",
                "jambase_venue_id": venue_id,
                "jambase_url": clean_string(venue.get("url")),
                "jambase_upcoming_events": venue.get("x-numUpcomingEvents") or 0,
                "qualifying_event_count": 0,
                "qualifying_event_ids": set(),
                "ticket_urls": set(),
                "artists": set(),
                "source_countries": {source_country_name},
                "source_cities": {source_city_name},
                "resolved_city_ids": {resolved_city_id},
            },
        )

        if website and not record["website"]:
            record["website"] = website
        if venue.get("url") and not record["jambase_url"]:
            record["jambase_url"] = clean_string(venue["url"])
        if venue.get("x-numUpcomingEvents"):
            record["jambase_upcoming_events"] = max(
                int(record["jambase_upcoming_events"] or 0),
                int(venue.get("x-numUpcomingEvents") or 0),
            )
        record["source_countries"].add(source_country_name)
        record["source_cities"].add(source_city_name)
        record["resolved_city_ids"].add(resolved_city_id)
        if not record["venue_name"] and venue.get("name"):
            record["venue_name"] = clean_string(venue["name"])
        if not record["address"]:
            record["address"] = build_combined_address(address)
        if not record["city"]:
            record["city"] = clean_string(address.get("addressLocality"))
        if not record["state"]:
            record["state"] = flatten_region(address.get("addressRegion"))
        if not record["zip_code"]:
            record["zip_code"] = clean_string(address.get("postalCode"))
        if not record["country"]:
            record["country"] = country_name or country_iso2 or country_iso3
        if not record["latitude"]:
            record["latitude"] = (venue.get("geo") or {}).get("latitude") or ""
        if not record["longitude"]:
            record["longitude"] = (venue.get("geo") or {}).get("longitude") or ""
        record["image_urls"].update(image_urls)

        for link_type, urls in social_links.items():
            record["social_links"].setdefault(link_type, set()).update(urls)

        qualifies, ticket_urls, artists = event_has_music_signal(event)
        if qualifies:
            event_id = event.get("identifier") or ""
            if event_id and event_id not in record["qualifying_event_ids"]:
                record["qualifying_event_ids"].add(event_id)
                record["qualifying_event_count"] += 1
            record["ticket_urls"].update(ticket_urls)
            record["artists"].update(artists)


def build_main_row(record: dict[str, Any]) -> dict[str, Any]:
    social_links = record["social_links"]
    known_social_keys = {"facebook", "instagram", "twitter", "youtube", "spotify"}
    other_links = set()
    for key, urls in social_links.items():
        if key not in known_social_keys:
            other_links.update(urls)

    return {
        "Venue Name": record["venue_name"],
        "Address": record["address"],
        "City": record["city"],
        "State": record["state"],
        "Zip Code": record["zip_code"],
        "Country": record["country"],
        "Latitude": record["latitude"],
        "Longitude": record["longitude"],
        "Website": record["website"],
        "Facebook": join_unique(social_links.get("facebook", set())),
        "Instagram": join_unique(social_links.get("instagram", set())),
        "X (Twitter)": join_unique(social_links.get("twitter", set())),
        "YouTube": join_unique(social_links.get("youtube", set())),
        "Spotify URL": join_unique(social_links.get("spotify", set())),
        "Other Links": join_unique(other_links),
        "Image URLs": join_unique(record["image_urls"]),
        "Description": record["description"],
    }


def build_audit_row(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "JamBase Venue ID": record["jambase_venue_id"],
        "JamBase URL": record["jambase_url"],
        "JamBase Upcoming Events": record["jambase_upcoming_events"],
        "Qualifying Event Count": record["qualifying_event_count"],
        "Qualifying Event IDs": join_unique(record["qualifying_event_ids"]),
        "Ticket URLs": join_unique(record["ticket_urls"]),
        "Artists": join_unique(record["artists"]),
        "Source Countries": join_unique(record["source_countries"]),
        "Source Cities": join_unique(record["source_cities"]),
        "Resolved Country ISO2": record["country_iso2"],
        "Resolved Country ISO3": record["country_iso3"],
        "Resolved City IDs": join_unique(record["resolved_city_ids"]),
    }


def write_worksheet(
    workbook: Workbook,
    title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(columns)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    widths = {index + 1: len(column) for index, column in enumerate(columns)}
    for row in rows:
        values = [row.get(column, "") for column in columns]
        worksheet.append(values)
        for index, value in enumerate(values, start=1):
            widths[index] = min(max(widths[index], sheet_row_width(value)), 80)

    for index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(index)].width = width + 2


def write_unresolved_report(path: Path, rows: list[dict[str, Any]]) -> None:
    ensure_parent(path)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "country",
                "country_iso2",
                "city",
                "reason",
                "candidate_names",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def harvest(args: argparse.Namespace) -> dict[str, Any]:
    api_key = get_secret("JAMBASE_API_KEY")
    client = JamBaseClient(
        api_key=api_key,
        cache_dir=args.cache_dir,
        refresh_cache=args.refresh_cache,
        request_pause=args.request_pause,
    )

    all_countries = paginate(
        client,
        "/geographies/countries",
        {},
        "countries",
    )
    city_plan = load_city_plan(args.input_yaml)

    if args.max_cities is not None:
        city_plan = city_plan[: args.max_cities]

    venues: OrderedDict[str, dict[str, Any]] = OrderedDict()
    unresolved: list[dict[str, Any]] = []

    for index, (input_country_name, input_city_name) in enumerate(city_plan, start=1):
        country = resolve_country(input_country_name, all_countries)
        if not country:
            unresolved.append(
                {
                    "country": input_country_name,
                    "country_iso2": "",
                    "city": input_city_name,
                    "reason": "country_not_resolved",
                    "candidate_names": "",
                }
            )
            print(
                f"[{index}/{len(city_plan)}] {input_country_name} / {input_city_name}: country not resolved",
                flush=True,
            )
            continue

        country_iso2 = country.get("identifier") or ""
        country_iso3 = country.get("alternateName") or ""

        city = resolve_city(client, country_iso2, input_city_name)
        if not city:
            city_candidates = client.get(
                "/geographies/cities",
                {
                    "geoCityName": input_city_name,
                    "geoCountryIso2": country_iso2,
                    "cityHasUpcomingEvents": True,
                    "page": 1,
                    "perPage": 10,
                },
            ).get("cities", [])
            unresolved.append(
                {
                    "country": input_country_name,
                    "country_iso2": country_iso2,
                    "city": input_city_name,
                    "reason": "city_not_resolved_or_no_upcoming_events",
                    "candidate_names": join_unique(
                        [candidate.get("name", "") for candidate in city_candidates],
                        sep=" | ",
                    ),
                }
            )
            print(
                f"[{index}/{len(city_plan)}] {input_country_name} / {input_city_name}: no resolved JamBase city",
                flush=True,
            )
            continue

        resolved_city_id = city.get("identifier") or ""
        resolved_city_name = city.get("name") or input_city_name
        print(
            f"[{index}/{len(city_plan)}] {input_country_name} / {input_city_name} -> {resolved_city_name} ({resolved_city_id})",
            flush=True,
        )

        events = paginate(
            client,
            "/events",
            {
                "geoCityId": resolved_city_id,
                "geoCountryIso2": country_iso2,
                "eventType": "concert",
                "eventDateFrom": dt.date.today().isoformat(),
            },
            "events",
        )
        print(
            f"  fetched {len(events)} concert events for {resolved_city_name}",
            flush=True,
        )

        collect_venue_records(
            events=events,
            source_country_name=input_country_name,
            source_city_name=input_city_name,
            resolved_country_iso2=country_iso2,
            resolved_country_iso3=country_iso3,
            resolved_city_id=resolved_city_id,
            registry=venues,
        )

    filtered_records = [
        record
        for record in venues.values()
        if int(record["jambase_upcoming_events"] or 0) >= 2
        and int(record["qualifying_event_count"] or 0) >= 2
    ]

    filtered_records.sort(
        key=lambda record: (
            normalize_text(record["country"]),
            normalize_text(record["city"]),
            normalize_text(record["venue_name"]),
        )
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_worksheet(
        workbook,
        "venues",
        MAIN_SHEET_COLUMNS,
        [build_main_row(record) for record in filtered_records],
    )
    write_worksheet(
        workbook,
        "audit",
        AUDIT_SHEET_COLUMNS,
        [build_audit_row(record) for record in filtered_records],
    )

    ensure_parent(args.output_xlsx)
    workbook.save(args.output_xlsx)
    write_unresolved_report(args.unresolved_csv, unresolved)

    summary = {
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "input_yaml": str(args.input_yaml),
        "output_xlsx": str(args.output_xlsx),
        "cache_dir": str(args.cache_dir),
        "cities_requested": len(city_plan),
        "cities_unresolved": len(unresolved),
        "venues_seen": len(venues),
        "venues_exported": len(filtered_records),
        "filter_rule": {
            "x_num_upcoming_events_min": 2,
            "qualifying_events_min": 2,
            "qualifying_event_definition": "event has at least one ticket URL and at least one performer",
            "event_type": "concert",
        },
    }
    ensure_parent(args.summary_json)
    args.summary_json.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return summary


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Harvest JamBase-backed music venues into an XLSX export."
    )
    parser.add_argument(
        "--input-yaml",
        type=Path,
        default=DEFAULT_INPUT,
        help="Country/city YAML input file.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Destination spreadsheet path.",
    )
    parser.add_argument(
        "--summary-json",
        type=Path,
        default=DEFAULT_SUMMARY,
        help="Harvest summary JSON path.",
    )
    parser.add_argument(
        "--unresolved-csv",
        type=Path,
        default=DEFAULT_UNRESOLVED,
        help="Unresolved country/city CSV path.",
    )
    parser.add_argument(
        "--cache-dir",
        type=Path,
        default=DEFAULT_CACHE_DIR,
        help="Directory for cached JamBase responses.",
    )
    parser.add_argument(
        "--refresh-cache",
        action="store_true",
        help="Ignore cached JamBase responses and fetch again.",
    )
    parser.add_argument(
        "--max-cities",
        type=int,
        default=None,
        help="Only process the first N country/city rows for a smoke test.",
    )
    parser.add_argument(
        "--request-pause",
        type=float,
        default=DEFAULT_REQUEST_PAUSE,
        help="Pause in seconds after each successful JamBase request.",
    )
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    try:
        summary = harvest(args)
    except KeyboardInterrupt:
        print("Harvest interrupted.", file=sys.stderr)
        return 130
    except Exception as exc:  # noqa: BLE001
        print(f"Harvest failed: {exc}", file=sys.stderr)
        return 1

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
